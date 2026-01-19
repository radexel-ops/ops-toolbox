# -*- coding: utf-8 -*-
"""
세금계산서 - AP List 매칭 엔진 (병렬 처리 적용)

매칭 방식: AI가 전체 AP 후보 중에서 최적 매칭을 찾음 (병렬 처리로 속도 개선)
"""

from pathlib import Path
from typing import Optional, Callable, Dict, List
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import openpyxl

from config import BASE_DIR, COMPANY_NAME
from utils.text_utils import safe_str
from utils.excel_utils import leave_comment, lookup_bank_by_supplier
from services.gemini_service import GeminiService


class MatchingEngine:
    """세금계산서와 AP List 매칭 처리 엔진"""

    # 세금계산서 내용 컬럼명
    TAX_CONTENT_COL = "세금 계산서 내용"

    # 병렬 처리 설정
    MAX_WORKERS = 5

    def __init__(self, ai_service: Optional[GeminiService] = None):
        """
        Args:
            ai_service: Gemini 서비스 인스턴스
        """
        self.ai = ai_service or GeminiService()

    @staticmethod
    def _combine_invoice_content(invoice_dict: Dict) -> str:
        """
        세금계산서의 모든 항목을 &로 결합

        Args:
            invoice_dict: 세금계산서 데이터

        Returns:
            "컬럼명:값 & 컬럼명:값 & ..." 형식 문자열
        """
        parts = []
        for key, value in invoice_dict.items():
            val_str = safe_str(value)
            if val_str:  # 빈 값 제외
                parts.append(f"{key}:{val_str}")
        return " & ".join(parts)

    def run_match(
        self,
        ap_path: str,
        inv_path: str,
        out_path: Optional[str] = None,
        on_progress: Optional[Callable[[int, int], None]] = None,
        on_match: Optional[Callable[[Dict], None]] = None,
        on_new: Optional[Callable[[Dict], None]] = None,
    ) -> Path:
        """
        AP List와 세금계산서 매칭 실행

        Args:
            ap_path: AP List 파일 경로
            inv_path: 세금계산서 파일 경로
            out_path: 출력 파일 경로 (None이면 자동 생성)
            on_progress: 진행률 콜백 (done, total)
            on_match: 매칭 발견 시 콜백
            on_new: 신규 행 생성 시 콜백

        Returns:
            출력 파일 경로
        """
        ap_file = Path(ap_path)
        inv_file = Path(inv_path)
        output_file = Path(out_path) if out_path else BASE_DIR / "AP List_RDXL_merged.xlsx"

        # 원본 로드 (Pandas)
        ap_book = pd.ExcelFile(ap_file)
        ap_df = pd.read_excel(ap_book, sheet_name="2025")
        inv_df = (
            pd.read_excel(inv_file, sheet_name=0, header=5)
            .rename(columns=lambda c: str(c).strip())
        )
        if "품목명" in inv_df.columns:
            inv_df = inv_df.rename(columns={"품목명": "품목"})

        # 계좌정보 시트 로드
        acc_df = (
            pd.read_excel(ap_book, sheet_name="계좌정보")
            .rename(columns=str)
            .fillna("")
        )

        # 엑셀 워크북 (openpyxl)
        wb = openpyxl.load_workbook(ap_file)
        ws = wb["2025"]
        header = [c.value for c in ws[1]]
        ap_cols = {h: i + 1 for i, h in enumerate(header)}

        # "세금 계산서 내용" 컬럼 없으면 추가
        if self.TAX_CONTENT_COL not in ap_cols:
            new_col_idx = len(header) + 1
            ws.cell(1, new_col_idx, self.TAX_CONTENT_COL)
            header.append(self.TAX_CONTENT_COL)
            ap_cols[self.TAX_CONTENT_COL] = new_col_idx

        # 인보이스 컬럼 매핑
        inv_cols = {
            "승인번호": [c for c in inv_df.columns if "승인번호" in c][0],
            "상호": [c for c in inv_df.columns if "상호" in c][0],
            "품목": "품목",
            "합계금액": [c for c in inv_df.columns if "합계" in c and "금액" in c][0],
        }

        total = len(inv_df)

        # AP 후보 리스트 (한 번만 생성)
        ap_candidate_dicts: List[Dict] = ap_df.apply(
            lambda r: {
                "row_index": int(r.name),
                **{col: safe_str(r[col]) for col in ap_df.columns},
                "내부메모": safe_str(r.get("내부 메모", "")),
                "금액": float(r["금액"]) if pd.notna(r["금액"]) else 0.0,
            },
            axis=1,
        ).tolist()

        # 1단계: 인보이스 데이터 준비
        invoice_list = []
        for idx_inv, inv in inv_df.iterrows():
            invoice_dict = {
                **{col: safe_str(inv[col]) for col in inv_df.columns},
                "합계금액": float(inv[inv_cols["합계금액"]]),
                "_idx": idx_inv,
            }
            invoice_list.append(invoice_dict)

        # 2단계: 병렬 매칭 처리 (AI가 전체 후보에서 매칭 찾음)
        def process_single_invoice(invoice_dict: Dict) -> Dict:
            """단일 인보이스 매칭 처리 (병렬용)"""
            # AI에게 전체 AP 후보를 전달하여 매칭
            match_idx = self.ai.find_matching_ap_row(invoice_dict, ap_candidate_dicts)

            if match_idx >= 0:
                return {
                    "type": "match",
                    "invoice": invoice_dict,
                    "match_idx": match_idx,
                    "candidate": ap_candidate_dicts[match_idx],
                }
            else:
                # 신규 행용 AI 데이터 생성
                ai_vals = self.ai.generate_new_ap_row(invoice_dict)
                ai_vals["은행"], ai_vals["계좌번호"] = lookup_bank_by_supplier(
                    invoice_dict["상호"], acc_df
                )
                return {
                    "type": "new",
                    "invoice": invoice_dict,
                    "ai_vals": ai_vals,
                }

        results = []
        completed = 0

        with ThreadPoolExecutor(max_workers=self.MAX_WORKERS) as executor:
            future_to_inv = {
                executor.submit(process_single_invoice, inv): inv
                for inv in invoice_list
            }

            for future in as_completed(future_to_inv):
                try:
                    result = future.result()
                    results.append(result)
                except Exception:
                    pass  # 에러 시 건너뜀

                completed += 1
                if on_progress:
                    on_progress(completed, total)

        # 3단계: 결과 순차 적용 (Excel 수정은 thread-safe하지 않으므로)
        for result in results:
            invoice_dict = result["invoice"]

            if result["type"] == "match":
                self._handle_match(
                    ws, ap_cols, invoice_dict, result["candidate"],
                    result["match_idx"], on_match
                )
            else:
                self._handle_new_row_with_ai_vals(
                    ws, header, invoice_dict, result["ai_vals"], on_new
                )

        # 저장
        wb.save(output_file)
        return output_file

    def _handle_match(
        self,
        ws,
        ap_cols: Dict[str, int],
        invoice_dict: Dict,
        cand: Dict,
        match_idx: int,
        on_match: Optional[Callable]
    ) -> None:
        """매칭된 행 처리"""
        ap_row_num = match_idx + 2  # 헤더 1줄 오프셋

        def swap(col, new_val):
            cell = ws.cell(ap_row_num, ap_cols[col])
            old = safe_str(cell.value)
            if col == "비고":
                if new_val in old:
                    return
                cell.value = f"{old}, {new_val}" if old else new_val
                return
            if old:
                leave_comment(ws, cell.coordinate, f"기존 : {old}")
            cell.value = new_val

        swap("세금계산서 수령", invoice_dict["승인번호"])
        swap("수취인", invoice_dict["상호"])
        swap("비고", invoice_dict["품목"])
        swap("본인 통장표시내용", invoice_dict["상호"])

        # 세금계산서 전체 내용 기록
        if self.TAX_CONTENT_COL in ap_cols:
            tax_content = self._combine_invoice_content(invoice_dict)
            ws.cell(ap_row_num, ap_cols[self.TAX_CONTENT_COL], tax_content)

        # 금액 차이 코멘트
        inv_amt = invoice_dict["합계금액"]
        ap_amt = cand["금액"]
        if abs(inv_amt - ap_amt) >= 1:
            leave_comment(
                ws,
                ws.cell(ap_row_num, ap_cols["금액"]).coordinate,
                f"세금계산서 금액 : {int(inv_amt):,}"
            )

        # UI용 dict
        if on_match:
            row_tv = {
                "inv_승인번호": invoice_dict["승인번호"],
                "ap_세금계산서수령": cand["세금계산서 수령"],
                "inv_상호": invoice_dict["상호"],
                "ap_수취인": cand["수취인"],
                "inv_품목": invoice_dict["품목"],
                "ap_품목": cand["품목"],
                "inv_합계금액": inv_amt,
                "ap_금액": ap_amt,
                "이체목표": cand.get("이체 목표", ""),
                "내부메모": cand.get("내부메모", ""),
                "비고": cand["비고"],
                "금액차이": inv_amt - ap_amt,
                "ap_row_num": ap_row_num,
            }
            on_match(row_tv)

    def _handle_new_row(
        self,
        ws,
        header: List[str],
        invoice_dict: Dict,
        acc_df: pd.DataFrame,
        on_new: Optional[Callable]
    ) -> None:
        """신규 행 처리"""
        ai_vals = self.ai.generate_new_ap_row(invoice_dict)
        # 계좌정보 시트 매핑 우선
        ai_vals["은행"], ai_vals["계좌번호"] = lookup_bank_by_supplier(
            invoice_dict["상호"], acc_df
        )

        # 신규 행 dict
        new_row = {
            "기안 날짜": invoice_dict["승인번호"][:8],
            "세금계산서 수령": invoice_dict["승인번호"],
            "이체 목표": "세금계산서 확인필요",
            "이체 날짜": "",
            "요청자": "정수현",
            "구분": ai_vals["구분"],
            "품목": invoice_dict["품목"],
            "문서번호": "",
            "비고": invoice_dict["품목"],
            "은행": ai_vals["은행"],
            "계좌번호": ai_vals["계좌번호"],
            "금액": invoice_dict["합계금액"],
            "수취인": invoice_dict["상호"],
            "공란": "",
            "CMS코드(공란)": "",
            "본인 통장표시내용": invoice_dict["상호"],
            "받는분 통장 표시": COMPANY_NAME,
            "내부 메모": ai_vals["내부메모"],
            self.TAX_CONTENT_COL: self._combine_invoice_content(invoice_dict),
        }
        # 실제 헤더 순서로 append
        ws.append([new_row.get(col, "") for col in header])

        # UI용 dict
        if on_new:
            row_tv = {
                "inv_승인번호": invoice_dict["승인번호"],
                "ap_세금계산서수령": "",
                "inv_상호": invoice_dict["상호"],
                "ap_수취인": "",
                "inv_품목": invoice_dict["품목"],
                "ap_품목": "",
                "inv_합계금액": invoice_dict["합계금액"],
                "ap_금액": 0,
                "금액차이": invoice_dict["합계금액"],
                "이체목표": ai_vals["구분"],
                "내부메모": ai_vals["내부메모"],
                "비고": "",
                "ap_row_num": None,
            }
            on_new(row_tv)

    def _handle_new_row_with_ai_vals(
        self,
        ws,
        header: List[str],
        invoice_dict: Dict,
        ai_vals: Dict,
        on_new: Optional[Callable]
    ) -> None:
        """신규 행 처리 (AI 값 미리 계산된 경우 - 병렬 처리용)"""
        # 신규 행 dict
        new_row = {
            "기안 날짜": invoice_dict["승인번호"][:8],
            "세금계산서 수령": invoice_dict["승인번호"],
            "이체 목표": "세금계산서 확인필요",
            "이체 날짜": "",
            "요청자": "정수현",
            "구분": ai_vals["구분"],
            "품목": invoice_dict["품목"],
            "문서번호": "",
            "비고": invoice_dict["품목"],
            "은행": ai_vals["은행"],
            "계좌번호": ai_vals["계좌번호"],
            "금액": invoice_dict["합계금액"],
            "수취인": invoice_dict["상호"],
            "공란": "",
            "CMS코드(공란)": "",
            "본인 통장표시내용": invoice_dict["상호"],
            "받는분 통장 표시": COMPANY_NAME,
            "내부 메모": ai_vals["내부메모"],
            self.TAX_CONTENT_COL: self._combine_invoice_content(invoice_dict),
        }
        # 실제 헤더 순서로 append
        ws.append([new_row.get(col, "") for col in header])

        # UI용 dict
        if on_new:
            row_tv = {
                "inv_승인번호": invoice_dict["승인번호"],
                "ap_세금계산서수령": "",
                "inv_상호": invoice_dict["상호"],
                "ap_수취인": "",
                "inv_품목": invoice_dict["품목"],
                "ap_품목": "",
                "inv_합계금액": invoice_dict["합계금액"],
                "ap_금액": 0,
                "금액차이": invoice_dict["합계금액"],
                "이체목표": ai_vals["구분"],
                "내부메모": ai_vals["내부메모"],
                "비고": "",
                "ap_row_num": None,
            }
            on_new(row_tv)
