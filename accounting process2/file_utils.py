"""
파일 처리 유틸리티
==================
다양한 파일 형식(PDF, Excel, 이미지 등)에서 데이터를 읽어오는 기능
"""

import os
import base64
from typing import Dict

import pandas as pd

# 선택적 라이브러리
try:
    import PyPDF2
    HAS_PYPDF2 = True
except ImportError:
    HAS_PYPDF2 = False

try:
    from docx import Document as DocxDocument
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

try:
    from PIL import Image
    import io
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class FileReader:
    """다양한 파일 형식에서 텍스트/이미지 추출"""

    @staticmethod
    def read_file(file_path: str) -> Dict:
        """
        파일을 읽어서 AI에 전달할 형식으로 변환

        Args:
            file_path: 파일 경로

        Returns:
            Dict:
                - type: "text" | "image" | "error"
                - content: 파일 내용
                - filename: 파일명
                - error: 오류 메시지 (있으면)
        """
        if not os.path.exists(file_path):
            return {
                "type": "error",
                "content": "",
                "filename": os.path.basename(file_path),
                "error": "파일을 찾을 수 없습니다"
            }

        ext = os.path.splitext(file_path)[1].lower().lstrip('.')
        filename = os.path.basename(file_path)

        try:
            if ext == 'pdf':
                return FileReader._read_pdf(file_path, filename)
            elif ext in ['png', 'jpg', 'jpeg']:
                return FileReader._read_image(file_path, filename)
            elif ext in ['xlsx', 'xls']:
                return FileReader._read_excel(file_path, filename)
            elif ext == 'csv':
                return FileReader._read_csv(file_path, filename)
            elif ext == 'docx':
                return FileReader._read_docx(file_path, filename)
            else:
                return {
                    "type": "error",
                    "content": "",
                    "filename": filename,
                    "error": f"지원하지 않는 파일 형식: {ext}"
                }
        except Exception as e:
            return {
                "type": "error",
                "content": "",
                "filename": filename,
                "error": str(e)
            }

    @staticmethod
    def _read_pdf(file_path: str, filename: str) -> Dict:
        """PDF 텍스트 추출"""
        if not HAS_PYPDF2:
            return {
                "type": "error",
                "content": "",
                "filename": filename,
                "error": "PyPDF2 라이브러리가 설치되지 않았습니다"
            }

        text_content = []
        with open(file_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            for page_num, page in enumerate(reader.pages[:10], 1):
                text = page.extract_text()
                if text:
                    text_content.append(f"[페이지 {page_num}]\n{text}")

        content = "\n\n".join(text_content)
        if not content.strip():
            return {
                "type": "text",
                "content": "(PDF에서 텍스트를 추출할 수 없습니다)",
                "filename": filename,
                "error": None
            }

        return {
            "type": "text",
            "content": content[:8000],
            "filename": filename,
            "error": None
        }

    @staticmethod
    def _read_image(file_path: str, filename: str) -> Dict:
        """이미지를 base64로 인코딩"""
        if not HAS_PIL:
            return {
                "type": "error",
                "content": "",
                "filename": filename,
                "error": "PIL 라이브러리가 설치되지 않았습니다"
            }

        with Image.open(file_path) as img:
            if img.mode == 'RGBA':
                img = img.convert('RGB')

            max_size = (1500, 1500)
            img.thumbnail(max_size, Image.Resampling.LANCZOS)

            buffer = io.BytesIO()
            img.save(buffer, format='JPEG', quality=85)
            base64_image = base64.b64encode(buffer.getvalue()).decode('utf-8')

        return {
            "type": "image",
            "content": base64_image,
            "filename": filename,
            "error": None
        }

    @staticmethod
    def _read_excel(file_path: str, filename: str) -> Dict:
        """Excel 파일 읽기"""
        df = pd.read_excel(file_path, sheet_name=0, nrows=50)
        content = f"[Excel 파일: {filename}]\n"
        content += df.to_string(index=False, max_rows=30, max_cols=10)

        return {
            "type": "text",
            "content": content[:6000],
            "filename": filename,
            "error": None
        }

    @staticmethod
    def _read_csv(file_path: str, filename: str) -> Dict:
        """CSV 파일 읽기"""
        df = pd.read_csv(file_path, nrows=50, encoding='utf-8-sig')
        content = f"[CSV 파일: {filename}]\n"
        content += df.to_string(index=False, max_rows=30, max_cols=10)

        return {
            "type": "text",
            "content": content[:6000],
            "filename": filename,
            "error": None
        }

    @staticmethod
    def _read_docx(file_path: str, filename: str) -> Dict:
        """Word 문서 읽기"""
        if not HAS_DOCX:
            return {
                "type": "error",
                "content": "",
                "filename": filename,
                "error": "python-docx 라이브러리가 설치되지 않았습니다"
            }

        doc = DocxDocument(file_path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        content = "\n".join(paragraphs[:100])

        return {
            "type": "text",
            "content": content[:6000],
            "filename": filename,
            "error": None
        }


def detect_highlighted_cells(file_path: str) -> Dict[str, set]:
    """
    Excel 파일에서 음영(배경색) 처리된 셀 감지

    Args:
        file_path: Excel 파일 경로

    Returns:
        Dict[str, set]: 시트명 -> {(row_idx, col_idx), ...}
    """
    if not HAS_OPENPYXL:
        return {}

    highlighted_cells = {}

    try:
        wb = load_workbook(file_path, data_only=True)

        for sheet_name in wb.sheetnames:
            highlighted_cells[sheet_name] = set()
            ws = wb[sheet_name]

            # 헤더(1행) 제외, 데이터 행부터 검사
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=0):
                for col_idx, cell in enumerate(row):
                    if cell.fill and cell.fill.fgColor:
                        color = cell.fill.fgColor

                        # RGB 타입 배경색 확인 (흰색/없음 제외)
                        if (color.type == 'rgb' and color.rgb and
                            color.rgb not in ['00000000', 'FFFFFFFF']):
                            highlighted_cells[sheet_name].add((row_idx, col_idx))

                        # 인덱스 타입 배경색 확인
                        elif (color.type == 'indexed' and color.indexed and
                              color.indexed not in [0, 64]):
                            highlighted_cells[sheet_name].add((row_idx, col_idx))

        wb.close()

    except Exception as e:
        print(f"음영 감지 오류: {e}")

    return highlighted_cells


def load_excel_with_sheets(file_path: str) -> Dict[str, pd.DataFrame]:
    """
    Excel 파일의 모든 시트 로드

    Args:
        file_path: Excel 파일 경로

    Returns:
        Dict[str, pd.DataFrame]: 시트명 -> DataFrame
    """
    sheets = {}

    if file_path.endswith('.csv'):
        df = pd.read_csv(file_path, encoding='utf-8-sig')
        sheets["Sheet1"] = df
    else:
        excel_file = pd.ExcelFile(file_path)
        for sheet_name in excel_file.sheet_names:
            sheets[sheet_name] = pd.read_excel(file_path, sheet_name=sheet_name)

    return sheets
