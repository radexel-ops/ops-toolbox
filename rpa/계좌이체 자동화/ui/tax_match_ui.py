# -*- coding: utf-8 -*-
"""
세금계산서 - AP List 매칭 UI
"""

import os
import threading
import queue
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from config import BASE_DIR, TAX_MATCH_COLUMNS, COMPANY_NAME
from services.matching_engine import MatchingEngine
from utils.excel_utils import leave_comment, find_file_by_pattern, lookup_bank_by_supplier
from utils.text_utils import normalize_header


class TaxMatchApp(tk.Tk):
    """세금계산서 매칭 애플리케이션"""

    def __init__(self):
        super().__init__()
        self._ap_cache: List[Tuple] = []
        self._ap_header_map: Optional[Dict[str, int]] = None
        self.title("AP-세금계산서 매칭 도우미")
        self.geometry("1520x840")
        self.queue = queue.Queue()
        self._data_match: Dict[str, Dict] = {}
        self._data_new: Dict[str, Dict] = {}
        self.matching_engine = MatchingEngine()
        self._build_ui()
        self._poll()

    def _build_ui(self):
        """UI 구성"""
        self._build_file_section()
        self._build_button_section()
        self._build_treeview_section()

    def _build_file_section(self):
        """파일 선택 섹션"""
        frm_top = ttk.Frame(self)
        frm_top.pack(fill="x", padx=8, pady=4)

        default_ap = Path(find_file_by_pattern("AP List_RDXL_updated*.xlsx")).name
        default_inv = Path(find_file_by_pattern("매입전자세금계산서목록*.xls")).name

        self.ap_var = tk.StringVar(value=default_ap)
        self.inv_var = tk.StringVar(value=default_inv)

        for i, (lbl, var, cmd) in enumerate([
            ("AP 파일", self.ap_var, self._sel_ap),
            ("세금계산서 파일", self.inv_var, self._sel_inv)
        ]):
            ttk.Label(frm_top, text=lbl).grid(row=i, column=0, sticky="e")
            ttk.Entry(frm_top, width=50, textvariable=var).grid(row=i, column=1, padx=4)
            ttk.Button(frm_top, text="찾기", command=cmd).grid(row=i, column=2, padx=4)

    def _build_button_section(self):
        """버튼 섹션"""
        frm_btn = ttk.Frame(self)
        frm_btn.pack(pady=4)
        ttk.Button(frm_btn, text="1. 실행", command=self._run_thread).pack(side="left", padx=4)
        ttk.Button(frm_btn, text="2. 최종 파일 생성", command=self._finalize).pack(side="left", padx=4)

        self.prog = ttk.Progressbar(frm_btn, length=200)
        self.prog.pack(side="left", padx=10)
        self.perc = ttk.Label(frm_btn, text="0 %")
        self.perc.pack(side="left")

    def _build_treeview_section(self):
        """트리뷰 섹션"""
        # 매칭 트리뷰
        self.tv_match = self._make_tree("매칭 결과")
        self.tv_match.pack(fill="both", expand=True, padx=8, pady=4)

        ttk.Label(self, text="신규 세금계산서").pack(anchor="w", padx=12)

        # 신규 트리뷰
        self.tv_new = self._make_tree("신규")
        self.tv_new.pack(fill="both", expand=False, padx=8, pady=(0, 8), ipady=80)

        # 삭제 버튼
        self.btn_del = ttk.Button(self, text="선택 행 삭제", command=self._delete_new_rows)
        self.btn_del.pack(pady=(0, 8))

        # 클릭 이벤트
        self.tv_match.bind("<Button-1>", self._on_click_match)
        self.tv_new.bind("<Button-1>", self._on_click_new)
        self.tv_match.bind("<Double-1>", self._start_edit)
        self.tv_new.bind("<Double-1>", self._start_edit)

    def _make_tree(self, title: str) -> ttk.Treeview:
        """트리뷰 생성"""
        tv = ttk.Treeview(
            self,
            columns=[c for c, _ in TAX_MATCH_COLUMNS],
            show="headings",
            height=10
        )
        for cid, w in TAX_MATCH_COLUMNS:
            tv.heading(cid, text=cid, anchor="center")
            tv.column(cid, width=w, anchor="center")
        tv.tag_configure("link", foreground="black")
        tv.tag_configure("diff", background="#FFB6B6")
        vsb = ttk.Scrollbar(tv, orient="vertical", command=tv.yview)
        tv.configure(yscroll=vsb.set)
        vsb.pack(side="right", fill="y")
        return tv

    def _sel_ap(self):
        """AP 파일 선택"""
        self.ap_var.set(Path(filedialog.askopenfilename()).name)

    def _sel_inv(self):
        """세금계산서 파일 선택"""
        self.inv_var.set(Path(filedialog.askopenfilename()).name)

    def _col(self, name: str) -> int:
        """열 인덱스 조회 (캐싱)"""
        if self._ap_header_map is None:
            wb = load_workbook(BASE_DIR / self.ap_var.get(), data_only=True)
            ws = wb["2025"]
            self._ap_header_map = {
                normalize_header(c.value): i + 1
                for i, c in enumerate(ws[1])
            }
        try:
            return self._ap_header_map[normalize_header(name)]
        except KeyError:
            raise KeyError(f"엑셀 헤더에서 '{name}'를 찾지 못했습니다.")

    def _load_ap_cache(self, ap_path: str):
        """AP 캐시 로드"""
        self._ap_header_map = None
        wb = load_workbook(ap_path, data_only=True)
        ws = wb["2025"]
        for r in range(2, ws.max_row + 1):
            self._ap_cache.append((
                r,
                str(ws.cell(r, self._col("수취인")).value or ""),
                str(ws.cell(r, self._col("세금계산서 수령")).value or ""),
                str(ws.cell(r, self._col("비고")).value or ""),
                ws.cell(r, self._col("금액")).value or 0,
            ))

    def _run_thread(self):
        """실행 쓰레드 시작"""
        full_ap = BASE_DIR / self.ap_var.get() if os.sep not in self.ap_var.get() else self.ap_var.get()
        self._ap_cache.clear()
        self._load_ap_cache(str(full_ap))

        for tv in (self.tv_match, self.tv_new):
            tv.delete(*tv.get_children())
        self._data_match.clear()
        self._data_new.clear()
        self.prog["value"] = 0
        self.perc.config(text="0 %")
        threading.Thread(target=self._worker, daemon=True).start()

    def _worker(self):
        """워커 쓰레드"""
        def on_progress(done, total):
            self.queue.put(("progress", done, total))

        def on_match(row):
            self.queue.put(("match", row))

        def on_new(row):
            self.queue.put(("new", row))

        full = lambda p: str(BASE_DIR / p) if p and os.sep not in p else p or None

        try:
            self.matching_engine.run_match(
                ap_path=full(self.ap_var.get()),
                inv_path=full(self.inv_var.get()),
                on_progress=on_progress,
                on_match=on_match,
                on_new=on_new,
            )
            self.queue.put(("done",))
        except Exception as e:
            self.queue.put(("error", str(e)))

    def _poll(self):
        """UI 쓰레드 큐 처리"""
        try:
            while True:
                tag, *payload = self.queue.get_nowait()
                if tag == "progress":
                    d, t = payload
                    pct = int(d * 100 / t)
                    self.prog["value"] = pct
                    self.perc.config(text=f"{pct} %")
                elif tag == "match":
                    self._insert_row(self.tv_match, self._data_match, payload[0], is_match=True)
                elif tag == "new":
                    self._insert_row(self.tv_new, self._data_new, payload[0], is_match=False)
                elif tag == "done":
                    messagebox.showinfo("완료", "미리보기가 준비되었습니다.")
                elif tag == "error":
                    messagebox.showerror("오류", payload[0])
        except queue.Empty:
            pass
        self.after(100, self._poll)

    def _insert_row(self, tv: ttk.Treeview, store: Dict, row: Dict, is_match: bool):
        """행 삽입"""
        row = row.copy()
        row["ACTION"] = "👉 거부" if is_match else "👉 승인"
        vals = [row.get(c, "") for c, _ in TAX_MATCH_COLUMNS]
        iid = tv.insert("", "end", values=vals, tags=("link",))
        store[iid] = row
        if is_match and abs(row.get("금액차이", 0)) >= 1:
            tv.item(iid, tags=("link", "diff"))

    def _on_click_match(self, event):
        """매칭 트리뷰 클릭"""
        iid = self.tv_match.identify_row(event.y)
        col = self.tv_match.identify_column(event.x)
        if not iid or col != f"#{len(TAX_MATCH_COLUMNS)}":
            return
        row = self._data_match.pop(iid)
        self.tv_match.delete(iid)
        self._insert_row(self.tv_new, self._data_new, row, is_match=False)

    def _on_click_new(self, event):
        """신규 트리뷰 클릭"""
        iid = self.tv_new.identify_row(event.y)
        col = self.tv_new.identify_column(event.x)
        if not iid or col != f"#{len(TAX_MATCH_COLUMNS)}":
            return

        row = self._data_new.get(iid)
        if row is None:
            return

        if not row["ap_세금계산서수령"]:
            self._popup_match(iid, row)
            return

        self._data_new.pop(iid)
        self.tv_new.delete(iid)
        self._insert_row(self.tv_match, self._data_match, row, is_match=True)

    def _popup_match(self, iid: str, row: Dict):
        """AP 행 검색 팝업"""
        top = tk.Toplevel(self)
        top.title("AP 행 검색")
        top.geometry("600x400")
        tk.Label(top, text="검색:").pack(anchor="w", padx=6, pady=4)

        sv = tk.StringVar()
        ent = ttk.Entry(top, textvariable=sv)
        ent.pack(fill="x", padx=8)

        lb = tk.Listbox(top)
        lb.pack(fill="both", expand=True, padx=8, pady=4)

        def refresh(*_):
            key = sv.get().lower()
            lb.delete(0, "end")
            for rnum, nm, acc, item, amt in self._ap_cache:
                line = f"{rnum:>4} | {nm} | {acc} | {item} | {amt:,.0f}"
                if key in line.lower():
                    lb.insert("end", line)

        sv.trace_add("write", refresh)
        refresh()

        def choose(_=None):
            if not lb.curselection():
                return
            sel = lb.get(lb.curselection()[0]).split(" | ")[0].strip()
            ap_row = int(sel)

            for d in self._data_match.values():
                if d["ap_row_num"] == ap_row:
                    messagebox.showwarning("중복", "이미 다른 항목과 매칭된 행입니다.")
                    return

            row_tuple = next(t for t in self._ap_cache if t[0] == ap_row)
            _, nm, acc, item, amt = row_tuple

            row.update({
                "ap_row_num": ap_row,
                "ap_세금계산서수령": acc,
                "ap_수취인": nm,
                "ap_품목": item,
                "ap_금액": amt,
                "금액차이": row["inv_합계금액"] - amt
            })

            self.tv_new.delete(iid)
            self._data_new.pop(iid, None)
            self._insert_row(self.tv_match, self._data_match, row, is_match=True)
            top.destroy()

        lb.bind("<Double-1>", choose)
        ent.focus()
        top.transient(self)
        top.grab_set()

    def _delete_new_rows(self):
        """신규 행 삭제"""
        iids = self.tv_new.selection()
        if not iids:
            messagebox.showwarning("경고", "삭제할 행을 먼저 선택하세요.")
            return

        if not messagebox.askokcancel("삭제 확인", "선택한 신규 세금계산서를 정말 삭제하시겠습니까?"):
            return

        for iid in iids:
            self.tv_new.delete(iid)
            self._data_new.pop(iid, None)

    def _start_edit(self, event):
        """셀 직접 편집"""
        tv = event.widget
        iid = tv.identify_row(event.y)
        col = tv.identify_column(event.x)
        if not iid or col == "#0":
            return
        col_idx = int(col[1:]) - 1
        col_id = TAX_MATCH_COLUMNS[col_idx][0]

        x, y, w, h = tv.bbox(iid, col)
        e = ttk.Entry(tv)
        e.place(x=x, y=y, width=w, height=h)
        e.insert(0, tv.set(iid, col_id))
        e.focus()

        def _save(_=None):
            val = e.get()
            e.destroy()
            tv.set(iid, col_id, val)
            store = self._data_match if tv is self.tv_match else self._data_new
            store[iid][col_id] = val

        e.bind("<Return>", _save)
        e.bind("<FocusOut>", _save)

    def _finalize(self):
        """최종 파일 생성"""
        if not self._data_match and not self._data_new:
            messagebox.showwarning("경고", "승인 또는 신규 항목이 없습니다.")
            return

        try:
            wb = load_workbook(Path(BASE_DIR) / self.ap_var.get())
            ws = wb["2025"]
            ap_cols = {c.value: i + 1 for i, c in enumerate(ws[1])}

            yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

            # 매칭된 행 업데이트
            for row in self._data_match.values():
                rnum = row.get("ap_row_num")
                if not rnum:
                    continue
                self._update_matched_row(ws, ap_cols, row, yellow_fill)

            # 신규 행 추가
            if self._data_new:
                self._add_new_rows(ws, ap_cols, yellow_fill)

            wb.save(Path(BASE_DIR) / "AP List_RDXL_merged.xlsx")
            messagebox.showinfo("완료", "AP List_RDXL_merged.xlsx 덮어썼습니다.")
        except Exception as e:
            messagebox.showerror("오류", str(e))

    def _update_matched_row(self, ws, ap_cols: Dict, row: Dict, fill):
        """매칭된 행 업데이트"""
        rnum = row["ap_row_num"]

        def swap(col, new_val):
            cell = ws.cell(rnum, ap_cols[col])
            old = str(cell.value or "")
            if col == "비고":
                cell.value = f"{old}, {new_val}" if old else new_val
                return
            if old:
                leave_comment(ws, cell.coordinate, f"기존 : {old}")
            cell.value = new_val

        swap("세금계산서 수령", row["inv_승인번호"])
        swap("수취인", row["inv_상호"])
        swap("비고", row["inv_품목"])
        swap("본인 통장표시내용", row["inv_상호"])

        inv_amt = float(row["inv_합계금액"])
        ap_amt = float(row["ap_금액"])
        if abs(inv_amt - ap_amt) >= 1:
            leave_comment(
                ws,
                ws.cell(rnum, ap_cols["금액"]).coordinate,
                f"세금계산서 금액 : {int(inv_amt):,}"
            )

        for col_idx in range(1, len(ap_cols) + 1):
            ws.cell(row=rnum, column=col_idx).fill = fill

    def _add_new_rows(self, ws, ap_cols: Dict, fill):
        """신규 행 추가"""
        header = [c.value for c in ws[1]]
        acc_df = pd.read_excel(
            Path(BASE_DIR) / self.ap_var.get(),
            sheet_name="계좌정보"
        ).rename(columns=str).fillna("")

        for tv in self._data_new.values():
            bank, acct = lookup_bank_by_supplier(tv["inv_상호"], acc_df)

            category = tv["이체목표"] or "검토 필요"
            if "용산센트럴파크 단지 관리단" in tv["inv_상호"]:
                category = "관리비"

            new_row = {
                "기안 날짜": tv["inv_승인번호"][:8],
                "세금계산서 수령": tv["inv_승인번호"],
                "이체 목표": "세금계산서 확인필요",
                "이체 날짜": "",
                "요청자": "정수현",
                "구분": category,
                "품목": tv["inv_품목"],
                "문서번호": "",
                "비고": tv["inv_품목"],
                "은행": bank,
                "계좌번호": acct,
                "금액": tv["inv_합계금액"],
                "수취인": tv["inv_상호"],
                "공란": "",
                "CMS코드(공란)": "",
                "본인 통장표시내용": tv["inv_상호"],
                "받는분 통장 표시": COMPANY_NAME,
                "내부 메모": f"({category}){tv['inv_품목']}"[:20],
            }
            ws.append([new_row.get(col, "") for col in header])

            new_row_num = ws.max_row
            for col_idx in range(1, len(header) + 1):
                ws.cell(row=new_row_num, column=col_idx).fill = fill


def main():
    """메인 함수"""
    app = TaxMatchApp()
    app.mainloop()


if __name__ == "__main__":
    main()
