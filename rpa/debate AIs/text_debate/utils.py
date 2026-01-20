"""
AI Debate System - Utility Functions
파일 처리, HTML 로깅, 헬퍼 함수들
"""

import os
import base64
import mimetypes
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path

import chardet
import markdown
from pygments.formatters import HtmlFormatter
from rich.console import Console

console = Console()


class FileProcessor:
    """다양한 파일 형식을 처리하는 클래스"""

    def __init__(self, config: dict):
        self.config = config
        self.max_text_size = config.get('file_processing', {}).get('max_text_size', 1048576)
        self.supported = config.get('file_processing', {}).get('supported_extensions', {})

    def get_file_type(self, file_path: str) -> str:
        """파일 확장자로 파일 타입 판별"""
        ext = Path(file_path).suffix.lower()

        for file_type, extensions in self.supported.items():
            if ext in extensions:
                return file_type

        return "unknown"

    def read_text_file(self, file_path: str) -> str:
        """텍스트 파일 읽기 (인코딩 자동 감지)"""
        try:
            with open(file_path, 'rb') as f:
                raw_data = f.read(self.max_text_size)

            # 인코딩 감지
            detected = chardet.detect(raw_data)
            encoding = detected.get('encoding', 'utf-8') or 'utf-8'

            return raw_data.decode(encoding, errors='replace')
        except Exception as e:
            return f"[파일 읽기 오류: {str(e)}]"

    def read_pdf(self, file_path: str) -> str:
        """PDF 파일에서 텍스트 추출"""
        try:
            from pypdf import PdfReader
            reader = PdfReader(file_path)
            text_parts = []
            for i, page in enumerate(reader.pages):
                text = page.extract_text()
                if text:
                    text_parts.append(f"[Page {i + 1}]\n{text}")
            return "\n\n".join(text_parts) if text_parts else "[PDF에서 텍스트를 추출할 수 없습니다]"
        except ImportError:
            return "[pypdf 라이브러리가 설치되지 않았습니다]"
        except Exception as e:
            return f"[PDF 읽기 오류: {str(e)}]"

    def read_docx(self, file_path: str) -> str:
        """Word 문서에서 텍스트 추출"""
        try:
            from docx import Document
            doc = Document(file_path)
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n\n".join(paragraphs) if paragraphs else "[문서에서 텍스트를 추출할 수 없습니다]"
        except ImportError:
            return "[python-docx 라이브러리가 설치되지 않았습니다]"
        except Exception as e:
            return f"[DOCX 읽기 오류: {str(e)}]"

    def read_xlsx(self, file_path: str) -> str:
        """Excel 파일에서 데이터 추출"""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, data_only=True)
            result_parts = []

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    row_data = [str(cell) if cell is not None else "" for cell in row]
                    if any(row_data):  # 빈 행 제외
                        rows.append(" | ".join(row_data))

                if rows:
                    result_parts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))

            return "\n\n".join(result_parts) if result_parts else "[Excel에서 데이터를 추출할 수 없습니다]"
        except ImportError:
            return "[openpyxl 라이브러리가 설치되지 않았습니다]"
        except Exception as e:
            return f"[XLSX 읽기 오류: {str(e)}]"

    def encode_image_base64(self, file_path: str, chunk_size: int = 8192) -> Tuple[str, str]:
        """이미지를 base64로 인코딩 (청크 단위 스트리밍으로 메모리 효율화)"""
        try:
            mime_type, _ = mimetypes.guess_type(file_path)
            mime_type = mime_type or "image/png"

            # 청크 단위로 읽어서 base64 인코딩 (메모리 효율화)
            chunks = []
            with open(file_path, 'rb') as f:
                while True:
                    chunk = f.read(chunk_size)
                    if not chunk:
                        break
                    chunks.append(chunk)

            # 전체 바이트를 한 번에 인코딩
            encoded = base64.b64encode(b''.join(chunks)).decode('utf-8')
            return encoded, mime_type
        except (IOError, OSError, PermissionError) as e:
            console.print(f"[red]이미지 인코딩 오류 ({file_path}): {e}[/red]")
            return "", ""

    def process_files_unified(self, file_paths: List[str]) -> Dict[str, Any]:
        """파일을 한 번만 읽고 Gemini/GPT 형식으로 변환 (I/O 최적화)

        Returns:
            Dict with 'gemini' and 'gpt' keys containing processed files for each API
        """
        try:
            import google.generativeai as genai
            has_genai = True
        except ImportError:
            has_genai = False

        result = {'gemini': [], 'gpt': []}

        for file_path in file_paths:
            if not os.path.exists(file_path):
                console.print(f"[yellow]파일을 찾을 수 없음: {file_path}[/yellow]")
                continue

            file_type = self.get_file_type(file_path)
            file_name = os.path.basename(file_path)

            try:
                if file_type == "text":
                    # 텍스트는 한 번만 읽고 두 API에 사용
                    content = self.read_text_file(file_path)
                    text_content = f"[File: {file_name}]\n```\n{content}\n```"

                    result['gemini'].append(text_content)
                    result['gpt'].append({
                        "type": "text",
                        "text": text_content
                    })

                elif file_type == "image":
                    # 이미지: Gemini는 File API, GPT는 base64
                    if has_genai:
                        console.print(f"[cyan]Gemini에 이미지 업로드 중: {file_name}[/cyan]")
                        uploaded = genai.upload_file(file_path)
                        result['gemini'].append(uploaded)

                    encoded, mime_type = self.encode_image_base64(file_path)
                    if encoded:
                        result['gpt'].append({
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:{mime_type};base64,{encoded}"
                            }
                        })

                elif file_type == "document":
                    ext = Path(file_path).suffix.lower()

                    if ext == ".pdf":
                        # Gemini: File API 사용, GPT: 텍스트 추출
                        if has_genai:
                            console.print(f"[cyan]Gemini에 PDF 업로드 중: {file_name}[/cyan]")
                            uploaded = genai.upload_file(file_path)
                            result['gemini'].append(uploaded)

                        content = self.read_pdf(file_path)
                        result['gpt'].append({
                            "type": "text",
                            "text": f"[File: {file_name}]\n{content}"
                        })
                    elif ext == ".docx":
                        content = self.read_docx(file_path)
                        text_content = f"[File: {file_name}]\n{content}"
                        result['gemini'].append(text_content)
                        result['gpt'].append({
                            "type": "text",
                            "text": text_content
                        })
                    elif ext == ".xlsx":
                        content = self.read_xlsx(file_path)
                        text_content = f"[File: {file_name}]\n{content}"
                        result['gemini'].append(text_content)
                        result['gpt'].append({
                            "type": "text",
                            "text": text_content
                        })
                    else:
                        content = self.read_text_file(file_path)
                        text_content = f"[File: {file_name}]\n{content}"
                        result['gemini'].append(text_content)
                        result['gpt'].append({
                            "type": "text",
                            "text": text_content
                        })

                elif file_type == "media":
                    # 미디어: Gemini만 File API 지원
                    if has_genai:
                        console.print(f"[cyan]Gemini에 미디어 파일 업로드 중: {file_name}[/cyan]")
                        uploaded = genai.upload_file(file_path)
                        result['gemini'].append(uploaded)

                    result['gpt'].append({
                        "type": "text",
                        "text": f"[File: {file_name}] - 미디어 파일입니다. GPT는 현재 비디오/오디오를 직접 처리할 수 없습니다."
                    })

                else:
                    # 알 수 없는 형식
                    content = self.read_text_file(file_path)
                    text_content = f"[File: {file_name}]\n{content}"
                    result['gemini'].append(text_content)
                    result['gpt'].append({
                        "type": "text",
                        "text": text_content
                    })

            except (IOError, OSError, ValueError) as e:
                console.print(f"[red]파일 처리 오류 ({file_name}): {e}[/red]")
                error_text = f"[File: {file_name}] - 처리 중 오류 발생: {str(e)}"
                result['gemini'].append(error_text)
                result['gpt'].append({
                    "type": "text",
                    "text": error_text
                })

        return result

    def process_for_gemini(self, file_paths: List[str]) -> List[Any]:
        """Gemini API용 파일 처리"""
        try:
            import google.generativeai as genai
        except ImportError:
            console.print("[red]google-generativeai 라이브러리가 설치되지 않았습니다[/red]")
            return []

        processed = []

        for file_path in file_paths:
            if not os.path.exists(file_path):
                console.print(f"[yellow]파일을 찾을 수 없음: {file_path}[/yellow]")
                continue

            file_type = self.get_file_type(file_path)
            file_name = os.path.basename(file_path)

            try:
                if file_type == "text":
                    content = self.read_text_file(file_path)
                    processed.append(f"[File: {file_name}]\n```\n{content}\n```")

                elif file_type == "image":
                    # Gemini File API로 업로드
                    console.print(f"[cyan]Gemini에 이미지 업로드 중: {file_name}[/cyan]")
                    uploaded = genai.upload_file(file_path)
                    processed.append(uploaded)

                elif file_type == "document":
                    ext = Path(file_path).suffix.lower()
                    if ext == ".pdf":
                        # PDF는 File API로 업로드 (Gemini가 직접 처리 가능)
                        console.print(f"[cyan]Gemini에 PDF 업로드 중: {file_name}[/cyan]")
                        uploaded = genai.upload_file(file_path)
                        processed.append(uploaded)
                    elif ext == ".docx":
                        content = self.read_docx(file_path)
                        processed.append(f"[File: {file_name}]\n{content}")
                    elif ext == ".xlsx":
                        content = self.read_xlsx(file_path)
                        processed.append(f"[File: {file_name}]\n{content}")
                    else:
                        content = self.read_text_file(file_path)
                        processed.append(f"[File: {file_name}]\n{content}")

                elif file_type == "media":
                    # 비디오/오디오는 File API로 업로드
                    console.print(f"[cyan]Gemini에 미디어 파일 업로드 중: {file_name}[/cyan]")
                    uploaded = genai.upload_file(file_path)
                    processed.append(uploaded)

                else:
                    # 알 수 없는 형식은 텍스트로 시도
                    content = self.read_text_file(file_path)
                    processed.append(f"[File: {file_name}]\n{content}")

            except Exception as e:
                console.print(f"[red]파일 처리 오류 ({file_name}): {e}[/red]")
                processed.append(f"[File: {file_name}] - 처리 중 오류 발생: {str(e)}")

        return processed

    def process_for_gpt(self, file_paths: List[str]) -> List[Dict[str, Any]]:
        """GPT API용 파일 처리"""
        processed = []

        for file_path in file_paths:
            if not os.path.exists(file_path):
                console.print(f"[yellow]파일을 찾을 수 없음: {file_path}[/yellow]")
                continue

            file_type = self.get_file_type(file_path)
            file_name = os.path.basename(file_path)

            try:
                if file_type == "text":
                    content = self.read_text_file(file_path)
                    processed.append({
                        "type": "text",
                        "text": f"[File: {file_name}]\n```\n{content}\n```"
                    })

                elif file_type == "image":
                    encoded, mime_type = self.encode_image_base64(file_path)
                    if encoded:
                        processed.append({
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:{mime_type};base64,{encoded}"
                            }
                        })

                elif file_type == "document":
                    ext = Path(file_path).suffix.lower()
                    if ext == ".pdf":
                        content = self.read_pdf(file_path)
                    elif ext == ".docx":
                        content = self.read_docx(file_path)
                    elif ext == ".xlsx":
                        content = self.read_xlsx(file_path)
                    else:
                        content = self.read_text_file(file_path)

                    processed.append({
                        "type": "text",
                        "text": f"[File: {file_name}]\n{content}"
                    })

                elif file_type == "media":
                    # GPT는 비디오/오디오 직접 처리 불가 (현재 기준)
                    processed.append({
                        "type": "text",
                        "text": f"[File: {file_name}] - 미디어 파일입니다. GPT는 현재 비디오/오디오를 직접 처리할 수 없습니다."
                    })

                else:
                    content = self.read_text_file(file_path)
                    processed.append({
                        "type": "text",
                        "text": f"[File: {file_name}]\n{content}"
                    })

            except Exception as e:
                console.print(f"[red]파일 처리 오류 ({file_name}): {e}[/red]")
                processed.append({
                    "type": "text",
                    "text": f"[File: {file_name}] - 처리 중 오류 발생: {str(e)}"
                })

        return processed


class HtmlLogger:
    """토론 기록을 HTML로 저장하는 클래스"""

    def __init__(self, output_dir: str, config: dict):
        self.output_dir = output_dir
        self.config = config

        # 디렉토리 생성
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        # 파일명 생성
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.filename = os.path.join(output_dir, f"debate_{timestamp}.html")

        # 세션 정보
        self.session_info = {
            "start_time": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "gemini_model": "",
            "gpt_model": "",
            "files": []
        }

        # 대화 기록
        self.entries: List[Dict[str, Any]] = []

        # 최종 결론
        self.final_conclusion: Optional[Dict[str, str]] = None

    def set_session_info(self, gemini_model: str, gpt_model: str, files: List[str]):
        """세션 정보 설정"""
        self.session_info["gemini_model"] = gemini_model
        self.session_info["gpt_model"] = gpt_model
        self.session_info["files"] = files

    def add_entry(self, role: str, model: str, content: str,
                  files: Optional[List[str]] = None, is_consensus: bool = False):
        """대화 항목 추가 (저장은 save() 호출 시 수행)"""
        entry = {
            "role": role,  # "user", "gemini", "gpt", "system"
            "model": model,
            "content": content,
            "timestamp": datetime.now().strftime('%H:%M:%S'),
            "files": files or [],
            "is_consensus": is_consensus
        }
        self.entries.append(entry)
        # 성능 최적화: 매번 저장하지 않고 save() 호출 시에만 저장

    def save(self):
        """HTML 파일 저장 (토론 종료 시 또는 필요 시 호출)"""
        self._save_html()

    def _get_css(self) -> str:
        """HTML 스타일시트"""
        theme = self.config.get('html_output', {}).get('theme', 'modern')
        code_style = self.config.get('html_output', {}).get('code_style', 'monokai')

        # Pygments 코드 하이라이팅 CSS
        try:
            formatter = HtmlFormatter(style=code_style)
            pygments_css = formatter.get_style_defs('.codehilite')
        except (ValueError, KeyError, ImportError) as e:
            # 스타일을 찾을 수 없거나 라이브러리 문제 시 기본값 사용
            pygments_css = ""

        return f"""
        :root {{
            --user-bg: #e3f2fd;
            --user-border: #2196f3;
            --gemini-bg: #fff3e0;
            --gemini-border: #ff9800;
            --gpt-bg: #e8f5e9;
            --gpt-border: #4caf50;
            --system-bg: #f5f5f5;
            --system-border: #9e9e9e;
            --consensus-bg: #fce4ec;
            --consensus-border: #e91e63;
        }}

        * {{
            box-sizing: border-box;
        }}

        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            max-width: 900px;
            margin: 0 auto;
            padding: 20px;
            background-color: #f8f9fa;
            line-height: 1.6;
        }}

        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px;
            border-radius: 15px;
            margin-bottom: 30px;
            box-shadow: 0 4px 15px rgba(0,0,0,0.2);
        }}

        .header h1 {{
            margin: 0 0 10px 0;
            font-size: 1.8em;
        }}

        .header .session-info {{
            font-size: 0.9em;
            opacity: 0.9;
        }}

        .chat-container {{
            display: flex;
            flex-direction: column;
            gap: 20px;
        }}

        .message {{
            padding: 20px;
            border-radius: 15px;
            position: relative;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}

        .message.user {{
            background-color: var(--user-bg);
            border-left: 5px solid var(--user-border);
            margin-right: 50px;
        }}

        .message.gemini {{
            background-color: var(--gemini-bg);
            border-left: 5px solid var(--gemini-border);
            margin-left: 30px;
            margin-right: 20px;
        }}

        .message.gpt {{
            background-color: var(--gpt-bg);
            border-left: 5px solid var(--gpt-border);
            margin-left: 20px;
            margin-right: 30px;
        }}

        .message.system {{
            background-color: var(--system-bg);
            border-left: 5px solid var(--system-border);
            text-align: center;
            font-style: italic;
        }}

        .message.consensus {{
            background-color: var(--consensus-bg);
            border: 3px solid var(--consensus-border);
            margin: 30px 0;
        }}

        .message-header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 10px;
            padding-bottom: 10px;
            border-bottom: 1px solid rgba(0,0,0,0.1);
        }}

        .message-role {{
            font-weight: bold;
            font-size: 1.1em;
        }}

        .message-role.user {{ color: #1565c0; }}
        .message-role.gemini {{ color: #e65100; }}
        .message-role.gpt {{ color: #2e7d32; }}

        .message-model {{
            font-size: 0.85em;
            color: #666;
            background: rgba(0,0,0,0.05);
            padding: 3px 8px;
            border-radius: 10px;
        }}

        .message-time {{
            font-size: 0.8em;
            color: #888;
        }}

        .message-content {{
            white-space: pre-wrap;
            word-wrap: break-word;
        }}

        .message-content p {{
            margin: 0.5em 0;
        }}

        .message-content pre {{
            background: #2d2d2d;
            color: #f8f8f2;
            padding: 15px;
            border-radius: 8px;
            overflow-x: auto;
            font-size: 0.9em;
        }}

        .message-content code {{
            background: rgba(0,0,0,0.1);
            padding: 2px 6px;
            border-radius: 4px;
            font-family: 'Consolas', 'Monaco', monospace;
        }}

        .message-content pre code {{
            background: none;
            padding: 0;
        }}

        .files-list {{
            margin-top: 10px;
            padding: 10px;
            background: rgba(0,0,0,0.05);
            border-radius: 8px;
            font-size: 0.9em;
        }}

        .files-list::before {{
            content: "📎 첨부파일: ";
            font-weight: bold;
        }}

        .consensus-badge {{
            display: inline-block;
            background: linear-gradient(135deg, #e91e63, #f06292);
            color: white;
            padding: 5px 15px;
            border-radius: 20px;
            font-weight: bold;
            margin-bottom: 15px;
        }}

        .divider {{
            text-align: center;
            margin: 30px 0;
            position: relative;
        }}

        .divider::before {{
            content: "";
            position: absolute;
            top: 50%;
            left: 0;
            right: 0;
            height: 1px;
            background: #ddd;
        }}

        .divider span {{
            background: #f8f9fa;
            padding: 0 20px;
            position: relative;
            color: #666;
            font-size: 0.9em;
        }}

        .footer {{
            margin-top: 40px;
            padding: 20px;
            text-align: center;
            color: #888;
            font-size: 0.85em;
            border-top: 1px solid #ddd;
        }}

        .final-conclusion-section {{
            margin-top: 40px;
            padding: 30px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            border-radius: 15px;
            box-shadow: 0 4px 20px rgba(0,0,0,0.2);
        }}

        .final-conclusion-section h2 {{
            color: white;
            margin: 0 0 20px 0;
            font-size: 1.5em;
        }}

        .final-conclusion-section .question-box {{
            background: rgba(255,255,255,0.15);
            color: white;
            padding: 15px;
            border-radius: 10px;
            margin-bottom: 15px;
        }}

        .final-conclusion-section .date-info {{
            color: rgba(255,255,255,0.8);
            font-size: 0.9em;
            margin-bottom: 15px;
        }}

        .final-conclusion-section .conclusion-content {{
            background: white;
            padding: 25px;
            border-radius: 10px;
            line-height: 1.8;
        }}

        .final-conclusion-section .conclusion-content h1,
        .final-conclusion-section .conclusion-content h2,
        .final-conclusion-section .conclusion-content h3 {{
            color: #667eea;
            margin-top: 20px;
        }}

        .final-conclusion-section .conclusion-content h1:first-child,
        .final-conclusion-section .conclusion-content h2:first-child,
        .final-conclusion-section .conclusion-content h3:first-child {{
            margin-top: 0;
        }}

        .final-conclusion-section .conclusion-content ul,
        .final-conclusion-section .conclusion-content ol {{
            margin-left: 20px;
        }}

        .final-conclusion-section .conclusion-content pre {{
            background: #2d2d2d;
            color: #f8f8f2;
            padding: 15px;
            border-radius: 8px;
            overflow-x: auto;
        }}

        {pygments_css}
        """

    def _render_content(self, content: str) -> str:
        """마크다운을 HTML로 변환"""
        try:
            html = markdown.markdown(
                content,
                extensions=['fenced_code', 'codehilite', 'tables', 'nl2br']
            )
            return html
        except (ValueError, TypeError, AttributeError) as e:
            # 마크다운 변환 실패 시 일반 텍스트로 (HTML 이스케이프 적용)
            import html as html_module
            escaped = html_module.escape(content)
            return f"<p>{escaped}</p>"

    def _save_html(self):
        """HTML 파일 저장"""
        css = self._get_css()

        html_parts = [
            "<!DOCTYPE html>",
            "<html lang='ko'>",
            "<head>",
            "  <meta charset='utf-8'>",
            "  <meta name='viewport' content='width=device-width, initial-scale=1'>",
            "  <title>AI Debate Log</title>",
            f"  <style>{css}</style>",
            "</head>",
            "<body>",
            "  <div class='header'>",
            "    <h1>AI Debate System - 토론 기록</h1>",
            "    <div class='session-info'>",
            f"      <div>시작 시간: {self.session_info['start_time']}</div>",
            f"      <div>Gemini: {self.session_info['gemini_model']} | GPT: {self.session_info['gpt_model']}</div>",
        ]

        if self.session_info['files']:
            files_str = ", ".join([os.path.basename(f) for f in self.session_info['files']])
            html_parts.append(f"      <div>첨부파일: {files_str}</div>")

        html_parts.extend([
            "    </div>",
            "  </div>",
            "  <div class='chat-container'>",
        ])

        # 대화 항목 렌더링
        for entry in self.entries:
            role_class = entry['role'].lower()
            consensus_class = " consensus" if entry['is_consensus'] else ""

            html_parts.append(f"    <div class='message {role_class}{consensus_class}'>")

            if entry['is_consensus']:
                html_parts.append("      <div class='consensus-badge'>합의 도출</div>")

            html_parts.append("      <div class='message-header'>")
            html_parts.append(f"        <span class='message-role {role_class}'>{entry['role'].upper()}</span>")

            if entry['model']:
                html_parts.append(f"        <span class='message-model'>{entry['model']}</span>")

            html_parts.append(f"        <span class='message-time'>{entry['timestamp']}</span>")
            html_parts.append("      </div>")

            # 콘텐츠
            rendered_content = self._render_content(entry['content'])
            html_parts.append(f"      <div class='message-content'>{rendered_content}</div>")

            # 첨부파일
            if entry['files']:
                files_str = ", ".join([os.path.basename(f) for f in entry['files']])
                html_parts.append(f"      <div class='files-list'>{files_str}</div>")

            html_parts.append("    </div>")

        # 최종 결론 섹션 추가
        if self.final_conclusion:
            conclusion_html = self._render_content(self.final_conclusion["conclusion"])
            html_parts.extend([
                "  </div>",  # chat-container 닫기
                "  <div class='final-conclusion-section'>",
                "    <h2>📋 최종 합의된 결론</h2>",
                "    <div class='question-box'>",
                f"      <strong>원래 질문:</strong> {self.final_conclusion['original_question']}",
                "    </div>",
                f"    <div class='date-info'>기준 날짜: {self.final_conclusion['date']}</div>",
                f"    <div class='conclusion-content'>{conclusion_html}</div>",
                "  </div>"
            ])
        else:
            html_parts.append("  </div>")  # chat-container 닫기

        html_parts.extend([
            "  <div class='footer'>",
            f"    Generated by AI Debate System | {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "  </div>",
            "</body>",
            "</html>"
        ])

        with open(self.filename, 'w', encoding='utf-8', errors='replace') as f:
            f.write("\n".join(html_parts))

    def add_final_conclusion(self, conclusion: str, original_question: str, date: str):
        """최종 합의된 결론을 별도 섹션으로 추가 (저장은 save() 호출 시 수행)"""
        self.final_conclusion = {
            "conclusion": conclusion,
            "original_question": original_question,
            "date": date
        }
        # 성능 최적화: save() 호출 시에만 저장

    def get_full_history_text(self) -> str:
        """전체 대화 기록을 텍스트로 반환 (컨텍스트 상속용)"""
        parts = []
        for entry in self.entries:
            role = entry['role'].upper()
            model = f" ({entry['model']})" if entry['model'] else ""
            parts.append(f"[{role}{model}]: {entry['content']}")
        return "\n\n".join(parts)
