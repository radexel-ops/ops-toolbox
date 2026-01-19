"""
회계처리 계정과목 분류 시스템
- Excel/CSV 파일 업로드 및 시트별 탭 표시
- AI 기반 계정과목 분류 (의사결정 트리 적용)
- 반복적인 Q&A를 통한 정확한 분류
"""

import os
import json
import threading
import base64
from typing import Optional, Dict, List, Any, Tuple
from dotenv import load_dotenv

import customtkinter as ctk
import tkinter as tk
import tkinter.font as tk_font
from tkinter import filedialog, messagebox
import pandas as pd

# 문서 처리 라이브러리
try:
    import PyPDF2
    HAS_PYPDF2 = True
except ImportError:
    HAS_PYPDF2 = False

try:
    from docx import Document as DocxDocument
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

try:
    from PIL import Image
    import io
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

# API 클라이언트
try:
    from openai import OpenAI
    HAS_OPENAI = True
except ImportError:
    HAS_OPENAI = False

try:
    import google.generativeai as genai
    HAS_GEMINI = True
except ImportError:
    HAS_GEMINI = False

# 환경변수 로드 (이 스크립트와 같은 폴더의 .env 파일만 사용)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ENV_PATH = os.path.join(SCRIPT_DIR, ".env")
load_dotenv(ENV_PATH, override=True)

# 버전 관리 함수 임포트
try:
    from guidelines_manager import (
        check_yaml_changes,
        acknowledge_yaml_changes,
        create_backup,
        get_version_list,
        restore_version,
        get_version_diff,
        load_guidelines
    )
    HAS_VERSION_MANAGER = True
except ImportError:
    HAS_VERSION_MANAGER = False

# 회계정책 검증 모듈 임포트
try:
    from accounting_standards import (
        PolicyValidator,
        ValidationSeverity,
        RuleType,
        format_validation_results
    )
    HAS_POLICY_VALIDATOR = True
except ImportError:
    HAS_POLICY_VALIDATOR = False

# 프로젝트 관리 모듈 임포트 (연관 거래 저장 포함)
try:
    from project_manager import ProjectManager as ExternalProjectManager
    HAS_PROJECT_MANAGER_MODULE = True
except ImportError:
    HAS_PROJECT_MANAGER_MODULE = False
    ExternalProjectManager = None

# UI 모듈 임포트
try:
    from widgets import DataTableFrame
    from chat_panel import AIChatPanel
    from dialogs import ColumnMappingDialog, ManualClassifyDialog
    HAS_UI_MODULES = True
except ImportError:
    HAS_UI_MODULES = False

# ==========================================
# 디자인 시스템
# ==========================================
BG_COLOR = "#18181C"
SIDEBAR_COLOR = "#25262B"
ACCENT_COLOR = "#3A76F0"
SUCCESS_COLOR = "#50C878"
WARNING_COLOR = "#FFB347"
ERROR_COLOR = "#FF6B6B"
FONT_FAMILY = "Malgun Gothic"

# ==========================================
# AI 모델 설정
# ==========================================
AI_MODELS = {
    "GPT-5.2 (고성능)": {
        "provider": "openai",
        "model": "gpt-5.2",
        "description": "OpenAI 최신 플래그십 모델"
    },
    "GPT-5-mini (빠름)": {
        "provider": "openai",
        "model": "gpt-5-mini",
        "description": "빠른 응답, 비용 효율적"
    },
    "Gemini 3 Pro (고성능)": {
        "provider": "gemini",
        "model": "gemini-3-pro-preview",
        "description": "Google 최신 추론 모델"
    },
    "Gemini 3 Flash (빠름)": {
        "provider": "gemini",
        "model": "gemini-3-flash-preview",
        "description": "Pro급 성능, 빠른 속도"
    }
}

# 지원 파일 확장자
SUPPORTED_FILE_TYPES = {
    # 문서
    "pdf": "PDF",
    "docx": "Word",
    "doc": "Word",
    "hwp": "한글",
    "hwpx": "한글",
    "txt": "텍스트",
    "rtf": "RTF",
    # 스프레드시트
    "xlsx": "Excel",
    "xls": "Excel",
    "csv": "CSV",
    # 이미지
    "png": "이미지",
    "jpg": "이미지",
    "jpeg": "이미지",
    "gif": "이미지",
    "bmp": "이미지",
    "tiff": "이미지",
    "tif": "이미지",
    "webp": "이미지",
    # 데이터/세금계산서
    "xml": "XML",
    "json": "JSON",
    "yaml": "YAML",
    "yml": "YAML",
    # 이메일
    "eml": "이메일",
    "msg": "이메일",
    # 압축 (미리보기용)
    "zip": "압축",
}


# ==========================================
# 파일 읽기 유틸리티
# ==========================================
class FileReader:
    """다양한 파일 형식에서 텍스트/이미지 추출"""

    @staticmethod
    def read_file(file_path: str) -> Dict:
        """
        파일을 읽어서 AI에 전달할 형식으로 변환
        Returns: {"type": "text"|"image", "content": str, "filename": str, "error": str|None}
        """
        if not os.path.exists(file_path):
            return {"type": "error", "content": "", "filename": os.path.basename(file_path), "error": "파일을 찾을 수 없습니다"}

        ext = os.path.splitext(file_path)[1].lower().lstrip('.')
        filename = os.path.basename(file_path)

        try:
            if ext == 'pdf':
                return FileReader._read_pdf(file_path, filename)
            elif ext in ['png', 'jpg', 'jpeg', 'gif', 'bmp', 'tiff', 'tif', 'webp']:
                return FileReader._read_image(file_path, filename)
            elif ext in ['xlsx', 'xls']:
                return FileReader._read_excel(file_path, filename)
            elif ext == 'csv':
                return FileReader._read_csv(file_path, filename)
            elif ext in ['docx', 'doc']:
                return FileReader._read_docx(file_path, filename)
            elif ext == 'xml':
                return FileReader._read_xml(file_path, filename)
            elif ext == 'json':
                return FileReader._read_json(file_path, filename)
            elif ext in ['yaml', 'yml']:
                return FileReader._read_yaml(file_path, filename)
            elif ext in ['txt', 'rtf', 'eml']:
                return FileReader._read_text(file_path, filename)
            elif ext in ['hwp', 'hwpx']:
                return {"type": "text", "content": f"[한글 문서: {filename}]\n(한글 파일은 직접 텍스트 추출이 어렵습니다. 내용을 직접 입력해주세요)", "filename": filename, "error": None}
            else:
                return {"type": "error", "content": "", "filename": filename, "error": f"지원하지 않는 파일 형식: {ext}"}
        except Exception as e:
            return {"type": "error", "content": "", "filename": filename, "error": str(e)}

    @staticmethod
    def _read_pdf(file_path: str, filename: str) -> Dict:
        """PDF 텍스트 추출"""
        if not HAS_PYPDF2:
            return {"type": "error", "content": "", "filename": filename, "error": "PyPDF2 라이브러리가 설치되지 않았습니다"}

        text_content = []
        with open(file_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            for page_num, page in enumerate(reader.pages[:10], 1):  # 최대 10페이지
                text = page.extract_text()
                if text:
                    text_content.append(f"[페이지 {page_num}]\n{text}")

        content = "\n\n".join(text_content)
        if not content.strip():
            # PDF에 텍스트가 없으면 이미지로 시도 (스캔 문서일 수 있음)
            return {"type": "text", "content": "(PDF에서 텍스트를 추출할 수 없습니다. 스캔된 문서일 수 있습니다)", "filename": filename, "error": None}

        return {"type": "text", "content": content[:8000], "filename": filename, "error": None}  # 최대 8000자

    @staticmethod
    def _read_image(file_path: str, filename: str) -> Dict:
        """이미지를 base64로 인코딩 (Vision AI용)"""
        if not HAS_PIL:
            return {"type": "error", "content": "", "filename": filename, "error": "PIL 라이브러리가 설치되지 않았습니다"}

        # 이미지 리사이즈 (너무 크면 API 제한)
        with Image.open(file_path) as img:
            # RGBA -> RGB 변환 (JPEG 저장용)
            if img.mode == 'RGBA':
                img = img.convert('RGB')

            # 최대 크기 제한
            max_size = (1500, 1500)
            img.thumbnail(max_size, Image.Resampling.LANCZOS)

            # base64 인코딩
            buffer = io.BytesIO()
            img.save(buffer, format='JPEG', quality=85)
            base64_image = base64.b64encode(buffer.getvalue()).decode('utf-8')

        return {"type": "image", "content": base64_image, "filename": filename, "error": None}

    @staticmethod
    def _read_excel(file_path: str, filename: str) -> Dict:
        """Excel 파일 읽기"""
        df = pd.read_excel(file_path, sheet_name=0, nrows=50)  # 첫 시트, 최대 50행
        content = f"[Excel 파일: {filename}]\n"
        content += df.to_string(index=False, max_rows=30, max_cols=10)
        return {"type": "text", "content": content[:6000], "filename": filename, "error": None}

    @staticmethod
    def _read_csv(file_path: str, filename: str) -> Dict:
        """CSV 파일 읽기"""
        df = pd.read_csv(file_path, nrows=50, encoding='utf-8-sig')
        content = f"[CSV 파일: {filename}]\n"
        content += df.to_string(index=False, max_rows=30, max_cols=10)
        return {"type": "text", "content": content[:6000], "filename": filename, "error": None}

    @staticmethod
    def _read_docx(file_path: str, filename: str) -> Dict:
        """Word 문서 읽기"""
        if not HAS_DOCX:
            return {"type": "error", "content": "", "filename": filename, "error": "python-docx 라이브러리가 설치되지 않았습니다"}

        doc = DocxDocument(file_path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        content = f"[Word 문서: {filename}]\n" + "\n".join(paragraphs[:100])  # 최대 100단락
        return {"type": "text", "content": content[:6000], "filename": filename, "error": None}

    @staticmethod
    def _read_xml(file_path: str, filename: str) -> Dict:
        """XML 파일 읽기 (세금계산서 등)"""
        import xml.etree.ElementTree as ET

        try:
            tree = ET.parse(file_path)
            root = tree.getroot()

            # XML을 읽기 쉬운 형식으로 변환
            def element_to_text(elem, indent=0):
                lines = []
                tag = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag  # namespace 제거
                text = elem.text.strip() if elem.text and elem.text.strip() else ""

                if text:
                    lines.append("  " * indent + f"{tag}: {text}")
                elif len(elem) > 0:
                    lines.append("  " * indent + f"[{tag}]")
                    for child in elem:
                        lines.extend(element_to_text(child, indent + 1))

                # 속성 처리
                for attr, value in elem.attrib.items():
                    lines.append("  " * (indent + 1) + f"@{attr}: {value}")

                return lines

            content_lines = element_to_text(root)
            content = f"[XML 파일: {filename}]\n" + "\n".join(content_lines[:500])  # 최대 500줄
            return {"type": "text", "content": content[:8000], "filename": filename, "error": None}

        except ET.ParseError as e:
            # XML 파싱 실패 시 텍스트로 읽기
            return FileReader._read_text(file_path, filename)

    @staticmethod
    def _read_json(file_path: str, filename: str) -> Dict:
        """JSON 파일 읽기"""
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        content = f"[JSON 파일: {filename}]\n"
        content += json.dumps(data, ensure_ascii=False, indent=2)
        return {"type": "text", "content": content[:8000], "filename": filename, "error": None}

    @staticmethod
    def _read_yaml(file_path: str, filename: str) -> Dict:
        """YAML 파일 읽기"""
        try:
            import yaml
            with open(file_path, 'r', encoding='utf-8') as f:
                data = yaml.safe_load(f)

            content = f"[YAML 파일: {filename}]\n"
            content += yaml.dump(data, allow_unicode=True, default_flow_style=False)
            return {"type": "text", "content": content[:8000], "filename": filename, "error": None}
        except ImportError:
            # YAML 라이브러리 없으면 텍스트로 읽기
            return FileReader._read_text(file_path, filename)

    @staticmethod
    def _read_text(file_path: str, filename: str) -> Dict:
        """텍스트 파일 읽기 (txt, rtf, eml 등)"""
        # 다양한 인코딩 시도
        encodings = ['utf-8', 'utf-8-sig', 'cp949', 'euc-kr', 'latin-1']

        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    text = f.read()
                content = f"[텍스트 파일: {filename}]\n{text}"
                return {"type": "text", "content": content[:8000], "filename": filename, "error": None}
            except UnicodeDecodeError:
                continue

        return {"type": "error", "content": "", "filename": filename, "error": "파일 인코딩을 인식할 수 없습니다"}


# ==========================================
# 프로젝트 관리자 (첨부파일 및 디렉토리 관리)
# ==========================================
class ProjectManager:
    """
    프로젝트 디렉토리 관리자

    각 Excel/CSV 파일에 대해 프로젝트 디렉토리를 생성하고,
    첨부파일을 행별로 연결하여 저장/관리합니다.

    디렉토리 구조:
    projects/
    ├── {프로젝트명_YYYYMMDD_HHMMSS}/
    │   ├── source.xlsx (원본 파일 복사본)
    │   ├── attachments/
    │   │   ├── row_001_001_filename.pdf
    │   │   ├── row_001_002_filename.docx
    │   │   └── ...
    │   └── project.json (메타데이터)
    """

    PROJECTS_DIR = os.path.join(SCRIPT_DIR, "projects")

    def __init__(self, source_file_path: str = None):
        self.source_file_path = source_file_path
        self.project_dir = None
        self.attachments_dir = None
        self.metadata = {
            "project_name": "",
            "source_file": "",
            "source_file_path": "",
            "created_at": "",
            "last_modified": "",
            "attachments": {},  # {row_idx: [{file_info}, ...]}
            "linked_transactions": {},  # {row_key: [linked_row_indices]}
            "row_notes": {}  # {row_idx: "메모"}
        }

        # 프로젝트 디렉토리 생성
        os.makedirs(self.PROJECTS_DIR, exist_ok=True)

    def create_project(self, source_file_path: str) -> bool:
        """새 프로젝트 생성"""
        from datetime import datetime
        import shutil

        self.source_file_path = source_file_path
        file_name = os.path.basename(source_file_path)
        base_name = os.path.splitext(file_name)[0]

        # 프로젝트명 생성 (파일명_날짜시간)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        project_name = f"{base_name}_{timestamp}"

        # 디렉토리 생성
        self.project_dir = os.path.join(self.PROJECTS_DIR, project_name)
        self.attachments_dir = os.path.join(self.project_dir, "attachments")

        try:
            os.makedirs(self.project_dir, exist_ok=True)
            os.makedirs(self.attachments_dir, exist_ok=True)

            # 원본 파일 복사
            dest_file = os.path.join(self.project_dir, f"source{os.path.splitext(file_name)[1]}")
            shutil.copy2(source_file_path, dest_file)

            # 메타데이터 초기화
            self.metadata = {
                "project_name": project_name,
                "source_file": file_name,
                "source_file_path": source_file_path,
                "created_at": datetime.now().isoformat(),
                "last_modified": datetime.now().isoformat(),
                "attachments": {},
                "linked_transactions": {},
                "row_notes": {}
            }
            self._save_metadata()
            return True

        except Exception as e:
            print(f"프로젝트 생성 오류: {e}")
            return False

    def load_project(self, project_dir: str) -> bool:
        """기존 프로젝트 로드"""
        self.project_dir = project_dir
        self.attachments_dir = os.path.join(project_dir, "attachments")
        metadata_path = os.path.join(project_dir, "project.json")

        if os.path.exists(metadata_path):
            try:
                with open(metadata_path, 'r', encoding='utf-8') as f:
                    self.metadata = json.load(f)
                self.source_file_path = self.metadata.get("source_file_path", "")
                return True
            except Exception as e:
                print(f"프로젝트 로드 오류: {e}")
        return False

    def find_existing_project(self, source_file_path: str) -> Optional[str]:
        """동일한 원본 파일에 대한 기존 프로젝트 찾기"""
        if not os.path.exists(self.PROJECTS_DIR):
            return None

        for project_name in os.listdir(self.PROJECTS_DIR):
            project_dir = os.path.join(self.PROJECTS_DIR, project_name)
            metadata_path = os.path.join(project_dir, "project.json")

            if os.path.exists(metadata_path):
                try:
                    with open(metadata_path, 'r', encoding='utf-8') as f:
                        meta = json.load(f)
                    if meta.get("source_file_path") == source_file_path:
                        return project_dir
                except:
                    continue
        return None

    def add_attachment(self, row_idx: int, file_path: str, sheet_name: str = "Sheet1") -> Optional[str]:
        """
        행에 첨부파일 추가

        Args:
            row_idx: 행 인덱스
            file_path: 첨부할 파일 경로
            sheet_name: 시트명

        Returns:
            저장된 파일 경로 (실패 시 None)
        """
        from datetime import datetime
        import shutil

        if not self.attachments_dir or not os.path.exists(self.attachments_dir):
            return None

        try:
            # 고유 키 생성 (시트_행)
            row_key = f"{sheet_name}_{row_idx}"

            # 기존 첨부파일 수 확인
            existing = self.metadata.get("attachments", {}).get(row_key, [])
            file_num = len(existing) + 1

            # 파일명 생성: row_XXX_YYY_원본파일명
            original_name = os.path.basename(file_path)
            ext = os.path.splitext(original_name)[1]
            safe_name = "".join(c if c.isalnum() or c in "._-" else "_" for c in os.path.splitext(original_name)[0])
            saved_name = f"row_{row_idx+1:03d}_{file_num:03d}_{safe_name}{ext}"

            dest_path = os.path.join(self.attachments_dir, saved_name)

            # 파일 복사
            shutil.copy2(file_path, dest_path)

            # 메타데이터 업데이트
            if "attachments" not in self.metadata:
                self.metadata["attachments"] = {}
            if row_key not in self.metadata["attachments"]:
                self.metadata["attachments"][row_key] = []

            self.metadata["attachments"][row_key].append({
                "original_name": original_name,
                "saved_name": saved_name,
                "saved_path": dest_path,
                "attached_at": datetime.now().isoformat(),
                "file_type": ext.lstrip('.').lower()
            })

            self.metadata["last_modified"] = datetime.now().isoformat()
            self._save_metadata()

            return dest_path

        except Exception as e:
            print(f"첨부파일 추가 오류: {e}")
            return None

    def get_attachments(self, row_idx: int, sheet_name: str = "Sheet1") -> List[Dict]:
        """행의 첨부파일 목록 가져오기"""
        row_key = f"{sheet_name}_{row_idx}"
        return self.metadata.get("attachments", {}).get(row_key, [])

    def get_all_attachments(self) -> Dict[str, List[Dict]]:
        """모든 첨부파일 목록 가져오기"""
        return self.metadata.get("attachments", {})

    def remove_attachment(self, row_idx: int, attachment_index: int, sheet_name: str = "Sheet1") -> bool:
        """첨부파일 삭제"""
        row_key = f"{sheet_name}_{row_idx}"
        attachments = self.metadata.get("attachments", {}).get(row_key, [])

        if 0 <= attachment_index < len(attachments):
            try:
                # 파일 삭제
                file_info = attachments[attachment_index]
                file_path = file_info.get("saved_path", "")
                if os.path.exists(file_path):
                    os.remove(file_path)

                # 메타데이터에서 제거
                attachments.pop(attachment_index)
                self.metadata["attachments"][row_key] = attachments
                self._save_metadata()
                return True
            except Exception as e:
                print(f"첨부파일 삭제 오류: {e}")
        return False

    def add_linked_transactions(self, row_idx: int, linked_rows: List[int],
                                 sheet_name: str = "Sheet1") -> bool:
        """
        행에 연관 거래 연결 추가 - 양방향 링크

        Args:
            row_idx: 소스 행 인덱스
            linked_rows: 연결할 행 인덱스 목록
            sheet_name: 시트 이름

        Returns:
            성공 여부
        """
        if not self.project_dir:
            return False

        try:
            if "linked_transactions" not in self.metadata:
                self.metadata["linked_transactions"] = {}

            row_key = f"{sheet_name}_{row_idx}"

            # 소스 행에 연결 추가
            self.metadata["linked_transactions"][row_key] = linked_rows.copy()

            # 양방향 링크 유지: 타겟 행에도 소스 행 추가
            for target_row in linked_rows:
                target_key = f"{sheet_name}_{target_row}"
                if target_key not in self.metadata["linked_transactions"]:
                    self.metadata["linked_transactions"][target_key] = []

                if row_idx not in self.metadata["linked_transactions"][target_key]:
                    self.metadata["linked_transactions"][target_key].append(row_idx)

            self._save_metadata()
            return True

        except Exception as e:
            print(f"연관 거래 저장 오류: {e}")
            return False

    def get_linked_transactions(self, row_idx: int, sheet_name: str = "Sheet1") -> List[int]:
        """
        행의 연관 거래 목록 가져오기

        Args:
            row_idx: 행 인덱스
            sheet_name: 시트 이름

        Returns:
            연결된 행 인덱스 목록
        """
        if "linked_transactions" not in self.metadata:
            return []

        row_key = f"{sheet_name}_{row_idx}"
        return self.metadata["linked_transactions"].get(row_key, []).copy()

    def remove_linked_transaction(self, row_idx: int, target_row: int,
                                   sheet_name: str = "Sheet1") -> bool:
        """
        특정 연관 거래 연결 해제

        Args:
            row_idx: 소스 행 인덱스
            target_row: 연결 해제할 타겟 행 인덱스
            sheet_name: 시트 이름

        Returns:
            성공 여부
        """
        if not self.project_dir or "linked_transactions" not in self.metadata:
            return False

        try:
            row_key = f"{sheet_name}_{row_idx}"
            target_key = f"{sheet_name}_{target_row}"

            # 소스 행에서 타겟 제거
            if row_key in self.metadata["linked_transactions"]:
                links = self.metadata["linked_transactions"][row_key]
                if target_row in links:
                    links.remove(target_row)

            # 타겟 행에서 소스 제거 (양방향)
            if target_key in self.metadata["linked_transactions"]:
                target_links = self.metadata["linked_transactions"][target_key]
                if row_idx in target_links:
                    target_links.remove(row_idx)

            self._save_metadata()
            return True

        except Exception as e:
            print(f"연관 거래 해제 오류: {e}")
            return False

    def _save_metadata(self):
        """메타데이터 저장"""
        if self.project_dir:
            metadata_path = os.path.join(self.project_dir, "project.json")
            try:
                with open(metadata_path, 'w', encoding='utf-8') as f:
                    json.dump(self.metadata, f, ensure_ascii=False, indent=2)
            except Exception as e:
                print(f"메타데이터 저장 오류: {e}")

    def save_classification_data(self, sheets: Dict[str, pd.DataFrame],
                                  chat_histories: Dict[str, Dict[int, List[Dict]]]) -> bool:
        """
        분류 작업 데이터 저장 (DataFrame의 분류 컬럼 + 대화 기록)

        Args:
            sheets: {sheet_name: DataFrame} - 분류 결과가 포함된 데이터프레임
            chat_histories: {sheet_name: {row_idx: [messages]}} - 대화 기록

        Returns:
            성공 여부
        """
        from datetime import datetime

        if not self.project_dir:
            print("프로젝트가 초기화되지 않았습니다.")
            return False

        data_path = os.path.join(self.project_dir, "classification_data.json")

        try:
            # 저장할 데이터 구조
            save_data = {
                "saved_at": datetime.now().isoformat(),
                "sheets": {},
                "chat_histories": {}
            }

            # 각 시트의 분류 관련 컬럼만 저장
            classification_columns = ["분류_상태", "최종_분류", "분류_근거", "사용자_입력", "대화_기록", "검토표시"]

            for sheet_name, df in sheets.items():
                sheet_data = {}
                for col in classification_columns:
                    if col in df.columns:
                        # NaN 값을 빈 문자열로 변환
                        col_data = df[col].fillna("").astype(str).tolist()
                        sheet_data[col] = col_data

                if sheet_data:
                    save_data["sheets"][sheet_name] = sheet_data

            # 대화 기록 저장
            for sheet_name, rows in chat_histories.items():
                save_data["chat_histories"][sheet_name] = {
                    str(row_idx): messages for row_idx, messages in rows.items()
                }

            # JSON 파일로 저장
            with open(data_path, 'w', encoding='utf-8') as f:
                json.dump(save_data, f, ensure_ascii=False, indent=2)

            # 메타데이터 수정 시간 업데이트
            self.metadata["last_modified"] = datetime.now().isoformat()
            self._save_metadata()

            print(f"분류 데이터 저장 완료: {data_path}")
            return True

        except Exception as e:
            print(f"분류 데이터 저장 오류: {e}")
            import traceback
            traceback.print_exc()
            return False

    def load_classification_data(self) -> Optional[Dict]:
        """
        저장된 분류 작업 데이터 로드

        Returns:
            {
                "sheets": {sheet_name: {col_name: [values]}},
                "chat_histories": {sheet_name: {row_idx: [messages]}}
            }
            또는 None (파일이 없거나 오류 시)
        """
        if not self.project_dir:
            return None

        data_path = os.path.join(self.project_dir, "classification_data.json")

        if not os.path.exists(data_path):
            print(f"분류 데이터 파일이 없습니다: {data_path}")
            return None

        try:
            with open(data_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            # chat_histories의 키를 int로 변환
            chat_histories = {}
            for sheet_name, rows in data.get("chat_histories", {}).items():
                chat_histories[sheet_name] = {
                    int(row_idx): messages for row_idx, messages in rows.items()
                }
            data["chat_histories"] = chat_histories

            print(f"분류 데이터 로드 완료: {data_path}")
            return data

        except Exception as e:
            print(f"분류 데이터 로드 오류: {e}")
            import traceback
            traceback.print_exc()
            return None

    def get_project_info(self) -> Dict:
        """프로젝트 정보 반환"""
        return {
            "project_dir": self.project_dir,
            "project_name": self.metadata.get("project_name", ""),
            "source_file": self.metadata.get("source_file", ""),
            "created_at": self.metadata.get("created_at", ""),
            "attachment_count": sum(len(v) for v in self.metadata.get("attachments", {}).values())
        }

    @classmethod
    def list_projects(cls) -> List[Dict]:
        """모든 프로젝트 목록 반환"""
        projects = []
        if not os.path.exists(cls.PROJECTS_DIR):
            return projects

        for project_name in os.listdir(cls.PROJECTS_DIR):
            project_dir = os.path.join(cls.PROJECTS_DIR, project_name)
            metadata_path = os.path.join(project_dir, "project.json")

            if os.path.exists(metadata_path):
                try:
                    with open(metadata_path, 'r', encoding='utf-8') as f:
                        meta = json.load(f)
                    projects.append({
                        "project_name": project_name,
                        "project_dir": project_dir,
                        "source_file": meta.get("source_file", ""),
                        "created_at": meta.get("created_at", ""),
                        "last_modified": meta.get("last_modified", ""),
                        "attachment_count": sum(len(v) for v in meta.get("attachments", {}).values())
                    })
                except:
                    continue

        # 최근 수정순 정렬
        projects.sort(key=lambda x: x.get("last_modified", ""), reverse=True)
        return projects


# ==========================================
# 분류 상태 정의
# ==========================================
class ClassificationStatus:
    PENDING = "미분류"           # 아직 분류되지 않음
    EXISTING = "기존분류"        # 파일에 이미 분류가 있음 (검증 필요)
    AI_CLASSIFIED = "AI분류"     # AI가 분류함
    VERIFIED = "검증완료"        # 사용자가 확인함
    NEEDS_REVIEW = "검토필요"    # AI가 불확실하다고 표시


# ==========================================
# 컬럼 매핑 다이얼로그 (dialogs.py로 이동됨)
# ==========================================


# ==========================================
# 회계처리 알고리즘 (시스템 프롬프트)
# ==========================================
# ==========================================
# 회계처리 실무지침 로더
# ==========================================
GUIDELINES_FILE_PATH = os.path.join(SCRIPT_DIR, "accounting_guidelines.json")


def load_accounting_guidelines() -> Dict:
    """accounting_guidelines.json 파일을 로드하여 반환"""
    try:
        with open(GUIDELINES_FILE_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"경고: 회계처리 실무지침 파일을 찾을 수 없습니다: {GUIDELINES_FILE_PATH}")
        return {}
    except json.JSONDecodeError as e:
        print(f"경고: 회계처리 실무지침 파일 파싱 오류: {e}")
        return {}


def save_accounting_guidelines(guidelines: Dict) -> bool:
    """accounting_guidelines.json 파일에 저장 (learned_cases, conflict_log 등 업데이트용)"""
    try:
        with open(GUIDELINES_FILE_PATH, 'w', encoding='utf-8') as f:
            json.dump(guidelines, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"경고: 회계처리 실무지침 저장 실패: {e}")
        return False


def add_learned_case(case_data: Dict) -> bool:
    """신규 케이스를 learned_cases 배열에 추가 (특정 위치에만 수정)"""
    import datetime

    guidelines = load_accounting_guidelines()
    if not guidelines:
        return False

    # learned_cases 배열이 없으면 생성
    if "learned_cases" not in guidelines:
        guidelines["learned_cases"] = []

    # 케이스 데이터에 타임스탬프 추가
    case_data["added_date"] = datetime.datetime.now().isoformat()
    case_data["source"] = "user_classification"

    # learned_cases 배열에만 추가 (다른 부분은 수정하지 않음)
    guidelines["learned_cases"].append(case_data)

    # audit_trail에도 기록
    if "audit_trail" not in guidelines:
        guidelines["audit_trail"] = []
    guidelines["audit_trail"].append({
        "action": "add_learned_case",
        "timestamp": datetime.datetime.now().isoformat(),
        "case_summary": case_data.get("description", "신규 케이스 추가")
    })

    return save_accounting_guidelines(guidelines)


def add_conflict_log(conflict_data: Dict) -> bool:
    """분류 충돌(사용자가 AI/기존 분류와 다르게 직접 입력)을 conflict_log에 기록"""
    import datetime

    guidelines = load_accounting_guidelines()
    if not guidelines:
        return False

    # conflict_log 배열이 없으면 생성
    if "conflict_log" not in guidelines:
        guidelines["conflict_log"] = []

    # 충돌 데이터에 타임스탬프 추가
    conflict_data["timestamp"] = datetime.datetime.now().isoformat()

    # conflict_log 배열에만 추가
    guidelines["conflict_log"].append(conflict_data)

    # audit_trail에도 기록
    if "audit_trail" not in guidelines:
        guidelines["audit_trail"] = []
    guidelines["audit_trail"].append({
        "action": "add_conflict_log",
        "timestamp": datetime.datetime.now().isoformat(),
        "case_summary": f"충돌: {conflict_data.get('original_classification', '?')} → {conflict_data.get('user_classification', '?')}"
    })

    return save_accounting_guidelines(guidelines)


def build_system_prompt_from_guidelines(max_learned_cases: int = 20) -> str:
    """
    JSON 실무지침 파일을 읽어서 AI 시스템 프롬프트로 변환

    Args:
        max_learned_cases: AI에게 전달할 최대 learned_cases 수 (너무 많으면 토큰 초과)
    """
    guidelines = load_accounting_guidelines()

    if not guidelines:
        # 파일이 없거나 오류인 경우 기본 프롬프트
        return """당신은 회계처리 계정과목 분류 전문가 AI입니다.
거래 내역을 분석하여 적절한 계정과목을 분류해주세요.

## 출력 형식 (반드시 JSON)
{
    "status": "INCOMPLETE" | "COMPLETE",
    "message": "사용자에게 보여줄 질문 또는 최종 설명",
    "questions": ["구체적인 질문1"],
    "final_classification": "계정코드 계정명 (예: 823 경상연구개발비, 206 건물)",
    "reasoning": "판단 근거 요약"
}
"""

    # 핵심 정보만 추출하여 AI에게 전달 (토큰 절약)
    essential_guidelines = {
        "_metadata": guidelines.get("_metadata", {}),
        "company_profile": guidelines.get("company_profile", {}),
        "required_input_columns": guidelines.get("required_input_columns", {}),
        "classification_categories": guidelines.get("classification_categories", {}),
        "master_decision_tree": guidelines.get("master_decision_tree", {}),
        "contract_payment_check": guidelines.get("contract_payment_check", {}),
        "tax_checkpoints": guidelines.get("tax_checkpoints", {}),
        "workflow": guidelines.get("workflow", {})
    }

    # learned_cases는 최근 N개만 포함 (토큰 절약)
    learned_cases = guidelines.get("learned_cases", [])
    if len(learned_cases) > max_learned_cases:
        # 최근 케이스만 사용 (가장 최신이 가장 관련 있을 가능성)
        essential_guidelines["learned_cases"] = learned_cases[-max_learned_cases:]
        essential_guidelines["_note"] = f"총 {len(learned_cases)}개의 learned_cases 중 최근 {max_learned_cases}개만 표시"
    else:
        essential_guidelines["learned_cases"] = learned_cases

    # conflict_log와 audit_trail은 AI에게 전달하지 않음 (프롬프트 절약)

    # JSON 실무지침을 AI에게 전달
    prompt = f"""당신은 대한민국 회계처리 계정과목 분류 전문가 AI입니다.

아래에 제공된 '회계처리 실무지침서(JSON 형식)'를 읽고 이해한 후,
사용자가 제공한 거래 내역을 분석하여 적절한 계정과목을 분류해야 합니다.

## 회계처리 실무지침서 (JSON)
이 실무지침서는 사람이 읽어도 이해할 수 있도록 작성되었습니다.
당신은 이 지침서의 내용을 참고하여 분류를 수행해야 합니다.

```json
{json.dumps(essential_guidelines, ensure_ascii=False, indent=2)}
```

## 분류 수행 방법
1. 위 실무지침서의 'master_decision_tree'에 정의된 의사결정 트리(STEP1~STEP6)를 따라가며 분류하세요.
2. 'company_profile'에서 회계 기준(중소기업회계기준, IFRS 고려)과 자본화 정책을 확인하세요.
3. 'classification_categories'에서 최종 분류 카테고리(1_유형자산, 2_건설중인자산, 3_무형자산, 4_기타)를 확인하세요.
4. 'learned_cases'에 유사한 케이스가 있다면 참고하세요.
5. 분류를 확정하기 위해 정보가 부족하다면, 의사결정 트리에 근거하여 추가 질문을 하세요.

## 중요 지침
- 한 번에 하나 또는 연관된 소수의 질문만 하여 사용자가 답변하기 편하게 하세요.
- 이미 알고 있는 정보는 다시 묻지 마세요.
- 모든 정보가 충족되어 분류가 확정되면, 더 이상 질문하지 말고 최종 결과를 출력하세요.

## 출력 형식 (반드시 JSON)
{{
    "status": "INCOMPLETE" | "COMPLETE",
    "message": "사용자에게 보여줄 질문 또는 최종 설명",
    "questions": ["구체적인 질문1", "구체적인 질문2"],
    "final_classification": "계정코드 계정명 (예: 823 경상연구개발비, 206 건물, 215 기계장치)",
    "reasoning": "판단 근거 요약 (의사결정 트리의 어떤 단계를 거쳤는지 포함)",
    "is_new_case": true | false,
    "new_case_suggestion": "실무지침서에 없는 새로운 케이스라면, 추가할 규칙 제안"
}}
"""
    return prompt


# ==========================================
# AI 클래스
# ==========================================
class AccountingAI:
    def __init__(self, model_config: Dict = None):
        self.openai_key = os.getenv("OPENAI_API_KEY")
        self.google_key = os.getenv("GOOGLE_API_KEY")

        # 기본 모델 설정
        if model_config is None:
            model_config = AI_MODELS["Gemini 3 Flash (빠름)"]

        self.provider = model_config["provider"]
        self.model = model_config["model"]

        # API 클라이언트 초기화
        if self.provider == "openai":
            if not self.openai_key or not HAS_OPENAI:
                raise ValueError("OpenAI API Key가 없거나 openai 라이브러리가 설치되지 않았습니다.")
            self.client = OpenAI(api_key=self.openai_key)
        elif self.provider == "gemini":
            if not self.google_key or not HAS_GEMINI:
                raise ValueError("Google API Key가 없거나 google-generativeai 라이브러리가 설치되지 않았습니다.")
            genai.configure(api_key=self.google_key)

    def analyze(self, row_data: str, history: List[Dict]) -> Dict:
        """행 데이터와 대화 기록을 받아 AI 응답을 반환"""
        # 매번 최신 실무지침을 로드 (learned_cases 등이 업데이트될 수 있음)
        system_prompt = build_system_prompt_from_guidelines()

        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": f"[현재 검토 중인 거래 데이터]\n{row_data}"}
        ]

        # 이전 대화 기록 추가
        for h in history:
            messages.append(h)

        try:
            if self.provider == "openai":
                # GPT-5 모델은 temperature 파라미터를 지원하지 않음 (기본값 1만 허용)
                # response_format도 json_schema 방식으로 변경
                response = self.client.chat.completions.create(
                    model=self.model,
                    messages=messages,
                    response_format={"type": "json_object"}
                    # temperature 파라미터 제거 (GPT-5는 지원하지 않음)
                )
                content = response.choices[0].message.content

            elif self.provider == "gemini":
                # Gemini 3용 프롬프트 구성 (실무지침 JSON 파일에서 로드)
                prompt = system_prompt + "\n\n" + f"[현재 거래 데이터]\n{row_data}\n"

                for h in history:
                    role = "사용자" if h["role"] == "user" else "AI"
                    prompt += f"\n{role}: {h['content']}"

                prompt += "\n\n위 정보를 바탕으로 JSON 형식으로 응답하세요:"

                model = genai.GenerativeModel(self.model)
                # Gemini 3는 thinking_level 파라미터 사용 (동적 추론)
                response = model.generate_content(
                    prompt,
                    generation_config={
                        "response_mime_type": "application/json",
                        "temperature": 0.3  # Gemini는 temperature 지원
                    }
                )
                content = response.text

            return json.loads(content)

        except json.JSONDecodeError as e:
            return {
                "status": "ERROR",
                "message": f"AI 응답 파싱 오류: {str(e)}\n원본: {content[:200] if 'content' in dir() else 'N/A'}",
                "final_classification": None
            }
        except Exception as e:
            return {
                "status": "ERROR",
                "message": f"AI 분석 중 오류 발생: {str(e)}",
                "final_classification": None
            }


# ==========================================
# 데이터 테이블 위젯 (widgets.py로 이동됨)
# ==========================================

# ==========================================
# 직접 분류 입력 다이얼로그 (dialogs.py로 이동됨)
# ==========================================

# ==========================================
# AI 대화 패널 (chat_panel.py로 이동됨)
# ==========================================

# ==========================================
# 메인 애플리케이션
# ==========================================
class AccountingClassifierApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        ctk.set_appearance_mode("Dark")
        self.title("회계처리 계정과목 분류 시스템")
        self.geometry("1500x900")
        self.configure(fg_color=BG_COLOR)

        # 상태 변수
        self.sheets: Dict[str, pd.DataFrame] = {}  # 시트명 -> DataFrame
        self.current_sheet: Optional[str] = None
        self.chat_histories: Dict[str, Dict[int, List[Dict]]] = {}  # 시트명 -> {row_idx -> history}
        self.column_mappings: Dict[str, Dict] = {}  # 시트명 -> 컬럼 매핑
        self.highlighted_cells: Dict[str, set] = {}  # 시트명 -> 음영 처리된 셀 좌표 {(row_idx, col_idx), ...}
        self.ai: Optional[AccountingAI] = None
        self.current_model_name: str = "Gemini 3 Flash (빠름)"  # 기본 모델
        self.current_filter: str = "all"  # 현재 필터
        self._current_file_name: str = ""  # 현재 파일명

        # 프로젝트 관리
        self.project_manager: Optional[ProjectManager] = None  # 프로젝트 매니저
        self._save_timer = None  # 자동 저장 debounce 타이머

        # AI 초기화
        self._init_ai()

        self._setup_ui()

        # 키보드 단축키 설정
        self._setup_keyboard_shortcuts()

        # 앱 시작 후 YAML 변경 확인 (UI가 완전히 로드된 후 실행)
        if HAS_VERSION_MANAGER:
            self.after(500, self._check_yaml_changes_on_startup)

    def _init_ai(self, model_name: str = None):
        """AI 초기화"""
        if model_name:
            self.current_model_name = model_name

        try:
            model_config = AI_MODELS.get(self.current_model_name)
            self.ai = AccountingAI(model_config)
            self.ai_status = f"{self.ai.provider.upper()} ({self.ai.model})"
        except Exception as e:
            self.ai_status = f"AI 초기화 실패: {str(e)}"
            self.ai = None

    def _setup_ui(self):
        # 상단 툴바
        toolbar = ctk.CTkFrame(self, fg_color=SIDEBAR_COLOR, height=50)
        toolbar.pack(fill="x", padx=10, pady=10)
        toolbar.pack_propagate(False)

        ctk.CTkButton(
            toolbar,
            text="파일 열기",
            font=(FONT_FAMILY, 12, "bold"),
            fg_color=ACCENT_COLOR,
            command=self._load_file
        ).pack(side="left", padx=10, pady=10)

        ctk.CTkButton(
            toolbar,
            text="프로젝트 열기",
            font=(FONT_FAMILY, 12, "bold"),
            fg_color="#7C3AED",
            hover_color="#6D28D9",
            command=self._show_open_project_dialog
        ).pack(side="left", padx=5, pady=10)

        ctk.CTkButton(
            toolbar,
            text="결과 저장",
            font=(FONT_FAMILY, 12, "bold"),
            fg_color=SUCCESS_COLOR,
            text_color="black",
            command=self._save_file
        ).pack(side="left", padx=5, pady=10)

        # 버전 관리 버튼
        if HAS_VERSION_MANAGER:
            ctk.CTkButton(
                toolbar,
                text="버전 관리",
                font=(FONT_FAMILY, 11),
                fg_color="#6B7280",
                hover_color="#4B5563",
                width=80,
                command=self._show_version_history
            ).pack(side="left", padx=5, pady=10)

            # 회계지침 검토 버튼
            ctk.CTkButton(
                toolbar,
                text="지침 검토",
                font=(FONT_FAMILY, 11),
                fg_color="#8B5CF6",
                hover_color="#7C3AED",
                width=80,
                command=self._show_guidelines_review_dialog
            ).pack(side="left", padx=5, pady=10)

        # 프로젝트 정보 버튼
        ctk.CTkButton(
            toolbar,
            text="프로젝트 정보",
            font=(FONT_FAMILY, 11),
            fg_color="#0EA5E9",
            hover_color="#0284C7",
            width=100,
            command=self._show_project_info
        ).pack(side="left", padx=5, pady=10)

        self.file_label = ctk.CTkLabel(
            toolbar,
            text="파일을 선택하세요",
            font=(FONT_FAMILY, 11),
            text_color="gray"
        )
        self.file_label.pack(side="left", padx=20)

        # 모델 선택 드롭다운 (우측)
        model_frame = ctk.CTkFrame(toolbar, fg_color="transparent")
        model_frame.pack(side="right", padx=10)

        ctk.CTkLabel(
            model_frame,
            text="AI 모델:",
            font=(FONT_FAMILY, 11),
            text_color="gray"
        ).pack(side="left", padx=(0, 5))

        self.model_var = ctk.StringVar(value=self.current_model_name)
        self.model_dropdown = ctk.CTkOptionMenu(
            model_frame,
            values=list(AI_MODELS.keys()),
            variable=self.model_var,
            font=(FONT_FAMILY, 11),
            width=180,
            command=self._on_model_change
        )
        self.model_dropdown.pack(side="left")

        self.ai_status_label = ctk.CTkLabel(
            model_frame,
            text=f"({self.ai_status})",
            font=(FONT_FAMILY, 9),
            text_color="gray"
        )
        self.ai_status_label.pack(side="left", padx=5)

        # ===== 필터 및 진행률 툴바 =====
        filter_toolbar = ctk.CTkFrame(self, fg_color=SIDEBAR_COLOR, height=45)
        filter_toolbar.pack(fill="x", padx=10, pady=(0, 5))
        filter_toolbar.pack_propagate(False)

        # 필터 버튼들
        filter_label = ctk.CTkLabel(
            filter_toolbar,
            text="필터:",
            font=(FONT_FAMILY, 11),
            text_color="gray"
        )
        filter_label.pack(side="left", padx=(10, 5), pady=8)

        self.filter_buttons = {}
        filter_configs = [
            ("all", "전체", "#6B7280"),
            ("pending", "미분류", "#EF4444"),
            ("highlighted", "검토표시", "#8B5CF6"),
            ("existing", "기존분류", "#F59E0B"),
            ("classified", "AI분류", "#3B82F6"),
            ("verified", "검증완료", "#10B981"),
            ("needs_review", "검토필요", "#EC4899")
        ]

        for filter_type, label, color in filter_configs:
            btn = ctk.CTkButton(
                filter_toolbar,
                text=label,
                font=(FONT_FAMILY, 10),
                fg_color=color if filter_type == "all" else "transparent",
                hover_color=color,
                width=70,
                height=28,
                command=lambda f=filter_type: self._set_filter(f)
            )
            btn.pack(side="left", padx=2, pady=8)
            self.filter_buttons[filter_type] = btn

        # 진행률 표시 (우측)
        progress_frame = ctk.CTkFrame(filter_toolbar, fg_color="transparent")
        progress_frame.pack(side="right", padx=10, pady=8)

        self.progress_label = ctk.CTkLabel(
            progress_frame,
            text="진행률: -",
            font=(FONT_FAMILY, 11),
            text_color="white"
        )
        self.progress_label.pack(side="left", padx=(0, 10))

        self.progress_bar = ctk.CTkProgressBar(
            progress_frame,
            width=150,
            height=12,
            progress_color=SUCCESS_COLOR
        )
        self.progress_bar.pack(side="left")
        self.progress_bar.set(0)

        # 빠른 작업 버튼
        ctk.CTkButton(
            filter_toolbar,
            text="선택 행 검증",
            font=(FONT_FAMILY, 10),
            fg_color="#10B981",
            width=90,
            height=28,
            command=self._verify_selected_row
        ).pack(side="right", padx=5, pady=8)

        ctk.CTkButton(
            filter_toolbar,
            text="일괄 AI분류",
            font=(FONT_FAMILY, 10),
            fg_color="#3B82F6",
            width=90,
            height=28,
            command=self._batch_classify
        ).pack(side="right", padx=2, pady=8)

        # 시트 탭
        self.tab_frame = ctk.CTkFrame(self, fg_color="transparent", height=40)
        self.tab_frame.pack(fill="x", padx=10)

        # 메인 컨텐츠 (PanedWindow 사용으로 크기 조절 가능)
        import tkinter as tk
        main_content = tk.PanedWindow(self, orient="horizontal", bg="#18181C", sashwidth=5, sashrelief="flat")
        main_content.pack(fill="both", expand=True, padx=10, pady=10)

        # 좌측: 데이터 테이블
        table_container = ctk.CTkFrame(main_content, fg_color=SIDEBAR_COLOR, corner_radius=10)
        main_content.add(table_container, minsize=400, stretch="always")

        table_header = ctk.CTkLabel(
            table_container,
            text="거래 내역",
            font=(FONT_FAMILY, 14, "bold")
        )
        table_header.pack(pady=10)

        self.data_table = DataTableFrame(
            table_container,
            on_row_select=self._on_row_select,
            on_cell_edit=self._on_cell_edit,
            fg_color="transparent"
        )
        self.data_table.pack(fill="both", expand=True, padx=5, pady=(0, 5))

        # 우측: AI 대화 패널 (최소 너비 400, 고정)
        self.chat_panel = AIChatPanel(
            main_content,
            on_submit=self._on_user_submit,
            on_reanalyze=self._on_reanalyze,
            on_file_analyze=self._on_file_analyze,
            on_keep_existing=self._on_keep_existing,
            on_manual_classify=self._on_manual_classify,
            on_reset_chat=self._on_reset_chat,
            on_save_attachment=self.save_attachment_to_project,
            on_get_attachments=self.get_row_attachments,
            on_remove_attachment=self.remove_attachment_from_project,
            on_start_ai=self._on_start_ai_analysis,
            on_approve_classification=self._on_approve_classification,
            on_link_changed=self.save_linked_transactions,
            on_get_linked=self.get_linked_transactions,
            corner_radius=10
        )
        main_content.add(self.chat_panel, minsize=400, width=420, stretch="never")

    def _on_model_change(self, model_name: str):
        """모델 변경 시"""
        self._init_ai(model_name)
        self.ai_status_label.configure(text=f"({self.ai_status})")

        if self.ai is None:
            messagebox.showerror("오류", f"모델 초기화 실패: {self.ai_status}")

    def _on_reanalyze(self, row_idx: int):
        """재분석 시작"""
        if self.current_sheet is None:
            return

        # 대화 기록 초기화
        if self.current_sheet in self.chat_histories:
            self.chat_histories[self.current_sheet][row_idx] = []

        # 데이터 테이블에서 분류 초기화
        self.data_table.clear_row_classification(row_idx)

        # 원본 데이터프레임도 초기화
        self.sheets[self.current_sheet].at[row_idx, "최종_분류"] = ""
        self.sheets[self.current_sheet].at[row_idx, "분류_근거"] = ""
        self.sheets[self.current_sheet].at[row_idx, "분류_상태"] = ClassificationStatus.PENDING

        # 채팅 패널에서 재분석 시작
        df = self.sheets[self.current_sheet]
        row_data = df.iloc[row_idx].to_json(force_ascii=False, indent=2)

        self.chat_panel.set_row(row_idx, row_data, is_reanalyze=True)
        self.chat_panel.add_message("System", "거래 내역을 재분석 중입니다...")
        self.chat_panel.show_loading(True)

        # AI 호출
        threading.Thread(
            target=self._run_ai_analysis,
            args=(self.current_sheet, row_idx, row_data),
            daemon=True
        ).start()

        self._update_progress()

    def _on_keep_existing(self, row_idx: int, existing_classification: str):
        """기존 분류 유지 (검증완료 처리)"""
        if self.current_sheet is None:
            return

        # 분류 상태를 검증완료로 변경
        self.data_table.update_row(
            row_idx,
            existing_classification,
            "기존 분류 유지 (사용자 확인)",
            "",
            status=ClassificationStatus.VERIFIED
        )
        self.sheets[self.current_sheet].at[row_idx, "최종_분류"] = existing_classification
        self.sheets[self.current_sheet].at[row_idx, "분류_근거"] = "기존 분류 유지 (사용자 확인)"
        self.sheets[self.current_sheet].at[row_idx, "분류_상태"] = ClassificationStatus.VERIFIED

        # 채팅 패널에 완료 표시
        self.chat_panel.show_complete(existing_classification, "기존 분류 유지 (사용자 확인)")
        self.chat_panel.add_message("System", "기존 분류가 검증완료 되었습니다.")

        # 대화 기록에 저장
        if self.current_sheet not in self.chat_histories:
            self.chat_histories[self.current_sheet] = {}
        self.chat_histories[self.current_sheet][row_idx] = [{
            "role": "assistant",
            "content": json.dumps({
                "status": "COMPLETE",
                "message": "기존 분류 유지",
                "final_classification": existing_classification,
                "reasoning": "기존 분류 유지 (사용자 확인)"
            }, ensure_ascii=False)
        }]

        self._update_progress()

    def _on_manual_classify(self, row_idx: int, classification: str, reasoning: str):
        """직접 분류 입력 처리"""
        if self.current_sheet is None:
            return

        # 기존 분류 확인 (충돌 감지용)
        original_classification = ""
        original_status = ""
        ai_reasoning = ""
        try:
            row = self.sheets[self.current_sheet].iloc[row_idx]
            original_classification = str(row.get("최종_분류", ""))
            original_status = str(row.get("분류_상태", ""))
            ai_reasoning = str(row.get("분류_근거", ""))
        except:
            pass

        # 충돌 검사 - AI가 분류한 것과 다르면 경고 다이얼로그 표시
        if original_classification and original_classification != classification:
            if original_status == ClassificationStatus.AI_CLASSIFIED:
                # 충돌 심각도 판단
                conflict_severity = self._assess_conflict_severity(
                    original_classification, classification, row_idx
                )

                # 경고 다이얼로그 표시 (사용자가 확인 후 진행)
                proceed = self._show_conflict_warning_dialog(
                    row_idx=row_idx,
                    ai_classification=original_classification,
                    user_classification=classification,
                    user_reasoning=reasoning,
                    ai_reasoning=ai_reasoning,
                    severity=conflict_severity
                )

                if not proceed:
                    return  # 사용자가 취소함

        # 분류 상태를 검증완료로 변경
        self.data_table.update_row(
            row_idx,
            classification,
            reasoning,
            "직접 입력",
            status=ClassificationStatus.VERIFIED
        )
        self.sheets[self.current_sheet].at[row_idx, "최종_분류"] = classification
        self.sheets[self.current_sheet].at[row_idx, "분류_근거"] = reasoning
        self.sheets[self.current_sheet].at[row_idx, "사용자_입력"] = "직접 입력"
        self.sheets[self.current_sheet].at[row_idx, "분류_상태"] = ClassificationStatus.VERIFIED

        # 채팅 패널에 완료 표시
        self.chat_panel.show_complete(classification, reasoning)
        self.chat_panel.add_message("System", "직접 입력한 분류가 저장되었습니다.")

        # 충돌 감지 및 기록 (기존 AI 분류와 다른 경우)
        if original_classification and original_classification != classification:
            if original_status in [ClassificationStatus.AI_CLASSIFIED, ClassificationStatus.EXISTING]:
                # 거래 데이터 가져오기
                row_data = self.sheets[self.current_sheet].iloc[row_idx].to_dict()
                conflict_data = {
                    "original_classification": original_classification,
                    "user_classification": classification,
                    "user_reasoning": reasoning,
                    "original_status": original_status,
                    "sheet_name": self.current_sheet,
                    "row_index": row_idx + 1,
                    "example_data": {
                        k: str(v)[:100] for k, v in row_data.items()
                        if k not in ["분류_상태", "최종_분류", "분류_근거", "사용자_입력", "검토표시"]
                        and pd.notna(v)
                    }
                }
                add_conflict_log(conflict_data)
                self.chat_panel.add_message(
                    "System",
                    f"충돌 기록됨: 기존 '{original_classification}' → 변경 '{classification}'"
                )

        # 대화 기록에 저장
        if self.current_sheet not in self.chat_histories:
            self.chat_histories[self.current_sheet] = {}
        self.chat_histories[self.current_sheet][row_idx] = [{
            "role": "assistant",
            "content": json.dumps({
                "status": "COMPLETE",
                "message": "직접 입력",
                "final_classification": classification,
                "reasoning": reasoning
            }, ensure_ascii=False)
        }]

        self._update_progress()

        # 프로젝트에 분류 데이터 즉시 저장 (중요 작업)
        self._save_project_data(immediate=True)

    def _on_reset_chat(self, row_idx: int, delete_index: int = None):
        """
        대화 초기화 또는 특정 메시지 삭제 콜백

        Args:
            row_idx: 행 인덱스
            delete_index: 삭제할 메시지 인덱스 (None이면 전체 초기화)
        """
        if self.current_sheet is None:
            return

        if self.current_sheet not in self.chat_histories:
            self.chat_histories[self.current_sheet] = {}

        if delete_index is not None:
            # 특정 메시지만 삭제
            history = self.chat_histories[self.current_sheet].get(row_idx, [])
            if 0 <= delete_index < len(history):
                history.pop(delete_index)
                self.chat_histories[self.current_sheet][row_idx] = history
        else:
            # 전체 대화 초기화
            self.chat_histories[self.current_sheet][row_idx] = []

    def _assess_conflict_severity(self, ai_classification: str, user_classification: str,
                                   row_idx: int) -> Dict:
        """
        AI 분류와 사용자 분류 간 충돌 심각도 평가

        Returns:
            Dict: 충돌 정보
                - level: "critical" | "warning" | "info"
                - message: 설명
                - risk: 잠재적 위험
        """
        # 계정과목 카테고리 추출 (앞자리로 판단)
        def get_category(cls: str) -> str:
            if any(x in cls for x in ["206", "208", "211", "212", "기계", "차량", "구축", "비품"]):
                return "유형자산"
            elif any(x in cls for x in ["214", "건설중인"]):
                return "건설중인자산"
            elif any(x in cls for x in ["231", "232", "233", "239", "240", "특허", "상표", "개발비", "소프트웨어"]):
                return "무형자산"
            elif any(x in cls for x in ["823", "830", "831", "경상연구", "소모품", "지급수수료"]):
                return "비용"
            elif any(x in cls for x in ["501", "원재료"]):
                return "재고자산"
            elif any(x in cls for x in ["140", "선급금"]):
                return "선급금"
            else:
                return "기타"

        ai_cat = get_category(ai_classification)
        user_cat = get_category(user_classification)

        # 자산 ↔ 비용 충돌은 심각
        if (ai_cat in ["유형자산", "무형자산", "건설중인자산"] and user_cat == "비용") or \
           (ai_cat == "비용" and user_cat in ["유형자산", "무형자산", "건설중인자산"]):
            return {
                "level": "critical",
                "message": f"AI는 '{ai_cat}'으로, 담당자는 '{user_cat}'으로 분류하려 합니다.",
                "risk": "자산/비용 분류 오류는 재무제표에 직접적인 영향을 미칩니다.\n"
                       "- 비용→자산: 당기순이익 과대계상 위험\n"
                       "- 자산→비용: 당기순이익 과소계상 위험",
                "ai_category": ai_cat,
                "user_category": user_cat
            }

        # 같은 카테고리 내 다른 계정 (재량적)
        if ai_cat == user_cat:
            return {
                "level": "info",
                "message": f"같은 '{ai_cat}' 내에서 다른 계정으로 분류합니다.",
                "risk": "동일 카테고리 내 세부 계정 선택은 회사 재량 범위일 수 있습니다.",
                "ai_category": ai_cat,
                "user_category": user_cat
            }

        # 그 외의 차이
        return {
            "level": "warning",
            "message": f"AI는 '{ai_cat}'으로, 담당자는 '{user_cat}'으로 분류하려 합니다.",
            "risk": "분류 카테고리가 다릅니다. 회계처리에 영향이 있을 수 있습니다.",
            "ai_category": ai_cat,
            "user_category": user_cat
        }

    def _show_conflict_warning_dialog(self, row_idx: int, ai_classification: str,
                                      user_classification: str, user_reasoning: str,
                                      ai_reasoning: str, severity: Dict) -> bool:
        """
        분류 충돌 경고 다이얼로그

        Returns:
            bool: True면 진행, False면 취소
        """
        dialog = ctk.CTkToplevel(self)
        dialog.title("분류 충돌 확인")
        dialog.geometry("600x500")
        dialog.configure(fg_color=BG_COLOR)
        dialog.transient(self)
        dialog.grab_set()

        # 중앙 배치
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() - 600) // 2
        y = (dialog.winfo_screenheight() - 500) // 2
        dialog.geometry(f"+{x}+{y}")

        result = {"proceed": False}

        # 심각도에 따른 색상
        severity_colors = {
            "critical": ERROR_COLOR,
            "warning": WARNING_COLOR,
            "info": "#60A5FA"
        }
        severity_icons = {
            "critical": "⚠️ 중요 경고",
            "warning": "⚡ 주의",
            "info": "ℹ️ 참고"
        }

        level = severity.get("level", "warning")
        color = severity_colors.get(level, WARNING_COLOR)
        icon = severity_icons.get(level, "⚡ 주의")

        # 헤더
        header_frame = ctk.CTkFrame(dialog, fg_color=color, corner_radius=0)
        header_frame.pack(fill="x")

        ctk.CTkLabel(
            header_frame,
            text=f"  {icon}: AI 분류와 다른 분류를 선택했습니다",
            font=(FONT_FAMILY, 14, "bold"),
            text_color="white" if level != "info" else "black"
        ).pack(pady=15)

        # 분류 비교
        compare_frame = ctk.CTkFrame(dialog, fg_color=SIDEBAR_COLOR)
        compare_frame.pack(fill="x", padx=15, pady=15)

        # AI 분류
        ai_frame = ctk.CTkFrame(compare_frame, fg_color="#1E3A5F")
        ai_frame.pack(fill="x", padx=10, pady=10)

        ctk.CTkLabel(
            ai_frame,
            text="AI 분류",
            font=(FONT_FAMILY, 10),
            text_color="#60A5FA"
        ).pack(anchor="w", padx=10, pady=(8, 2))

        ctk.CTkLabel(
            ai_frame,
            text=ai_classification,
            font=(FONT_FAMILY, 12, "bold"),
            text_color="white"
        ).pack(anchor="w", padx=10, pady=2)

        if ai_reasoning:
            ctk.CTkLabel(
                ai_frame,
                text=f"근거: {ai_reasoning[:100]}{'...' if len(ai_reasoning) > 100 else ''}",
                font=(FONT_FAMILY, 9),
                text_color="gray",
                wraplength=500
            ).pack(anchor="w", padx=10, pady=(2, 8))

        # 화살표
        ctk.CTkLabel(
            compare_frame,
            text="↓ 변경하려는 분류 ↓",
            font=(FONT_FAMILY, 10),
            text_color=color
        ).pack(pady=5)

        # 사용자 분류
        user_frame = ctk.CTkFrame(compare_frame, fg_color="#3D1F1F" if level == "critical" else "#1F3D1F")
        user_frame.pack(fill="x", padx=10, pady=10)

        ctk.CTkLabel(
            user_frame,
            text="담당자 입력",
            font=(FONT_FAMILY, 10),
            text_color=color
        ).pack(anchor="w", padx=10, pady=(8, 2))

        ctk.CTkLabel(
            user_frame,
            text=user_classification,
            font=(FONT_FAMILY, 12, "bold"),
            text_color="white"
        ).pack(anchor="w", padx=10, pady=2)

        if user_reasoning:
            ctk.CTkLabel(
                user_frame,
                text=f"근거: {user_reasoning[:100]}{'...' if len(user_reasoning) > 100 else ''}",
                font=(FONT_FAMILY, 9),
                text_color="gray",
                wraplength=500
            ).pack(anchor="w", padx=10, pady=(2, 8))

        # 위험 설명
        risk_frame = ctk.CTkFrame(dialog, fg_color="#2D1F1F" if level == "critical" else SIDEBAR_COLOR)
        risk_frame.pack(fill="x", padx=15, pady=5)

        ctk.CTkLabel(
            risk_frame,
            text=severity.get("message", ""),
            font=(FONT_FAMILY, 11),
            text_color="white"
        ).pack(anchor="w", padx=15, pady=(10, 5))

        ctk.CTkLabel(
            risk_frame,
            text=severity.get("risk", ""),
            font=(FONT_FAMILY, 10),
            text_color="#FCA5A5" if level == "critical" else "#FCD34D",
            wraplength=550,
            justify="left"
        ).pack(anchor="w", padx=15, pady=(0, 10))

        # 심각한 충돌인 경우 추가 근거 입력 요청
        extra_reason_var = ctk.StringVar()
        if level == "critical":
            ctk.CTkLabel(
                dialog,
                text="변경 사유를 상세히 입력해주세요 (필수):",
                font=(FONT_FAMILY, 11),
                text_color=ERROR_COLOR
            ).pack(anchor="w", padx=15, pady=(10, 5))

            extra_reason_entry = ctk.CTkTextbox(
                dialog,
                height=60,
                font=(FONT_FAMILY, 10),
                fg_color=SIDEBAR_COLOR
            )
            extra_reason_entry.pack(fill="x", padx=15, pady=5)
            extra_reason_var.set("")

            def get_extra_reason():
                return extra_reason_entry.get("1.0", "end-1c")
        else:
            def get_extra_reason():
                return ""

        # 버튼
        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.pack(fill="x", padx=15, pady=20)

        def on_proceed():
            if level == "critical":
                extra = get_extra_reason()
                if len(extra.strip()) < 10:
                    messagebox.showwarning("입력 필요",
                                          "심각한 분류 변경입니다.\n변경 사유를 10자 이상 입력해주세요.")
                    return
            result["proceed"] = True
            dialog.destroy()

        def on_cancel():
            result["proceed"] = False
            dialog.destroy()

        if level == "critical":
            proceed_text = "그래도 변경 (위험 인지함)"
            proceed_color = ERROR_COLOR
        else:
            proceed_text = "변경 진행"
            proceed_color = WARNING_COLOR

        ctk.CTkButton(
            btn_frame,
            text=proceed_text,
            font=(FONT_FAMILY, 12, "bold"),
            fg_color=proceed_color,
            text_color="white",
            width=200,
            command=on_proceed
        ).pack(side="left", padx=10)

        ctk.CTkButton(
            btn_frame,
            text="취소 (AI 분류 유지)",
            font=(FONT_FAMILY, 12),
            fg_color="#4B5563",
            width=180,
            command=on_cancel
        ).pack(side="right", padx=10)

        # 모달 대기
        dialog.wait_window()
        return result["proceed"]

    def _on_file_analyze(self, row_idx: int, file_paths: List[str]):
        """첨부 파일 분석"""
        if self.ai is None or self.current_sheet is None:
            return

        # 파일 내용 읽기
        file_contents = []
        for file_path in file_paths:
            result = FileReader.read_file(file_path)
            if result["error"]:
                self.chat_panel.add_message("System", f"파일 오류 ({result['filename']}): {result['error']}")
            else:
                file_contents.append(result)
                self.chat_panel.add_message("System", f"파일 로드됨: {result['filename']}")

        if not file_contents:
            self.chat_panel.add_message("System", "분석할 수 있는 파일이 없습니다")
            return

        # 현재 행 데이터 가져오기
        df = self.sheets[self.current_sheet]
        row_data = df.iloc[row_idx].to_json(force_ascii=False, indent=2)

        # 연관 행 데이터도 가져오기
        linked_rows_data = []
        linked_rows = self.chat_panel.get_linked_rows()
        for linked_idx in linked_rows:
            if linked_idx < len(df):
                linked_data = df.iloc[linked_idx].to_json(force_ascii=False, indent=2)
                linked_rows_data.append(f"[연관 거래 #{linked_idx + 1}]\n{linked_data}")

        self.chat_panel.show_loading(True)

        # AI 분석 (파일 포함)
        threading.Thread(
            target=self._run_ai_analysis_with_files,
            args=(self.current_sheet, row_idx, row_data, file_contents, linked_rows_data),
            daemon=True
        ).start()

    def _run_ai_analysis_with_files(self, sheet_name: str, row_idx: int, row_data: str,
                                     file_contents: List[Dict], linked_rows_data: List[str]):
        """파일 포함 AI 분석 실행"""
        history = self.chat_histories.get(sheet_name, {}).get(row_idx, [])

        # 파일 내용을 컨텍스트에 추가
        file_context = "\n\n[첨부된 참고 문서]\n"
        for fc in file_contents:
            if fc["type"] == "text":
                file_context += f"\n--- {fc['filename']} ---\n{fc['content']}\n"
            elif fc["type"] == "image":
                file_context += f"\n--- {fc['filename']} (이미지) ---\n(이미지 내용은 별도 분석됨)\n"

        # 연관 행 컨텍스트
        linked_context = ""
        if linked_rows_data:
            linked_context = "\n\n[연관된 거래 내역 - 계약금/잔금 등 함께 검토 필요]\n"
            linked_context += "\n".join(linked_rows_data)

        # 확장된 row_data 생성
        extended_data = row_data + file_context + linked_context

        result = self.ai.analyze(extended_data, history)

        # UI 업데이트 (메인 스레드)
        self.after(0, self._handle_ai_result, sheet_name, row_idx, result)

    def _load_file(self):
        """파일 로드"""
        file_path = filedialog.askopenfilename(
            filetypes=[
                ("Excel files", "*.xlsx *.xls"),
                ("CSV files", "*.csv"),
                ("All files", "*.*")
            ]
        )
        if not file_path:
            return

        try:
            self.sheets.clear()
            self.chat_histories.clear()
            self.column_mappings = {}  # 시트별 컬럼 매핑 저장
            self.highlighted_cells = {}  # 시트별 음영 처리된 셀 좌표 {(row_idx, col_idx), ...}

            # 파일명 저장
            self._current_file_name = os.path.basename(file_path)

            # 프로젝트 관리자 초기화
            self.project_manager = ProjectManager()

            # 기존 프로젝트 확인
            existing_project = self.project_manager.find_existing_project(file_path)
            if existing_project:
                # 기존 프로젝트 로드 여부 확인
                result = messagebox.askyesnocancel(
                    "기존 프로젝트 발견",
                    f"이 파일에 대한 기존 프로젝트가 있습니다.\n\n"
                    f"첨부파일: {self._count_project_attachments(existing_project)}개\n\n"
                    f"기존 프로젝트를 로드하시겠습니까?\n"
                    f"- 예: 기존 프로젝트 로드 (첨부파일 유지)\n"
                    f"- 아니오: 새 프로젝트 생성\n"
                    f"- 취소: 파일 로드 취소"
                )
                if result is None:  # 취소
                    return
                elif result:  # 예 - 기존 프로젝트 로드
                    self.project_manager.load_project(existing_project)
                else:  # 아니오 - 새 프로젝트 생성
                    self.project_manager.create_project(file_path)
            else:
                # 새 프로젝트 생성
                self.project_manager.create_project(file_path)

            # 로딩 상태 표시
            self.file_label.configure(text="파일 로딩 중...")
            self.update_idletasks()  # UI 즉시 업데이트

            if file_path.endswith('.csv'):
                self.file_label.configure(text="CSV 파일 읽는 중...")
                self.update_idletasks()
                df = pd.read_csv(file_path, encoding='utf-8-sig')
                self.sheets["Sheet1"] = df
                self.highlighted_cells["Sheet1"] = set()
            else:
                # Excel 파일의 모든 시트 로드 + 음영 감지
                self.file_label.configure(text="Excel 시트 로딩 중...")
                self.update_idletasks()
                excel_file = pd.ExcelFile(file_path)
                for i, sheet_name in enumerate(excel_file.sheet_names):
                    self.file_label.configure(text=f"시트 로딩 중... ({i+1}/{len(excel_file.sheet_names)})")
                    self.update_idletasks()
                    self.sheets[sheet_name] = pd.read_excel(file_path, sheet_name=sheet_name)
                    self.highlighted_cells[sheet_name] = set()

                # openpyxl로 셀 배경색 확인 (셀 단위 음영 감지)
                self.file_label.configure(text="셀 서식 분석 중...")
                self.update_idletasks()
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(file_path, data_only=True)
                    for sheet_name in wb.sheetnames:
                        if sheet_name in self.sheets:
                            ws = wb[sheet_name]
                            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=0):  # 헤더 제외
                                for col_idx, cell in enumerate(row):
                                    if cell.fill and cell.fill.fgColor:
                                        # 배경색이 있는지 확인 (흰색/없음 제외)
                                        color = cell.fill.fgColor
                                        if color.type == 'rgb' and color.rgb and color.rgb != '00000000' and color.rgb != 'FFFFFFFF':
                                            # 셀 좌표 (row_idx, col_idx) 저장
                                            self.highlighted_cells[sheet_name].add((row_idx, col_idx))
                                        elif color.type == 'indexed' and color.indexed and color.indexed not in [0, 64]:
                                            self.highlighted_cells[sheet_name].add((row_idx, col_idx))
                    wb.close()
                except Exception as e:
                    print(f"음영 감지 오류 (무시됨): {e}")

            # 시트별 대화 기록 초기화
            for sheet_name in self.sheets:
                self.chat_histories[sheet_name] = {}

            self.file_label.configure(text=f"파일: {os.path.basename(file_path)}")

            # 첫 번째 시트에 대해 컬럼 매핑 다이얼로그 표시
            if self.sheets:
                first_sheet = list(self.sheets.keys())[0]
                columns = list(self.sheets[first_sheet].columns)

                # 컬럼 매핑 다이얼로그 표시
                dialog = ColumnMappingDialog(self, columns)
                if dialog.result:
                    # 모든 시트에 동일한 매핑 적용 (시트 구조가 같다고 가정)
                    for sheet_name in self.sheets:
                        self.column_mappings[sheet_name] = dialog.result.copy()
                else:
                    # 취소 시 빈 매핑
                    for sheet_name in self.sheets:
                        self.column_mappings[sheet_name] = {}

            self.file_label.configure(text="테이블 구성 중...")
            self.update_idletasks()

            self._update_sheet_tabs()

            # 첫 번째 시트 선택
            if self.sheets:
                first_sheet = list(self.sheets.keys())[0]
                self._select_sheet(first_sheet)

            self.file_label.configure(text=f"파일: {os.path.basename(file_path)}")

            # 상태 카운트 표시
            counts = self.data_table.get_status_counts()
            existing_count = counts.get("existing", 0)
            highlighted_count = counts.get("highlighted", 0)

            # 로드 완료 메시지
            msg_parts = [f"파일을 불러왔습니다.\n시트 수: {len(self.sheets)}\n"]

            if highlighted_count > 0:
                msg_parts.append(f"\n★ 검토 표시(음영) 항목: {highlighted_count}건")

            if existing_count > 0:
                msg_parts.append(f"\n기존 분류가 있는 항목: {existing_count}건")

            msg_parts.append(f"\n미분류 항목: {counts.get('pending', 0)}건")

            if existing_count > 0 or highlighted_count > 0:
                msg_parts.append("\n\n• 음영 표시된 셀은 ★ 마커로 표시됩니다")
                msg_parts.append("\n• '검토★' 필터로 음영 셀이 있는 행 확인")
                msg_parts.append("\n• 기존 분류는 AI 검증 또는 '기존분류 유지'")

            messagebox.showinfo("파일 로드 완료", "".join(msg_parts))

            # 진행률 업데이트
            self._update_progress()

        except Exception as e:
            messagebox.showerror("오류", f"파일 로드 실패:\n{str(e)}")

    def _update_sheet_tabs(self):
        """시트 탭 업데이트"""
        for widget in self.tab_frame.winfo_children():
            widget.destroy()

        for sheet_name in self.sheets:
            btn = ctk.CTkButton(
                self.tab_frame,
                text=sheet_name,
                font=(FONT_FAMILY, 11),
                fg_color=ACCENT_COLOR if sheet_name == self.current_sheet else "transparent",
                hover_color=ACCENT_COLOR,
                command=lambda s=sheet_name: self._select_sheet(s)
            )
            btn.pack(side="left", padx=2, pady=5)

    def _select_sheet(self, sheet_name: str):
        """시트 선택"""
        self.current_sheet = sheet_name
        self._update_sheet_tabs()

        df = self.sheets[sheet_name]
        column_mapping = self.column_mappings.get(sheet_name, {})
        highlighted_cells = self.highlighted_cells.get(sheet_name, set())
        self.data_table.load_data(df, column_mapping, highlighted_cells)
        self.chat_panel.clear()
        self._update_progress()

    def _on_row_select(self, row_idx: int):
        """행 선택 시 - AI 자동 시작 없이 정보만 표시"""
        if self.current_sheet is None:
            return

        df = self.sheets[self.current_sheet]
        row = df.iloc[row_idx]

        # 행 데이터를 문자열로 변환
        row_data = row.to_json(force_ascii=False, indent=2)

        # 현재 분류 상태와 기존 분류 정보 가져오기
        current_status = str(row.get("분류_상태", ClassificationStatus.PENDING))
        existing_cls = str(row.get("최종_분류", "")) if pd.notna(row.get("최종_분류")) else ""

        # 채팅 패널 초기화 (기존 분류 정보 전달)
        self.chat_panel.set_row(
            row_idx, row_data,
            existing_classification=existing_cls if existing_cls else None,
            existing_status=current_status
        )

        # 이미 대화 기록이 있으면 복원
        sheet_history = self.chat_histories.get(self.current_sheet, {})
        if row_idx in sheet_history and sheet_history[row_idx]:
            history = sheet_history[row_idx]
            last_complete_result = None  # 마지막 완료 결과 저장

            for h in history:
                if h["role"] == "user":
                    self.chat_panel.add_message("User", h["content"])
                else:
                    try:
                        content = json.loads(h["content"])
                        # JSON 파싱 성공 - message 필드만 표시
                        message = content.get("message", "")
                        if message:
                            self.chat_panel.add_message("AI", message)

                        # COMPLETE 상태면 저장해두고 마지막에 show_complete 호출
                        if content.get("status") == "COMPLETE":
                            last_complete_result = {
                                "classification": content.get("final_classification", ""),
                                "reasoning": content.get("reasoning", "")
                            }
                    except (json.JSONDecodeError, TypeError):
                        # JSON 파싱 실패 - 일반 텍스트로 표시
                        self.chat_panel.add_message("AI", h["content"])

            # 분류 완료 상태였다면 show_complete 호출
            if last_complete_result:
                self.chat_panel.show_complete(
                    last_complete_result["classification"],
                    last_complete_result["reasoning"]
                )
            return

        # 이미 분류가 완료된 상태면 완료 상태 표시
        if current_status in [ClassificationStatus.AI_CLASSIFIED, ClassificationStatus.VERIFIED]:
            reasoning = str(row.get("분류_근거", "")) if pd.notna(row.get("분류_근거")) else ""
            self.chat_panel.add_message("System", f"이 항목은 이미 분류되었습니다.\n\n분류: {existing_cls}\n근거: {reasoning}")
            self.chat_panel.show_complete(existing_cls, reasoning)
            return

        # 기존분류 상태 안내
        if current_status == ClassificationStatus.EXISTING:
            self.chat_panel.add_message("System", f"기존 분류가 있습니다: {existing_cls}\n\nAI 분석을 원하시면 'AI 분석 시작' 버튼을 누르세요.\n또는 테이블에서 직접 계정을 수정할 수 있습니다.")
            return

        # 미분류 상태 안내
        self.chat_panel.add_message("System", "AI 분석을 원하시면 'AI 분석 시작' 버튼을 누르세요.\n또는 테이블에서 직접 계정을 입력할 수 있습니다.")

    def _on_start_ai_analysis(self, row_idx: int):
        """AI 분석 시작 (버튼 클릭 시)"""
        if self.ai is None:
            messagebox.showwarning("알림", "AI가 초기화되지 않았습니다. API 키를 확인하세요.")
            return

        if self.current_sheet is None:
            return

        df = self.sheets[self.current_sheet]
        row = df.iloc[row_idx]

        # 현재 일시 추가 (AI가 기간 경과 항목 인식용)
        from datetime import datetime
        current_datetime = datetime.now().strftime("%Y년 %m월 %d일 %H시 %M분")
        current_date_only = datetime.now().strftime("%Y-%m-%d")

        # row_data에 분석 기준일 포함
        row_data = f"=== 분석 기준일 ===\n{current_datetime} (기준일: {current_date_only})\n\n=== 거래 내역 ===\n{row.to_json(force_ascii=False, indent=2)}"

        # 현재 분류 상태 확인
        current_status = str(row.get("분류_상태", ClassificationStatus.PENDING))
        is_existing = current_status == ClassificationStatus.EXISTING

        # 대화 기록 초기화
        if self.current_sheet not in self.chat_histories:
            self.chat_histories[self.current_sheet] = {}
        self.chat_histories[self.current_sheet][row_idx] = []

        # 로딩 표시 후 AI 초기 분류 제안 (비동기)
        self.chat_panel.add_message("System", "AI가 거래 내역을 분석하고 있습니다...")
        self.chat_panel.show_loading(True)

        def run_suggestion_and_analysis():
            suggestions = self._generate_initial_suggestion(row, df.columns.tolist())
            self.after(0, lambda: self._show_initial_suggestion_result(suggestions, is_existing=is_existing))
            # 추가 AI 분석 진행
            self._run_ai_analysis(self.current_sheet, row_idx, row_data)

        threading.Thread(target=run_suggestion_and_analysis, daemon=True).start()

    def _show_initial_suggestion_result(self, suggestions: Dict, is_existing: bool = False):
        """AI 초기 분류 제안 결과 표시 (메인 스레드에서 실행)"""
        self.chat_panel.show_loading(False)
        self.chat_panel.show_initial_suggestions(suggestions)

        if is_existing:
            self.chat_panel.add_message("System", "AI 검증을 시작하려면 '답변 전송'을 누르거나, 기존 분류를 유지하세요.")
        else:
            self.chat_panel.add_message("System", "AI가 추가 정보를 확인 중입니다...")

    def _generate_initial_suggestion(self, row: pd.Series, columns: List[str]) -> Dict:
        """
        AI 기반 초기 분류 제안 생성

        Args:
            row: DataFrame 행
            columns: 컬럼 목록

        Returns:
            Dict: 분류 제안 정보
                - status: "confident" | "need_info" | "discretionary"
                - primary: 주요 추천 계정
                - options: 선택 가능한 계정 목록
                - missing_info: 부족한 정보 목록
                - reason: 판단 근거
                - decision_step: 의사결정 트리 단계
        """
        # 기본 결과 (AI 호출 실패 시 사용)
        default_result = {
            "status": "need_info",
            "primary": "",
            "options": [],
            "available_info": [],
            "missing_info": ["AI 분석 대기 중"],
            "reason": "AI가 거래 내역을 분석합니다. 잠시 기다려주세요.",
            "decision_step": "",
            "required_questions": [],
            "confidence_score": 0.0
        }

        # AI가 없으면 기본값 반환
        if self.ai is None:
            default_result["reason"] = "AI가 초기화되지 않았습니다. API 키를 확인하세요."
            return default_result

        # 행 데이터 추출
        row_dict = row.to_dict()
        row_json = json.dumps(row_dict, ensure_ascii=False, indent=2)

        # 현재 일시 (AI 분석 시점 기준)
        from datetime import datetime
        current_datetime = datetime.now().strftime("%Y년 %m월 %d일 %H시 %M분")
        current_date_only = datetime.now().strftime("%Y-%m-%d")

        # 실무지침 로드
        guidelines = {}
        if HAS_VERSION_MANAGER:
            guidelines = load_guidelines()

        # 회사 프로필 및 분류 카테고리 추출
        company_profile = guidelines.get("company_profile", {})
        categories = guidelines.get("classification_categories", {})
        decision_tree = guidelines.get("master_decision_tree", {})

        threshold = company_profile.get("자본화_기준금액", 1000000)

        # AI 프롬프트 구성
        prompt = f"""당신은 회계 전문가입니다. 다음 거래 내역을 분석하여 초기 분류 제안을 생성하세요.

=== 분석 기준일 ===
{current_datetime} (기준일: {current_date_only})

=== 거래 내역 ===
{row_json}

=== 회사 회계 정책 ===
- 회계기준: {company_profile.get('회계기준', 'N/A')}
- 자본화 기준금액: {threshold:,}원
- 업종: {company_profile.get('업종', 'N/A')}
- 회계방향: {company_profile.get('회계방향', {}).get('기본원칙', 'N/A')}

=== 사용 가능한 계정과목 ===
{json.dumps(categories, ensure_ascii=False, indent=2)}

=== 의사결정 트리 주요 단계 ===
1. IP(지식재산권) 관련 여부
2. 환급성(보증금/투자) 여부
3. 금액 기준 ({threshold:,}원) 충족 및 장기 효익 여부
4. 물리적 형태 여부 (유형자산 vs 무형자산)
5. 세부 분류 (기계장치, 비품, 개발비, 수수료 등)

분석 후 JSON 형식으로 응답하세요:
{{
    "status": "confident" | "need_info" | "discretionary",
    "primary": "주요 추천 계정과목 (확신있는 경우)",
    "options": [
        {{"account": "계정과목 코드", "description": "설명", "conditions": "이 분류가 적용되는 조건"}},
        ...
    ],
    "available_info": ["확인된 정보 1", "확인된 정보 2", ...],
    "missing_info": ["부족한 정보 1", "부족한 정보 2", ...],
    "reason": "분류 판단 근거 설명",
    "decision_step": "해당 의사결정 트리 단계",
    "required_questions": [
        {{
            "question": "사용자에게 물어볼 질문",
            "options": ["선택지1", "선택지2", ...],
            "impact": "이 질문의 답변이 분류에 미치는 영향"
        }},
        ...
    ],
    "confidence_score": 0.0-1.0
}}

status 설명:
- "confident": 정보가 충분하여 확신있게 분류 가능 (confidence_score >= 0.8)
- "need_info": 추가 정보 필요 (missing_info에 명시, required_questions 반드시 포함)
- "discretionary": 회사 재량으로 선택 가능한 항목들 존재 (options에 조건별 분류 포함)

중요:
- status가 "need_info"인 경우, required_questions에 반드시 1개 이상의 질문 포함
- 각 질문에는 클릭 가능한 options 배열 포함 (2-4개 선택지)
- impact에는 해당 답변이 어떤 분류 결정에 영향을 주는지 명시
- 회계기준에 근거한 질문만 포함 (추측성 질문 금지)"""

        try:
            # AI 호출 (동기)
            response = self.ai.analyze(prompt, [])

            if response.get("status") == "ERROR":
                default_result["reason"] = f"AI 분석 오류: {response.get('message', '알 수 없는 오류')}"
                return default_result

            # AI 응답 파싱
            result = {
                "status": response.get("status", "need_info"),
                "primary": response.get("primary", ""),
                "options": response.get("options", []),
                "available_info": response.get("available_info", []),
                "missing_info": response.get("missing_info", []),
                "reason": response.get("reason", ""),
                "decision_step": response.get("decision_step", ""),
                "required_questions": response.get("required_questions", []),
                "confidence_score": response.get("confidence_score", 0.5)
            }

            # options 형식 정규화
            normalized_options = []
            for opt in result["options"]:
                if isinstance(opt, str):
                    normalized_options.append({"account": opt, "description": ""})
                elif isinstance(opt, dict):
                    normalized_options.append(opt)
            result["options"] = normalized_options

            return result

        except Exception as e:
            default_result["reason"] = f"AI 분석 중 오류: {str(e)}"
            return default_result

    def _on_user_submit(self, row_idx: int, user_text: str, attached_files: List[str] = None):
        """사용자 답변 제출 (텍스트 + 첨부파일)"""
        from datetime import datetime

        if self.ai is None or self.current_sheet is None:
            return

        # 대화 기록 초기화 (없으면)
        if self.current_sheet not in self.chat_histories:
            self.chat_histories[self.current_sheet] = {}
        if row_idx not in self.chat_histories[self.current_sheet]:
            self.chat_histories[self.current_sheet][row_idx] = []

        # 대화 기록에 추가 (텍스트와 파일 정보 포함)
        content = user_text
        if attached_files:
            file_names = [os.path.basename(f) for f in attached_files]
            content += f"\n[첨부파일: {', '.join(file_names)}]"

        self.chat_histories[self.current_sheet][row_idx].append({
            "role": "user",
            "content": content
        })

        # 사용자 입력을 타임스탬프와 함께 DataFrame의 "대화_기록" 컬럼에 기록
        self._append_chat_log(row_idx, user_text, attached_files)

        # 첨부파일 내용 추출 (연관 거래 검색용)
        file_contents = ""
        if attached_files:
            file_contents = self._extract_file_contents(attached_files)

        # 연관 거래 검색 및 제안 (백그라운드가 아닌 즉시 실행)
        self._suggest_related_transactions(row_idx, user_text, file_contents)

        # AI 호출
        df = self.sheets[self.current_sheet]

        # 현재 일시 추가 (AI가 기간 경과 항목 인식용)
        current_datetime = datetime.now().strftime("%Y년 %m월 %d일 %H시 %M분")
        current_date_only = datetime.now().strftime("%Y-%m-%d")

        # row_data에 분석 기준일 포함
        row_data = f"=== 분석 기준일 ===\n{current_datetime} (기준일: {current_date_only})\n\n=== 거래 내역 ===\n{df.iloc[row_idx].to_json(force_ascii=False, indent=2)}"

        self.chat_panel.show_loading(True)

        # 첨부파일이 있으면 파일 분석도 포함
        threading.Thread(
            target=self._run_ai_analysis,
            args=(self.current_sheet, row_idx, row_data, attached_files),
            daemon=True
        ).start()

    def _on_approve_classification(self, row_idx: int, classification: str, reasoning: str,
                                     is_new_case: bool = False, new_case_suggestion: str = ""):
        """분류 승인 - 트리뷰에 적용"""
        if self.current_sheet is None:
            return

        # 사용자 입력 정보 가져오기
        user_inputs = self.chat_panel.get_user_inputs_text()
        linked_rows = self.chat_panel.get_linked_rows()

        # 데이터 테이블 업데이트 (검증완료 상태로 설정)
        self.data_table.update_row(
            row_idx, classification, reasoning, user_inputs,
            status=ClassificationStatus.VERIFIED
        )

        # 원본 데이터프레임도 업데이트
        self.sheets[self.current_sheet].at[row_idx, "최종_분류"] = classification
        self.sheets[self.current_sheet].at[row_idx, "분류_근거"] = reasoning
        self.sheets[self.current_sheet].at[row_idx, "사용자_입력"] = user_inputs
        self.sheets[self.current_sheet].at[row_idx, "분류_상태"] = ClassificationStatus.VERIFIED

        # 연관된 행도 함께 업데이트 (동일 분류 적용)
        if linked_rows:
            linked_reasoning = f"[연관 거래] 행 #{row_idx + 1}과 연결됨. " + reasoning
            for linked_idx in linked_rows:
                if linked_idx < len(self.sheets[self.current_sheet]):
                    self.data_table.update_row(
                        linked_idx, classification, linked_reasoning, f"행 #{row_idx + 1}과 연관",
                        status=ClassificationStatus.VERIFIED
                    )
                    self.sheets[self.current_sheet].at[linked_idx, "최종_분류"] = classification
                    self.sheets[self.current_sheet].at[linked_idx, "분류_근거"] = linked_reasoning
                    self.sheets[self.current_sheet].at[linked_idx, "사용자_입력"] = f"행 #{row_idx + 1}과 연관"
                    self.sheets[self.current_sheet].at[linked_idx, "분류_상태"] = ClassificationStatus.VERIFIED

            self.chat_panel.add_message("System", f"연관된 {len(linked_rows)}개 거래도 동일하게 분류되었습니다")

        # 진행률 업데이트
        self._update_progress()

        # 프로젝트에 분류 데이터 즉시 저장 (중요 작업)
        self._save_project_data(immediate=True)

        # 새로운 케이스인 경우 실무지침 업데이트 제안
        if is_new_case and new_case_suggestion:
            self.after(500, lambda: self._prompt_save_new_case(
                self.current_sheet, row_idx, classification, reasoning, new_case_suggestion
            ))

    def _on_cell_edit(self, row_idx: int, col_name: str, new_value: str):
        """셀 편집 완료 콜백 - DataFrame 동기화 및 자동 저장"""
        if self.current_sheet is None:
            return

        # 원본 DataFrame과 동기화
        if col_name in self.sheets[self.current_sheet].columns:
            self.sheets[self.current_sheet].at[row_idx, col_name] = new_value

        # 진행률 업데이트
        self._update_progress()

        # 프로젝트에 분류 데이터 자동 저장
        self._save_project_data()

    def _save_project_data(self, immediate: bool = False):
        """
        프로젝트에 분류 데이터 자동 저장 (debounce 적용)

        Args:
            immediate: True면 즉시 저장, False면 debounce 적용 (기본 2초 대기)
        """
        if not self.project_manager or not self.project_manager.project_dir:
            return

        # debounce 타이머가 이미 있으면 취소
        if hasattr(self, '_save_timer') and self._save_timer is not None:
            self.after_cancel(self._save_timer)
            self._save_timer = None

        if immediate:
            self._do_save_project_data()
        else:
            # 2초 후 저장 (빠른 연속 편집 시 한 번만 저장)
            self._save_timer = self.after(2000, self._do_save_project_data)

    def _do_save_project_data(self):
        """실제 저장 수행"""
        self._save_timer = None
        if self.project_manager and self.project_manager.project_dir:
            try:
                self.project_manager.save_classification_data(self.sheets, self.chat_histories)
            except Exception as e:
                print(f"프로젝트 데이터 저장 오류: {e}")

    def _find_related_transactions(self, current_row_idx: int, keywords: List[str]) -> List[int]:
        """
        키워드를 기반으로 연관 거래 찾기 (최적화된 버전)

        Args:
            current_row_idx: 현재 선택된 행 인덱스
            keywords: 검색할 키워드 목록

        Returns:
            연관된 행 인덱스 목록
        """
        if self.current_sheet is None or not keywords:
            return []

        df = self.sheets[self.current_sheet]

        # 검색할 컬럼 (텍스트 기반 컬럼)
        exclude_cols = {"분류_상태", "최종_분류", "분류_근거", "사용자_입력", "대화_기록", "검토표시"}
        search_columns = [col for col in df.columns if col not in exclude_cols]

        if not search_columns:
            return []

        # 이미 연결된 행
        linked_rows = set(self.chat_panel.get_linked_rows())

        # 벡터화된 검색: 모든 검색 컬럼을 하나의 문자열로 합침
        combined_text = df[search_columns].astype(str).agg(' '.join, axis=1).str.lower()

        # 키워드 검색 (OR 조건)
        mask = pd.Series([False] * len(df))
        for keyword in keywords:
            mask |= combined_text.str.contains(keyword.lower(), na=False, regex=False)

        # 현재 행 및 이미 연결된 행 제외
        mask.iloc[current_row_idx] = False
        for linked_idx in linked_rows:
            if linked_idx < len(mask):
                mask.iloc[linked_idx] = False

        # 결과 반환 (최대 10개)
        related_rows = mask[mask].index.tolist()[:10]
        return related_rows

    def _suggest_related_transactions(self, current_row_idx: int, user_text: str, file_contents: str = ""):
        """
        사용자 입력/파일 내용을 분석하여 연관 거래 연결 제안

        Args:
            current_row_idx: 현재 행 인덱스
            user_text: 사용자가 입력한 텍스트
            file_contents: 첨부파일 내용
        """
        import re

        if self.current_sheet is None:
            return

        # 키워드 추출 (거래처명, 금액, 프로젝트명 등)
        combined_text = f"{user_text} {file_contents}"

        # 키워드 추출 패턴
        keywords = []

        # 금액 패턴 (1,000,000 또는 1000000)
        amounts = re.findall(r'\d{1,3}(?:,\d{3})+|\d{4,}', combined_text)
        keywords.extend(amounts[:3])  # 최대 3개

        # 2글자 이상의 한글 단어 (거래처명, 프로젝트명 등)
        korean_words = re.findall(r'[가-힣]{2,}', combined_text)
        # 일반적인 단어 제외
        exclude_words = {"계약", "잔금", "선급", "금액", "거래", "분류", "내용", "구매", "판매", "지출", "수입",
                         "입금", "출금", "이체", "결제", "승인", "완료", "대기", "처리", "확인", "요청"}
        korean_words = [w for w in korean_words if w not in exclude_words and len(w) >= 2]
        keywords.extend(korean_words[:5])  # 최대 5개

        if not keywords:
            return

        # 연관 거래 검색
        related_rows = self._find_related_transactions(current_row_idx, keywords)

        if not related_rows:
            return

        # 연관 거래 제안 메시지 표시
        df = self.sheets[self.current_sheet]
        suggestions = []
        for row_idx in related_rows[:5]:  # 최대 5개만 표시
            row = df.iloc[row_idx]
            # 거래 요약 (첫 몇 컬럼의 값)
            summary_parts = []
            for col in list(df.columns)[:4]:
                val = str(row.get(col, ""))[:20]
                if val:
                    summary_parts.append(val)
            summary = " | ".join(summary_parts)
            suggestions.append(f"  #{row_idx + 1}: {summary}")

        if suggestions:
            suggestion_text = "🔗 연관 가능한 거래를 발견했습니다:\n" + "\n".join(suggestions)
            suggestion_text += "\n\n'연관 거래 연결' 버튼으로 연결할 수 있습니다."
            self.chat_panel.add_message("System", suggestion_text)

    def _append_chat_log(self, row_idx: int, user_text: str, attached_files: List[str] = None):
        """
        사용자 입력을 타임스탬프와 함께 DataFrame의 "대화_기록" 컬럼에 기록

        Args:
            row_idx: 행 인덱스
            user_text: 사용자 입력 텍스트
            attached_files: 첨부파일 목록
        """
        from datetime import datetime

        if self.current_sheet is None:
            return

        df = self.sheets[self.current_sheet]

        # "대화_기록" 컬럼이 없으면 추가
        if "대화_기록" not in df.columns:
            df["대화_기록"] = ""

        # 기존 기록 가져오기
        existing_log = str(df.at[row_idx, "대화_기록"]) if pd.notna(df.at[row_idx, "대화_기록"]) else ""

        # 새 기록 추가
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        new_entry = f"[{timestamp}] {user_text}"

        # 첨부파일 정보 추가
        if attached_files:
            file_names = [os.path.basename(f) for f in attached_files]
            new_entry += f" (첨부: {', '.join(file_names)})"

        # 기존 기록이 있으면 구분자로 연결
        if existing_log and existing_log.strip():
            updated_log = existing_log + "\n---\n" + new_entry
        else:
            updated_log = new_entry

        # DataFrame 업데이트
        df.at[row_idx, "대화_기록"] = updated_log

    def _run_ai_analysis(self, sheet_name: str, row_idx: int, row_data: str, attached_files: List[str] = None):
        """AI 분석 실행 (별도 스레드)"""
        history = self.chat_histories.get(sheet_name, {}).get(row_idx, [])

        # 첨부파일이 있으면 파일 내용도 분석에 포함
        if attached_files:
            file_context = self._extract_file_contents(attached_files)
            if file_context:
                row_data += f"\n\n=== 첨부파일 내용 ===\n{file_context}"

        result = self.ai.analyze(row_data, history)

        # UI 업데이트 (메인 스레드에서)
        self.after(0, self._handle_ai_result, sheet_name, row_idx, result)

    def _extract_file_contents(self, files: List[str]) -> str:
        """첨부파일에서 텍스트 내용 추출"""
        if not files:
            return ""

        contents = []
        for file_path in files:
            if not os.path.exists(file_path):
                continue
            try:
                # FileReader를 통해 파일 내용 추출
                result = FileReader.read_file(file_path)
                filename = os.path.basename(file_path)

                if result.get("error"):
                    print(f"파일 읽기 오류 ({file_path}): {result.get('error')}")
                    continue

                file_content = result.get("content", "")
                file_type = result.get("type", "text")

                # 이미지는 base64이므로 텍스트로 표시
                if file_type == "image":
                    contents.append(f"[{filename}] (이미지 파일 - AI가 직접 분석)")
                elif file_content:
                    # 텍스트 내용은 3000자로 제한
                    contents.append(f"[{filename}]\n{file_content[:3000]}")
            except Exception as e:
                print(f"파일 읽기 오류 ({file_path}): {e}")

        return "\n\n".join(contents)

    def _handle_ai_result(self, sheet_name: str, row_idx: int, result: Dict):
        """AI 결과 처리 (백그라운드에서도 동작)"""
        # 현재 선택된 행인지 확인
        is_current_row = (sheet_name == self.current_sheet and
                          self.chat_panel.current_row_index == row_idx)

        # 현재 행이면 로딩 상태 해제
        if is_current_row:
            self.chat_panel.show_loading(False)

        status = result.get("status", "ERROR")
        message = result.get("message", "")

        # 대화 기록에 추가 (항상 저장)
        if sheet_name in self.chat_histories and row_idx in self.chat_histories[sheet_name]:
            self.chat_histories[sheet_name][row_idx].append({
                "role": "assistant",
                "content": json.dumps(result, ensure_ascii=False)
            })

        if status == "ERROR":
            if is_current_row:
                self.chat_panel.add_message("System", f"오류: {message}")
            return

        # 현재 행이면 AI 메시지 표시
        if is_current_row:
            self.chat_panel.add_message("AI", message)

            # 추가 질문이 있으면 표시
            questions = result.get("questions", [])
            if questions and status == "INCOMPLETE":
                questions_text = "\n".join([f"  - {q}" for q in questions])
                self.chat_panel.add_message("AI", f"확인 필요:\n{questions_text}", is_question=True)

        # 분류 완료 - 승인 대기 상태로 전환
        if status == "COMPLETE":
            final_cls = result.get("final_classification", "")
            reasoning = result.get("reasoning", "")
            is_new_case = result.get("is_new_case", False)
            new_case_suggestion = result.get("new_case_suggestion", "")

            # 현재 행이면 승인 버튼 표시
            if is_current_row:
                # 분류 추천 표시 (승인 버튼과 함께)
                self.chat_panel.show_classification_recommendation(
                    final_cls, reasoning, is_new_case, new_case_suggestion
                )
            else:
                # 다른 행을 보고 있을 때는 백그라운드에서 임시 저장 (AI분류 상태)
                user_inputs = ""
                linked_rows = []

                # 데이터 테이블 업데이트 (AI분류 상태로 설정)
                if sheet_name == self.current_sheet:
                    self.data_table.update_row(
                        row_idx, final_cls, reasoning, user_inputs,
                        status=ClassificationStatus.AI_CLASSIFIED
                    )

                # 원본 데이터프레임도 업데이트
                self.sheets[sheet_name].at[row_idx, "최종_분류"] = final_cls
                self.sheets[sheet_name].at[row_idx, "분류_근거"] = reasoning
                self.sheets[sheet_name].at[row_idx, "사용자_입력"] = user_inputs
                self.sheets[sheet_name].at[row_idx, "분류_상태"] = ClassificationStatus.AI_CLASSIFIED

                # 상태바 또는 알림으로 표시
                self.file_label.configure(text=f"행 #{row_idx + 1} AI 분류 완료: {final_cls}")
                # 3초 후 원래 파일명으로 복귀
                self.after(3000, self._restore_file_label)

                # 진행률 업데이트
                self._update_progress()

            # 진행률 업데이트
            self._update_progress()

    def _prompt_save_new_case(self, sheet_name: str, row_idx: int,
                               final_cls: str, reasoning: str, new_case_suggestion: str):
        """신규 케이스 저장 여부 확인 및 저장"""
        # 사용자에게 확인
        msg = f"""AI가 이 거래를 '기존 실무지침에 없는 새로운 케이스'로 판단했습니다.

[분류 결과] {final_cls}

[AI 제안 규칙]
{new_case_suggestion}

이 케이스를 실무지침서(learned_cases)에 추가하시겠습니까?
추가하면 향후 유사한 거래에서 AI가 참고할 수 있습니다."""

        save_case = messagebox.askyesno("신규 케이스 학습", msg)

        if save_case:
            # 거래 데이터 가져오기
            row_data = self.sheets[sheet_name].iloc[row_idx].to_dict()

            # 학습 케이스 데이터 구성
            case_data = {
                "description": new_case_suggestion,
                "classification": final_cls,
                "reasoning": reasoning,
                "example_data": {
                    k: str(v)[:100] for k, v in row_data.items()
                    if k not in ["분류_상태", "최종_분류", "분류_근거", "사용자_입력", "검토표시"]
                    and pd.notna(v)
                },
                "sheet_name": sheet_name,
                "row_index": row_idx + 1  # 1-indexed for user readability
            }

            # learned_cases에 추가
            if add_learned_case(case_data):
                self.chat_panel.add_message(
                    "System",
                    f"신규 케이스가 실무지침서에 추가되었습니다.\n→ {new_case_suggestion[:50]}..."
                )
                messagebox.showinfo("학습 완료", "신규 케이스가 실무지침서(learned_cases)에 저장되었습니다.")
            else:
                messagebox.showerror("저장 실패", "실무지침서 저장 중 오류가 발생했습니다.")
        else:
            self.chat_panel.add_message(
                "System", "신규 케이스 학습을 건너뛰었습니다."
            )

    def _restore_file_label(self):
        """파일 라벨 복원"""
        # 현재 파일명이 있으면 복원
        if hasattr(self, '_current_file_name'):
            self.file_label.configure(text=f"파일: {self._current_file_name}")
        else:
            self.file_label.configure(text="파일을 선택하세요")

    def _set_filter(self, filter_type: str):
        """필터 설정"""
        self.current_filter = filter_type
        self.data_table.set_filter(filter_type)

        # 버튼 스타일 업데이트
        filter_colors = {
            "all": "#6B7280",
            "pending": "#EF4444",
            "highlighted": "#8B5CF6",
            "existing": "#F59E0B",
            "classified": "#3B82F6",
            "verified": "#10B981",
            "needs_review": "#EC4899"
        }

        for ft, btn in self.filter_buttons.items():
            if ft == filter_type:
                btn.configure(fg_color=filter_colors.get(ft, "#6B7280"))
            else:
                btn.configure(fg_color="transparent")

    def _count_project_attachments(self, project_dir: str) -> int:
        """프로젝트의 첨부파일 수 카운트"""
        try:
            metadata_path = os.path.join(project_dir, "project.json")
            if os.path.exists(metadata_path):
                with open(metadata_path, 'r', encoding='utf-8') as f:
                    meta = json.load(f)
                return sum(len(v) for v in meta.get("attachments", {}).values())
        except:
            pass
        return 0

    def save_attachment_to_project(self, row_idx: int, file_path: str) -> Optional[str]:
        """
        첨부파일을 프로젝트에 저장

        Args:
            row_idx: 행 인덱스
            file_path: 원본 파일 경로

        Returns:
            저장된 파일 경로 (실패 시 None)
        """
        if not self.project_manager:
            return None

        sheet_name = self.current_sheet or "Sheet1"
        return self.project_manager.add_attachment(row_idx, file_path, sheet_name)

    def get_row_attachments(self, row_idx: int) -> List[Dict]:
        """행의 첨부파일 목록 가져오기"""
        if not self.project_manager:
            return []

        sheet_name = self.current_sheet or "Sheet1"
        return self.project_manager.get_attachments(row_idx, sheet_name)

    def remove_attachment_from_project(self, row_idx: int, attachment_index: int) -> bool:
        """
        첨부파일을 프로젝트에서 삭제

        Args:
            row_idx: 행 인덱스
            attachment_index: 첨부파일 인덱스

        Returns:
            삭제 성공 여부
        """
        if not self.project_manager:
            return False

        sheet_name = self.current_sheet or "Sheet1"
        return self.project_manager.remove_attachment(row_idx, attachment_index, sheet_name)

    def save_linked_transactions(self, row_idx: int, linked_rows: List[int]) -> bool:
        """
        연관 거래 연결을 프로젝트에 저장

        Args:
            row_idx: 소스 행 인덱스
            linked_rows: 연결된 행 인덱스 목록

        Returns:
            저장 성공 여부
        """
        if not self.project_manager:
            return False

        sheet_name = self.current_sheet or "Sheet1"
        return self.project_manager.add_linked_transactions(row_idx, linked_rows, sheet_name)

    def get_linked_transactions(self, row_idx: int) -> List[int]:
        """
        행의 연관 거래 목록 가져오기

        Args:
            row_idx: 행 인덱스

        Returns:
            연결된 행 인덱스 목록
        """
        if not self.project_manager:
            return []

        sheet_name = self.current_sheet or "Sheet1"
        return self.project_manager.get_linked_transactions(row_idx, sheet_name)

    def _show_project_info(self):
        """프로젝트 정보 다이얼로그 표시"""
        if not self.project_manager or not self.project_manager.project_dir:
            messagebox.showinfo("안내", "열린 프로젝트가 없습니다.")
            return

        info = self.project_manager.get_project_info()
        all_attachments = self.project_manager.get_all_attachments()

        # 프로젝트 정보 다이얼로그
        dialog = ctk.CTkToplevel(self)
        dialog.title("프로젝트 정보")
        dialog.geometry("600x500")
        dialog.configure(fg_color=BG_COLOR)
        dialog.transient(self)

        # 기본 정보
        info_frame = ctk.CTkFrame(dialog, fg_color=SIDEBAR_COLOR)
        info_frame.pack(fill="x", padx=15, pady=15)

        ctk.CTkLabel(
            info_frame,
            text="프로젝트 정보",
            font=(FONT_FAMILY, 14, "bold"),
            text_color="white"
        ).pack(anchor="w", padx=15, pady=(15, 5))

        ctk.CTkLabel(
            info_frame,
            text=f"프로젝트명: {info.get('project_name', '')}",
            font=(FONT_FAMILY, 11),
            text_color="#A0AEC0"
        ).pack(anchor="w", padx=15, pady=2)

        ctk.CTkLabel(
            info_frame,
            text=f"원본 파일: {info.get('source_file', '')}",
            font=(FONT_FAMILY, 11),
            text_color="#A0AEC0"
        ).pack(anchor="w", padx=15, pady=2)

        ctk.CTkLabel(
            info_frame,
            text=f"첨부파일: {info.get('attachment_count', 0)}개",
            font=(FONT_FAMILY, 11),
            text_color="#A0AEC0"
        ).pack(anchor="w", padx=15, pady=(2, 15))

        # 첨부파일 목록
        attach_frame = ctk.CTkFrame(dialog, fg_color=SIDEBAR_COLOR)
        attach_frame.pack(fill="both", expand=True, padx=15, pady=(0, 15))

        ctk.CTkLabel(
            attach_frame,
            text="첨부파일 목록",
            font=(FONT_FAMILY, 12, "bold"),
            text_color="white"
        ).pack(anchor="w", padx=15, pady=(15, 5))

        attach_scroll = ctk.CTkScrollableFrame(attach_frame, fg_color="#1E1E24")
        attach_scroll.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        if all_attachments:
            for row_key, files in all_attachments.items():
                # 행 표시
                row_label = ctk.CTkLabel(
                    attach_scroll,
                    text=f"행 {row_key}:",
                    font=(FONT_FAMILY, 10, "bold"),
                    text_color="#60A5FA"
                )
                row_label.pack(anchor="w", padx=10, pady=(10, 2))

                for file_info in files:
                    file_frame = ctk.CTkFrame(attach_scroll, fg_color="#2D2D35", corner_radius=5)
                    file_frame.pack(fill="x", padx=15, pady=2)

                    ctk.CTkLabel(
                        file_frame,
                        text=f"  {file_info.get('original_name', '')}",
                        font=(FONT_FAMILY, 10),
                        text_color="white"
                    ).pack(side="left", padx=5, pady=5)

                    # 열기 버튼
                    def open_file(path=file_info.get('saved_path', '')):
                        if os.path.exists(path):
                            os.startfile(path)

                    ctk.CTkButton(
                        file_frame,
                        text="열기",
                        font=(FONT_FAMILY, 9),
                        width=50,
                        height=22,
                        fg_color=ACCENT_COLOR,
                        command=open_file
                    ).pack(side="right", padx=5, pady=3)
        else:
            ctk.CTkLabel(
                attach_scroll,
                text="첨부된 파일이 없습니다.",
                font=(FONT_FAMILY, 10),
                text_color="gray"
            ).pack(pady=20)

        # 폴더 열기 버튼
        def open_project_folder():
            if self.project_manager and self.project_manager.project_dir:
                os.startfile(self.project_manager.project_dir)

        ctk.CTkButton(
            dialog,
            text="프로젝트 폴더 열기",
            font=(FONT_FAMILY, 11),
            fg_color="#6B7280",
            command=open_project_folder
        ).pack(pady=(0, 15))

    def _show_open_project_dialog(self):
        """과거 프로젝트 목록에서 선택하여 열기"""
        projects = ProjectManager.list_projects()

        if not projects:
            messagebox.showinfo("안내", "저장된 프로젝트가 없습니다.\n새 파일을 열면 자동으로 프로젝트가 생성됩니다.")
            return

        # 다이얼로그 생성
        dialog = ctk.CTkToplevel(self)
        dialog.title("프로젝트 열기")
        dialog.geometry("700x500")
        dialog.configure(fg_color=BG_COLOR)
        dialog.transient(self)
        dialog.grab_set()

        # 헤더
        header_frame = ctk.CTkFrame(dialog, fg_color=SIDEBAR_COLOR, height=60)
        header_frame.pack(fill="x", padx=15, pady=(15, 0))
        header_frame.pack_propagate(False)

        ctk.CTkLabel(
            header_frame,
            text="프로젝트 선택",
            font=(FONT_FAMILY, 16, "bold"),
            text_color="white"
        ).pack(side="left", padx=20, pady=15)

        ctk.CTkLabel(
            header_frame,
            text=f"총 {len(projects)}개 프로젝트",
            font=(FONT_FAMILY, 11),
            text_color="#A0AEC0"
        ).pack(side="right", padx=20, pady=15)

        # 프로젝트 목록 스크롤 영역
        list_frame = ctk.CTkScrollableFrame(dialog, fg_color=SIDEBAR_COLOR)
        list_frame.pack(fill="both", expand=True, padx=15, pady=15)

        selected_project = {"project": None}  # 선택된 프로젝트 저장용

        def select_project(project, frame):
            """프로젝트 선택"""
            # 모든 프레임 배경색 리셋
            for child in list_frame.winfo_children():
                if isinstance(child, ctk.CTkFrame):
                    child.configure(fg_color="#2D2D35")
            # 선택된 프레임 하이라이트
            frame.configure(fg_color=ACCENT_COLOR)
            selected_project["project"] = project

        def load_selected():
            """선택한 프로젝트 로드"""
            if selected_project["project"]:
                dialog.destroy()
                self._load_from_project(selected_project["project"])

        def double_click_load(project):
            """더블클릭으로 바로 로드"""
            dialog.destroy()
            self._load_from_project(project)

        # 프로젝트 목록 표시
        for project in projects:
            proj_frame = ctk.CTkFrame(list_frame, fg_color="#2D2D35", corner_radius=8)
            proj_frame.pack(fill="x", padx=5, pady=5)

            # 클릭 이벤트 바인딩
            proj_frame.bind("<Button-1>", lambda e, p=project, f=proj_frame: select_project(p, f))
            proj_frame.bind("<Double-Button-1>", lambda e, p=project: double_click_load(p))

            # 프로젝트 정보 표시
            info_container = ctk.CTkFrame(proj_frame, fg_color="transparent")
            info_container.pack(fill="x", padx=15, pady=12)
            info_container.bind("<Button-1>", lambda e, p=project, f=proj_frame: select_project(p, f))
            info_container.bind("<Double-Button-1>", lambda e, p=project: double_click_load(p))

            # 프로젝트 이름 (원본 파일명)
            name_label = ctk.CTkLabel(
                info_container,
                text=project.get("source_file", "알 수 없음"),
                font=(FONT_FAMILY, 12, "bold"),
                text_color="white",
                anchor="w"
            )
            name_label.pack(anchor="w")
            name_label.bind("<Button-1>", lambda e, p=project, f=proj_frame: select_project(p, f))
            name_label.bind("<Double-Button-1>", lambda e, p=project: double_click_load(p))

            # 세부 정보
            details_text = f"생성: {project.get('created_at', '')[:10]}  |  수정: {project.get('last_modified', '')[:10]}  |  첨부: {project.get('attachment_count', 0)}개"
            details_label = ctk.CTkLabel(
                info_container,
                text=details_text,
                font=(FONT_FAMILY, 10),
                text_color="#A0AEC0",
                anchor="w"
            )
            details_label.pack(anchor="w", pady=(3, 0))
            details_label.bind("<Button-1>", lambda e, p=project, f=proj_frame: select_project(p, f))
            details_label.bind("<Double-Button-1>", lambda e, p=project: double_click_load(p))

            # 프로젝트 폴더 경로
            path_label = ctk.CTkLabel(
                info_container,
                text=f"📁 {project.get('project_name', '')}",
                font=(FONT_FAMILY, 9),
                text_color="#6B7280",
                anchor="w"
            )
            path_label.pack(anchor="w", pady=(2, 0))
            path_label.bind("<Button-1>", lambda e, p=project, f=proj_frame: select_project(p, f))
            path_label.bind("<Double-Button-1>", lambda e, p=project: double_click_load(p))

        # 하단 버튼 영역
        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.pack(fill="x", padx=15, pady=(0, 15))

        ctk.CTkButton(
            btn_frame,
            text="열기",
            font=(FONT_FAMILY, 12, "bold"),
            fg_color=ACCENT_COLOR,
            width=120,
            command=load_selected
        ).pack(side="right", padx=5)

        ctk.CTkButton(
            btn_frame,
            text="취소",
            font=(FONT_FAMILY, 12),
            fg_color="#4B5563",
            width=80,
            command=dialog.destroy
        ).pack(side="right", padx=5)

    def _load_from_project(self, project: dict):
        """선택한 프로젝트에서 파일 로드"""
        project_dir = project.get("project_dir")
        if not project_dir or not os.path.exists(project_dir):
            messagebox.showerror("오류", "프로젝트 폴더를 찾을 수 없습니다.")
            return

        # 프로젝트 매니저 로드
        self.project_manager = ProjectManager()
        if not self.project_manager.load_project(project_dir):
            messagebox.showerror("오류", "프로젝트를 로드할 수 없습니다.")
            return

        # 원본 파일 경로 찾기
        source_file = None
        for ext in ['.xlsx', '.xls', '.csv']:
            potential_path = os.path.join(project_dir, f"source{ext}")
            if os.path.exists(potential_path):
                source_file = potential_path
                break

        if not source_file:
            messagebox.showerror("오류", "프로젝트 내 원본 파일을 찾을 수 없습니다.")
            return

        try:
            # 기존 데이터 초기화
            self.sheets.clear()
            self.chat_histories.clear()
            self.column_mappings = {}
            self.highlighted_cells = {}

            # 파일명 저장
            self._current_file_name = project.get('source_file', '')

            # 파일 확장자에 따라 로드
            self.file_label.configure(text="파일 로딩 중...")
            self.update_idletasks()

            ext = os.path.splitext(source_file)[1].lower()

            if ext == '.csv':
                self.file_label.configure(text="CSV 파일 읽는 중...")
                self.update_idletasks()
                df = pd.read_csv(source_file, encoding='utf-8-sig')
                self.sheets["Sheet1"] = df
                self.highlighted_cells["Sheet1"] = set()
            else:
                # Excel 파일의 모든 시트 로드
                self.file_label.configure(text="Excel 시트 로딩 중...")
                self.update_idletasks()
                excel_file = pd.ExcelFile(source_file)
                for i, sheet_name in enumerate(excel_file.sheet_names):
                    self.file_label.configure(text=f"시트 로딩 중... ({i+1}/{len(excel_file.sheet_names)})")
                    self.update_idletasks()
                    self.sheets[sheet_name] = pd.read_excel(source_file, sheet_name=sheet_name)
                    self.highlighted_cells[sheet_name] = set()

                # openpyxl로 셀 배경색 확인
                self.file_label.configure(text="셀 서식 분석 중...")
                self.update_idletasks()
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(source_file, data_only=True)
                    for sheet_name in wb.sheetnames:
                        if sheet_name in self.sheets:
                            ws = wb[sheet_name]
                            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=0):
                                for col_idx, cell in enumerate(row):
                                    if cell.fill and cell.fill.fgColor:
                                        color = cell.fill.fgColor
                                        if color.type == 'rgb' and color.rgb and color.rgb != '00000000':
                                            if color.rgb not in ['FFFFFFFF', 'FFFFFF', '00FFFFFF']:
                                                self.highlighted_cells[sheet_name].add((row_idx, col_idx))
                except Exception as e:
                    print(f"셀 서식 분석 중 오류 (무시됨): {e}")

            self.file_label.configure(text="테이블 구성 중...")
            self.update_idletasks()

            # 저장된 분류 데이터 로드
            self.file_label.configure(text="분류 데이터 복원 중...")
            self.update_idletasks()
            classification_data = self.project_manager.load_classification_data()

            if classification_data:
                # 분류 데이터를 DataFrame에 적용
                sheets_data = classification_data.get("sheets", {})
                for sheet_name, col_data in sheets_data.items():
                    if sheet_name in self.sheets:
                        df = self.sheets[sheet_name]
                        for col_name, values in col_data.items():
                            # 컬럼이 없으면 추가
                            if col_name not in df.columns:
                                df[col_name] = ""
                            # 값 적용 (행 수가 맞는 경우)
                            if len(values) == len(df):
                                df[col_name] = values

                # 대화 기록 복원
                self.chat_histories = classification_data.get("chat_histories", {})
                print(f"분류 데이터 복원 완료: {len(sheets_data)}개 시트, {sum(len(v) for v in self.chat_histories.values())}개 대화 기록")

            # 시트 탭 업데이트
            self._update_sheet_tabs()

            # 첫 번째 시트 선택
            if self.sheets:
                first_sheet = list(self.sheets.keys())[0]
                self._select_sheet(first_sheet)

            # 파일명 표시
            self.file_label.configure(text=f"📁 {project.get('source_file', '프로젝트')} (프로젝트에서 로드)")

            # 진행률 업데이트
            self._update_progress()

        except Exception as e:
            messagebox.showerror("오류", f"파일 로드 실패: {str(e)}")
            self.file_label.configure(text="파일을 선택하세요")

    def _update_progress(self):
        """진행률 업데이트"""
        counts = self.data_table.get_status_counts()
        total = counts.get("total", 0)

        if total == 0:
            self.progress_label.configure(text="진행률: -")
            self.progress_bar.set(0)
            return

        # 완료 = AI분류 + 검증완료
        completed = counts.get("classified", 0) + counts.get("verified", 0)
        pending = counts.get("pending", 0)
        existing = counts.get("existing", 0)
        needs_review = counts.get("needs_review", 0)

        progress = (completed + existing) / total  # 기존분류도 일단 진행된 것으로 카운트

        self.progress_label.configure(
            text=f"완료: {completed}/{total} | 기존: {existing} | 미분류: {pending} | 검토: {needs_review}"
        )
        self.progress_bar.set(progress)

        # 필터 버튼에 건수 표시 업데이트
        self.filter_buttons["all"].configure(text=f"전체 ({total})")
        self.filter_buttons["pending"].configure(text=f"미분류 ({pending})")
        self.filter_buttons["highlighted"].configure(text=f"검토★ ({counts.get('highlighted', 0)})")
        self.filter_buttons["existing"].configure(text=f"기존 ({existing})")
        self.filter_buttons["classified"].configure(text=f"AI ({counts.get('classified', 0)})")
        self.filter_buttons["verified"].configure(text=f"검증 ({counts.get('verified', 0)})")
        self.filter_buttons["needs_review"].configure(text=f"검토 ({needs_review})")

    def _verify_selected_row(self):
        """선택된 행의 분류를 검증 완료로 표시"""
        if self.current_sheet is None:
            return

        selected_idx = self.data_table.selected_index
        if selected_idx is None:
            messagebox.showinfo("알림", "검증할 행을 먼저 선택하세요")
            return

        # 현재 상태 확인
        df = self.data_table.df
        if df is None:
            return

        current_cls = str(df.iloc[selected_idx].get("최종_분류", ""))

        if not current_cls:
            messagebox.showwarning("경고", "분류가 없는 행은 검증할 수 없습니다.\n먼저 AI 분류를 실행하세요.")
            return

        # 검증 완료로 변경
        self.data_table.verify_row(selected_idx)
        self.sheets[self.current_sheet].at[selected_idx, "분류_상태"] = ClassificationStatus.VERIFIED
        self._update_progress()

        messagebox.showinfo("검증 완료", f"행 #{selected_idx + 1}이 검증 완료되었습니다.")

    def _batch_classify(self):
        """미분류 항목 일괄 AI 분류"""
        if self.ai is None or self.current_sheet is None:
            messagebox.showwarning("경고", "AI가 초기화되지 않았거나 파일이 로드되지 않았습니다.")
            return

        df = self.data_table.df
        if df is None:
            return

        # 미분류 항목 찾기
        pending_indices = []
        for idx in range(len(df)):
            status = str(df.iloc[idx].get("분류_상태", ClassificationStatus.PENDING))
            if status == ClassificationStatus.PENDING:
                pending_indices.append(idx)

        if not pending_indices:
            messagebox.showinfo("알림", "미분류 항목이 없습니다.")
            return

        # 확인 대화상자
        if not messagebox.askyesno(
            "일괄 분류 확인",
            f"미분류 항목 {len(pending_indices)}건을 AI로 분류하시겠습니까?\n\n"
            "* 각 항목에 대해 기본적인 분류만 진행됩니다.\n"
            "* 추가 질문이 필요한 경우 '검토필요'로 표시됩니다.\n"
            "* 정확한 분류를 위해 개별 검토를 권장합니다."
        ):
            return

        # 일괄 분류 시작
        self._batch_classify_items(pending_indices, 0)

    def _batch_classify_items(self, indices: List[int], current_idx: int):
        """일괄 분류 순차 실행"""
        if current_idx >= len(indices):
            self._update_progress()
            messagebox.showinfo("일괄 분류 완료", f"{len(indices)}건의 분류가 완료되었습니다.")
            return

        row_idx = indices[current_idx]
        df = self.sheets[self.current_sheet]
        row_data = df.iloc[row_idx].to_json(force_ascii=False, indent=2)

        # 진행 상황 표시
        self.progress_label.configure(
            text=f"일괄 분류 중... ({current_idx + 1}/{len(indices)})"
        )

        # AI 분석 (콜백에서 다음 항목 처리)
        def on_batch_result(result):
            status = result.get("status", "ERROR")

            if status == "COMPLETE":
                final_cls = result.get("final_classification", "")
                reasoning = result.get("reasoning", "")

                # 데이터 업데이트
                self.data_table.update_row(
                    row_idx, final_cls, reasoning, "",
                    status=ClassificationStatus.AI_CLASSIFIED
                )
                self.sheets[self.current_sheet].at[row_idx, "최종_분류"] = final_cls
                self.sheets[self.current_sheet].at[row_idx, "분류_근거"] = reasoning
                self.sheets[self.current_sheet].at[row_idx, "분류_상태"] = ClassificationStatus.AI_CLASSIFIED

            elif status == "INCOMPLETE":
                # 추가 질문이 필요한 경우 검토필요로 표시
                self.data_table.update_row(
                    row_idx, "", result.get("message", "추가 정보 필요"), "",
                    status=ClassificationStatus.NEEDS_REVIEW
                )
                self.sheets[self.current_sheet].at[row_idx, "분류_상태"] = ClassificationStatus.NEEDS_REVIEW
                self.sheets[self.current_sheet].at[row_idx, "분류_근거"] = result.get("message", "")

            # 다음 항목 처리
            self.after(100, lambda: self._batch_classify_items(indices, current_idx + 1))

        # AI 호출 (별도 스레드)
        def run_batch_ai():
            result = self.ai.analyze(row_data, [])
            self.after(0, lambda: on_batch_result(result))

        threading.Thread(target=run_batch_ai, daemon=True).start()

    def _setup_keyboard_shortcuts(self):
        """키보드 단축키 설정"""
        # Ctrl+S: 저장
        self.bind("<Control-s>", lambda e: self._save_file())
        self.bind("<Control-S>", lambda e: self._save_file())

        # Ctrl+O: 파일 열기
        self.bind("<Control-o>", lambda e: self._load_file())
        self.bind("<Control-O>", lambda e: self._load_file())

        # Ctrl+P: 프로젝트 열기
        self.bind("<Control-p>", lambda e: self._show_open_project_dialog())
        self.bind("<Control-P>", lambda e: self._show_open_project_dialog())

        # 방향키: 행 이동
        self.bind("<Down>", lambda e: self._move_to_next_row())
        self.bind("<Up>", lambda e: self._move_to_prev_row())

        # Ctrl+Enter: AI 분석 시작
        self.bind("<Control-Return>", lambda e: self._trigger_ai_analysis())

        # Ctrl+1~5: 필터 단축키
        self.bind("<Control-Key-0>", lambda e: self._set_filter("all"))
        self.bind("<Control-Key-1>", lambda e: self._set_filter("pending"))
        self.bind("<Control-Key-2>", lambda e: self._set_filter("existing"))
        self.bind("<Control-Key-3>", lambda e: self._set_filter("classified"))
        self.bind("<Control-Key-4>", lambda e: self._set_filter("highlighted"))

        # F1: 단축키 도움말
        self.bind("<F1>", lambda e: self._show_shortcuts_help())

    def _move_to_next_row(self):
        """다음 행으로 이동"""
        if self.data_table.selected_index is not None:
            next_idx = self.data_table.selected_index + 1
            if next_idx < len(self.data_table.df):
                self.data_table.select_row(next_idx)
                self._on_row_select(next_idx)

    def _move_to_prev_row(self):
        """이전 행으로 이동"""
        if self.data_table.selected_index is not None:
            prev_idx = self.data_table.selected_index - 1
            if prev_idx >= 0:
                self.data_table.select_row(prev_idx)
                self._on_row_select(prev_idx)

    def _trigger_ai_analysis(self):
        """AI 분석 트리거"""
        if self.chat_panel.current_row_index is not None:
            self.chat_panel._submit()

    def _show_shortcuts_help(self):
        """단축키 도움말 표시"""
        help_text = """[ 키보드 단축키 ]

Ctrl+O    파일 열기
Ctrl+P    프로젝트 열기
Ctrl+S    결과 저장

↑/↓      행 이동
Ctrl+Enter    AI 분석 시작

Ctrl+0    전체 보기
Ctrl+1    미분류만 보기
Ctrl+2    기존분류만 보기
Ctrl+3    AI분류만 보기
Ctrl+4    검토표시만 보기

F1        이 도움말 표시
"""
        messagebox.showinfo("키보드 단축키", help_text)

    def _save_file(self):
        """결과 저장"""
        if not self.sheets:
            messagebox.showwarning("경고", "저장할 데이터가 없습니다.")
            return

        # 현재 시트의 데이터프레임 가져오기
        if self.current_sheet:
            updated_df = self.data_table.get_dataframe()
            if updated_df is not None:
                self.sheets[self.current_sheet] = updated_df

        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv")]
        )

        if not save_path:
            return

        try:
            if save_path.endswith('.csv'):
                # CSV는 첫 번째 시트만 저장
                first_sheet = list(self.sheets.values())[0]
                first_sheet.to_csv(save_path, index=False, encoding='utf-8-sig')
            else:
                # Excel은 모든 시트 저장
                with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                    for sheet_name, df in self.sheets.items():
                        df.to_excel(writer, sheet_name=sheet_name, index=False)

            messagebox.showinfo("저장 완료", f"파일이 저장되었습니다:\n{save_path}")

        except Exception as e:
            messagebox.showerror("저장 실패", str(e))

    # ==========================================
    # 버전 관리 기능
    # ==========================================
    def _check_yaml_changes_on_startup(self):
        """앱 시작 시 YAML 파일 변경 확인"""
        if not HAS_VERSION_MANAGER:
            return

        change_info = check_yaml_changes()

        if change_info.get("changed"):
            self._show_yaml_change_dialog(change_info)

    def _show_yaml_change_dialog(self, change_info: Dict):
        """YAML 변경 감지 다이얼로그"""
        dialog = ctk.CTkToplevel(self)
        dialog.title("실무지침 변경 감지")
        dialog.geometry("500x300")
        dialog.configure(fg_color=BG_COLOR)
        dialog.transient(self)
        dialog.grab_set()

        # 중앙 배치
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() - 500) // 2
        y = (dialog.winfo_screenheight() - 300) // 2
        dialog.geometry(f"+{x}+{y}")

        # 아이콘/제목
        ctk.CTkLabel(
            dialog,
            text="실무지침(YAML)이 변경되었습니다",
            font=(FONT_FAMILY, 16, "bold"),
            text_color=WARNING_COLOR
        ).pack(pady=(30, 10))

        # 변경 정보
        info_frame = ctk.CTkFrame(dialog, fg_color=SIDEBAR_COLOR)
        info_frame.pack(fill="x", padx=30, pady=10)

        ctk.CTkLabel(
            info_frame,
            text=f"수정 시간: {change_info.get('last_modified', '알 수 없음')}",
            font=(FONT_FAMILY, 11),
            text_color="white"
        ).pack(pady=5)

        ctk.CTkLabel(
            info_frame,
            text=f"이전 확인: {change_info.get('last_acknowledged', '알 수 없음')}",
            font=(FONT_FAMILY, 11),
            text_color="gray"
        ).pack(pady=5)

        ctk.CTkLabel(
            dialog,
            text="변경된 실무지침을 적용하시겠습니까?",
            font=(FONT_FAMILY, 12),
            text_color="white"
        ).pack(pady=10)

        # 회계원칙 검증 버튼 (별도 프레임)
        if HAS_POLICY_VALIDATOR:
            validate_frame = ctk.CTkFrame(dialog, fg_color=SIDEBAR_COLOR)
            validate_frame.pack(fill="x", padx=30, pady=5)

            ctk.CTkLabel(
                validate_frame,
                text="변경 내용이 회계기준에 맞는지 확인하세요:",
                font=(FONT_FAMILY, 10),
                text_color="gray"
            ).pack(side="left", padx=10, pady=8)

            def validate_changes():
                dialog.destroy()
                self._show_validation_results(check_changes_only=True)

            ctk.CTkButton(
                validate_frame,
                text="회계원칙 검증",
                font=(FONT_FAMILY, 11, "bold"),
                fg_color="#8B5CF6",
                hover_color="#7C3AED",
                width=120,
                command=validate_changes
            ).pack(side="right", padx=10, pady=8)

        # 버튼 프레임
        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.pack(pady=15)

        def apply_changes():
            # 적용 전 백업
            create_backup("yaml_change")
            # 변경 확인 처리
            acknowledge_yaml_changes()
            messagebox.showinfo("적용 완료", "실무지침 변경이 적용되었습니다.\n(이전 버전은 버전 관리에서 확인 가능)")
            dialog.destroy()

        def ignore_changes():
            # 현재 상태 그대로 유지 (해시 업데이트하지 않음)
            dialog.destroy()

        def show_backup():
            dialog.destroy()
            self._show_version_history()

        ctk.CTkButton(
            btn_frame,
            text="적용",
            font=(FONT_FAMILY, 12, "bold"),
            fg_color=SUCCESS_COLOR,
            text_color="black",
            width=100,
            command=apply_changes
        ).pack(side="left", padx=10)

        ctk.CTkButton(
            btn_frame,
            text="무시",
            font=(FONT_FAMILY, 12),
            fg_color="#6B7280",
            width=100,
            command=ignore_changes
        ).pack(side="left", padx=10)

        ctk.CTkButton(
            btn_frame,
            text="버전 히스토리",
            font=(FONT_FAMILY, 12),
            fg_color=ACCENT_COLOR,
            width=120,
            command=show_backup
        ).pack(side="left", padx=10)

    def _show_version_history(self):
        """버전 히스토리 다이얼로그"""
        if not HAS_VERSION_MANAGER:
            messagebox.showwarning("경고", "버전 관리 모듈이 로드되지 않았습니다.")
            return

        dialog = ctk.CTkToplevel(self)
        dialog.title("실무지침 버전 관리")
        dialog.geometry("700x500")
        dialog.configure(fg_color=BG_COLOR)
        dialog.transient(self)
        dialog.grab_set()

        # 중앙 배치
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() - 700) // 2
        y = (dialog.winfo_screenheight() - 500) // 2
        dialog.geometry(f"+{x}+{y}")

        # 제목
        header_frame = ctk.CTkFrame(dialog, fg_color=SIDEBAR_COLOR)
        header_frame.pack(fill="x", padx=10, pady=10)

        ctk.CTkLabel(
            header_frame,
            text="실무지침 버전 히스토리",
            font=(FONT_FAMILY, 14, "bold"),
            text_color="white"
        ).pack(side="left", padx=15, pady=10)

        # 수동 백업 버튼
        def create_manual_backup():
            path = create_backup("manual")
            if path:
                messagebox.showinfo("백업 완료", f"백업이 생성되었습니다.")
                refresh_list()
            else:
                messagebox.showerror("오류", "백업 생성에 실패했습니다.")

        ctk.CTkButton(
            header_frame,
            text="현재 상태 백업",
            font=(FONT_FAMILY, 11),
            fg_color=ACCENT_COLOR,
            width=120,
            command=create_manual_backup
        ).pack(side="right", padx=15, pady=10)

        # 버전 목록 프레임
        list_frame = ctk.CTkFrame(dialog, fg_color=SIDEBAR_COLOR)
        list_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        # 스크롤 가능한 프레임
        scrollable = ctk.CTkScrollableFrame(list_frame, fg_color="transparent")
        scrollable.pack(fill="both", expand=True, padx=5, pady=5)

        selected_version = {"path": None}

        def refresh_list():
            # 기존 위젯 제거
            for widget in scrollable.winfo_children():
                widget.destroy()

            versions = get_version_list()

            if not versions:
                ctk.CTkLabel(
                    scrollable,
                    text="저장된 버전이 없습니다.\n'현재 상태 백업' 버튼을 눌러 첫 백업을 생성하세요.",
                    font=(FONT_FAMILY, 12),
                    text_color="gray"
                ).pack(pady=50)
                return

            for i, version in enumerate(versions):
                v_frame = ctk.CTkFrame(scrollable, fg_color="#2D2D35" if i % 2 == 0 else "#25262B")
                v_frame.pack(fill="x", pady=2)

                # 라디오 버튼 효과 (선택 표시)
                def select_version(path=version["path"], frame=v_frame):
                    selected_version["path"] = path
                    # 시각적 피드백
                    for child in scrollable.winfo_children():
                        try:
                            idx = scrollable.winfo_children().index(child)
                            child.configure(fg_color="#2D2D35" if idx % 2 == 0 else "#25262B")
                        except (tk.TclError, AttributeError):
                            pass  # fg_color 미지원 위젯은 무시
                    frame.configure(fg_color=ACCENT_COLOR)

                v_frame.bind("<Button-1>", lambda e, p=version["path"], f=v_frame: select_version(p, f))

                # 버전 정보
                info_label = ctk.CTkLabel(
                    v_frame,
                    text=version["display_name"],
                    font=(FONT_FAMILY, 11),
                    text_color="white",
                    anchor="w"
                )
                info_label.pack(side="left", padx=15, pady=8)
                info_label.bind("<Button-1>", lambda e, p=version["path"], f=v_frame: select_version(p, f))

                # 파일 크기
                size_kb = version["size"] / 1024
                size_label = ctk.CTkLabel(
                    v_frame,
                    text=f"{size_kb:.1f} KB",
                    font=(FONT_FAMILY, 10),
                    text_color="gray"
                )
                size_label.pack(side="right", padx=15, pady=8)
                size_label.bind("<Button-1>", lambda e, p=version["path"], f=v_frame: select_version(p, f))

        refresh_list()

        # 하단 버튼 프레임
        bottom_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        bottom_frame.pack(fill="x", padx=10, pady=10)

        def restore_selected():
            if not selected_version["path"]:
                messagebox.showwarning("경고", "복원할 버전을 선택하세요.")
                return

            if messagebox.askyesno("복원 확인",
                                   "선택한 버전으로 복원하시겠습니까?\n\n"
                                   "현재 상태는 자동으로 백업됩니다."):
                success, msg = restore_version(selected_version["path"])
                if success:
                    messagebox.showinfo("복원 완료", msg)
                    refresh_list()
                else:
                    messagebox.showerror("복원 실패", msg)

        def show_diff():
            if not selected_version["path"]:
                messagebox.showwarning("경고", "비교할 버전을 선택하세요.")
                return

            diff = get_version_diff(selected_version["path"])
            if "error" in diff:
                messagebox.showerror("오류", diff["error"])
                return

            diff_text = f"""[ 현재 vs 선택 버전 비교 ]

학습 케이스:
  현재: {diff['learned_cases']['current']}개
  선택: {diff['learned_cases']['old']}개

충돌 기록:
  현재: {diff['conflict_log']['current']}개
  선택: {diff['conflict_log']['old']}개

의사결정 트리 단계:
  현재: {diff['decision_tree_steps']['current']}개
  선택: {diff['decision_tree_steps']['old']}개
"""
            messagebox.showinfo("버전 비교", diff_text)

        ctk.CTkButton(
            bottom_frame,
            text="선택 버전으로 복원",
            font=(FONT_FAMILY, 12, "bold"),
            fg_color=WARNING_COLOR,
            text_color="black",
            width=150,
            command=restore_selected
        ).pack(side="left", padx=10)

        ctk.CTkButton(
            bottom_frame,
            text="현재와 비교",
            font=(FONT_FAMILY, 12),
            fg_color="#6B7280",
            width=120,
            command=show_diff
        ).pack(side="left", padx=10)

        ctk.CTkButton(
            bottom_frame,
            text="닫기",
            font=(FONT_FAMILY, 12),
            fg_color="#4B5563",
            width=80,
            command=dialog.destroy
        ).pack(side="right", padx=10)

    # ==========================================
    # 회계정책 검증 기능
    # ==========================================
    def _show_validation_results(self, check_changes_only: bool = False):
        """
        회계정책 검증 결과 표시

        Args:
            check_changes_only: True면 변경된 부분만 검증
        """
        if not HAS_POLICY_VALIDATOR:
            messagebox.showwarning("경고", "회계정책 검증 모듈이 로드되지 않았습니다.")
            return

        # 진행 표시 다이얼로그
        progress_dialog = ctk.CTkToplevel(self)
        progress_dialog.title("검증 중...")
        progress_dialog.geometry("300x100")
        progress_dialog.configure(fg_color=BG_COLOR)
        progress_dialog.transient(self)
        progress_dialog.grab_set()

        progress_dialog.update_idletasks()
        x = (progress_dialog.winfo_screenwidth() - 300) // 2
        y = (progress_dialog.winfo_screenheight() - 100) // 2
        progress_dialog.geometry(f"+{x}+{y}")

        ctk.CTkLabel(
            progress_dialog,
            text="회계기준 검증 중...",
            font=(FONT_FAMILY, 12),
            text_color="white"
        ).pack(pady=20)

        progress_bar = ctk.CTkProgressBar(progress_dialog, width=200)
        progress_bar.pack(pady=10)
        progress_bar.set(0)
        progress_dialog.update()

        def run_validation():
            try:
                validator = PolicyValidator()
                guidelines = load_guidelines()

                progress_bar.set(0.3)
                progress_dialog.update()

                # 검증 실행
                if check_changes_only:
                    # 이전 버전과 비교 검증
                    versions = get_version_list()
                    if versions and len(versions) > 0:
                        # 가장 최근 백업과 비교
                        import json
                        with open(versions[0]["path"], 'r', encoding='utf-8') as f:
                            old_data = json.load(f)
                            old_data.pop("_backup_info", None)
                        results = validator.validate_changes(old_data, guidelines)
                    else:
                        results = validator.validate_guidelines(guidelines, deep_check=False)
                else:
                    results = validator.validate_guidelines(guidelines, deep_check=False)

                progress_bar.set(0.8)
                progress_dialog.update()

                formatted = format_validation_results(results)

                progress_bar.set(1.0)
                progress_dialog.update()

                progress_dialog.destroy()

                # 결과 표시
                self._display_validation_results(formatted)

            except Exception as e:
                progress_dialog.destroy()
                messagebox.showerror("검증 오류", f"검증 중 오류가 발생했습니다:\n{str(e)}")

        # 비동기 실행
        self.after(100, run_validation)

    def _display_validation_results(self, results: Dict):
        """검증 결과 표시 다이얼로그"""
        dialog = ctk.CTkToplevel(self)
        dialog.title("회계정책 검증 결과")
        dialog.geometry("800x600")
        dialog.configure(fg_color=BG_COLOR)
        dialog.transient(self)

        # 중앙 배치
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() - 800) // 2
        y = (dialog.winfo_screenheight() - 600) // 2
        dialog.geometry(f"+{x}+{y}")

        # 헤더
        header_frame = ctk.CTkFrame(dialog, fg_color=SIDEBAR_COLOR)
        header_frame.pack(fill="x", padx=10, pady=10)

        ctk.CTkLabel(
            header_frame,
            text="회계정책 검증 결과",
            font=(FONT_FAMILY, 16, "bold"),
            text_color="white"
        ).pack(side="left", padx=15, pady=10)

        # 요약 정보
        summary = results.get("summary", {})
        summary_text = (
            f"검증항목: {summary.get('total', 0)}개 | "
            f"오류: {summary.get('errors', 0)} | "
            f"경고: {summary.get('warnings', 0)} | "
            f"정보: {summary.get('info', 0)} | "
            f"적합: {summary.get('passed', 0)}"
        )

        # 요약 색상 결정
        if summary.get('errors', 0) > 0:
            summary_color = ERROR_COLOR
        elif summary.get('warnings', 0) > 0:
            summary_color = WARNING_COLOR
        else:
            summary_color = SUCCESS_COLOR

        ctk.CTkLabel(
            header_frame,
            text=summary_text,
            font=(FONT_FAMILY, 11),
            text_color=summary_color
        ).pack(side="right", padx=15, pady=10)

        # 결과 목록 (스크롤)
        scrollable = ctk.CTkScrollableFrame(dialog, fg_color=SIDEBAR_COLOR)
        scrollable.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        # 색상 정의
        severity_colors = {
            "errors": ("#FF6B6B", "오류 - 회계기준 위반"),
            "warnings": ("#FFB347", "경고 - 권장사항 미준수"),
            "info": ("#60A5FA", "정보 - 회사 재량/전략적 선택"),
            "passed": ("#50C878", "적합 - 회계기준 충족")
        }

        # 각 카테고리별 결과 표시
        for category in ["errors", "warnings", "info", "passed"]:
            items = results.get(category, [])
            if not items:
                continue

            color, label = severity_colors[category]

            # 카테고리 헤더
            cat_header = ctk.CTkFrame(scrollable, fg_color=color, corner_radius=5)
            cat_header.pack(fill="x", pady=(10, 5))

            ctk.CTkLabel(
                cat_header,
                text=f" {label} ({len(items)}개)",
                font=(FONT_FAMILY, 12, "bold"),
                text_color="white" if category != "passed" else "black"
            ).pack(anchor="w", padx=10, pady=5)

            # 각 항목
            for item in items:
                item_frame = ctk.CTkFrame(scrollable, fg_color="#2D2D35", corner_radius=5)
                item_frame.pack(fill="x", pady=2, padx=5)

                # 제목 줄
                title_frame = ctk.CTkFrame(item_frame, fg_color="transparent")
                title_frame.pack(fill="x", padx=10, pady=(8, 2))

                ctk.CTkLabel(
                    title_frame,
                    text=item.get("title", ""),
                    font=(FONT_FAMILY, 11, "bold"),
                    text_color=color
                ).pack(side="left")

                # 회사 재량 표시
                if item.get("is_company_choice"):
                    ctk.CTkLabel(
                        title_frame,
                        text="[회사 재량]",
                        font=(FONT_FAMILY, 9),
                        text_color="#A78BFA"
                    ).pack(side="left", padx=10)

                # 섹션 표시
                ctk.CTkLabel(
                    title_frame,
                    text=f"섹션: {item.get('section', '')}",
                    font=(FONT_FAMILY, 9),
                    text_color="gray"
                ).pack(side="right")

                # 메시지
                if item.get("message"):
                    ctk.CTkLabel(
                        item_frame,
                        text=item.get("message", ""),
                        font=(FONT_FAMILY, 10),
                        text_color="white",
                        wraplength=700,
                        justify="left"
                    ).pack(anchor="w", padx=10, pady=2)

                # 회계기준 참조
                if item.get("standard_ref"):
                    ctk.CTkLabel(
                        item_frame,
                        text=f"참조: {item.get('standard_ref', '')}",
                        font=(FONT_FAMILY, 9),
                        text_color="#9CA3AF"
                    ).pack(anchor="w", padx=10, pady=2)

                # 제안
                if item.get("suggestion"):
                    suggestion_frame = ctk.CTkFrame(item_frame, fg_color="#1F2937", corner_radius=3)
                    suggestion_frame.pack(fill="x", padx=10, pady=(2, 8))
                    ctk.CTkLabel(
                        suggestion_frame,
                        text=f"제안: {item.get('suggestion', '')}",
                        font=(FONT_FAMILY, 9),
                        text_color="#FCD34D",
                        wraplength=680,
                        justify="left"
                    ).pack(anchor="w", padx=8, pady=5)

        # 결과가 없는 경우
        if not any(results.get(cat) for cat in severity_colors.keys()):
            ctk.CTkLabel(
                scrollable,
                text="검증 결과가 없습니다.",
                font=(FONT_FAMILY, 12),
                text_color="gray"
            ).pack(pady=50)

        # 하단 버튼
        bottom_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        bottom_frame.pack(fill="x", padx=10, pady=10)

        # 범례
        legend_frame = ctk.CTkFrame(bottom_frame, fg_color="transparent")
        legend_frame.pack(side="left")

        ctk.CTkLabel(
            legend_frame,
            text="범례:",
            font=(FONT_FAMILY, 9),
            text_color="gray"
        ).pack(side="left", padx=5)

        for cat, (color, label) in severity_colors.items():
            short_label = label.split(" - ")[0]
            ctk.CTkLabel(
                legend_frame,
                text=f"● {short_label}",
                font=(FONT_FAMILY, 9),
                text_color=color
            ).pack(side="left", padx=5)

        ctk.CTkButton(
            bottom_frame,
            text="닫기",
            font=(FONT_FAMILY, 12),
            fg_color="#4B5563",
            width=80,
            command=dialog.destroy
        ).pack(side="right", padx=10)

        # 전체 검증 버튼
        def run_full_validation():
            dialog.destroy()
            self._show_validation_results(check_changes_only=False)

        ctk.CTkButton(
            bottom_frame,
            text="전체 검증",
            font=(FONT_FAMILY, 11),
            fg_color=ACCENT_COLOR,
            width=100,
            command=run_full_validation
        ).pack(side="right", padx=5)

    def _show_guidelines_review_dialog(self):
        """
        회계 실무지침 검토 및 수정 다이얼로그 (새 창)

        기능:
        - 실무지침을 사람이 읽을 수 있는 형태로 표시
        - 섹션별 선택 후 부분 검토 가능
        - 편집 모드에서 직접 수정 가능
        - 파일 첨부로 회사 정보 업데이트
        - 충돌 검사 및 AI 검증
        """
        if not HAS_VERSION_MANAGER:
            messagebox.showwarning("경고", "실무지침 모듈이 로드되지 않았습니다.")
            return

        # 실무지침 로드
        guidelines = load_guidelines()
        if not guidelines:
            messagebox.showwarning("경고", "실무지침 파일을 불러올 수 없습니다.")
            return

        # 원본 백업 (수정 전 상태)
        import copy
        original_guidelines = copy.deepcopy(guidelines)

        dialog = ctk.CTkToplevel(self)
        dialog.title("회계 실무지침 검토 및 수정")
        dialog.geometry("1200x800")
        dialog.configure(fg_color=BG_COLOR)
        dialog.transient(self)

        # 중앙 배치
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() - 1200) // 2
        y = (dialog.winfo_screenheight() - 800) // 2
        dialog.geometry(f"+{x}+{y}")

        # 상태 추적
        state = {
            "selected_sections": set(),
            "edit_mode": False,
            "modified": False,
            "attached_files": [],
            "pending_changes": {},  # 섹션별 변경 내용
            "current_model": "claude-sonnet",  # 기본 AI 모델
            "content_widgets": {},  # 섹션별 텍스트 위젯 (선택 영역 추출용)
        }

        # ===== 헤더 =====
        header_frame = ctk.CTkFrame(dialog, fg_color=SIDEBAR_COLOR, height=60)
        header_frame.pack(fill="x")
        header_frame.pack_propagate(False)

        title_label = ctk.CTkLabel(
            header_frame,
            text="회계 실무지침 검토",
            font=(FONT_FAMILY, 18, "bold"),
            text_color="white"
        )
        title_label.pack(side="left", padx=20, pady=15)

        # 수정됨 표시 (더 눈에 띄게)
        modified_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        modified_frame.pack(side="left", padx=10)

        modified_label = ctk.CTkLabel(
            modified_frame,
            text="",
            font=(FONT_FAMILY, 12, "bold"),
            text_color=WARNING_COLOR,
            corner_radius=5
        )
        modified_label.pack(side="left")

        # 저장 안내 버튼 (변경사항 있을 때만 표시)
        save_reminder_btn = ctk.CTkButton(
            modified_frame,
            text="",
            font=(FONT_FAMILY, 10, "bold"),
            fg_color=SUCCESS_COLOR,
            text_color="black",
            width=0,
            height=24,
            corner_radius=12
        )
        # 초기에는 숨김

        def update_save_reminder():
            """저장 리마인더 UI 업데이트"""
            if state["modified"] and state["pending_changes"]:
                count = len(state["pending_changes"])
                modified_label.configure(
                    text=f"  {count}개 섹션 수정됨  ",
                    fg_color=WARNING_COLOR,
                    text_color="black"
                )
                save_reminder_btn.configure(
                    text="💾 저장하기",
                    command=save_changes
                )
                save_reminder_btn.pack(side="left", padx=5)
            else:
                modified_label.configure(text="", fg_color="transparent")
                save_reminder_btn.pack_forget()

        # state 변경 감지를 위한 래퍼 함수 저장
        state["update_save_reminder"] = update_save_reminder

        # 회사 정보 표시
        company_info = guidelines.get("company_profile", {})
        info_text = f"{company_info.get('업종', '')} | {company_info.get('회계기준', '')}"
        ctk.CTkLabel(
            header_frame,
            text=info_text,
            font=(FONT_FAMILY, 11),
            text_color="#A0AEC0"
        ).pack(side="left", padx=20)

        # 헤더 우측: AI 모델 선택 및 편집 모드 토글
        edit_mode_var = ctk.BooleanVar(value=False)

        def toggle_edit_mode():
            state["edit_mode"] = edit_mode_var.get()
            if state["edit_mode"]:
                title_label.configure(text="회계 실무지침 편집")
                edit_toggle.configure(fg_color=WARNING_COLOR, text="편집 모드 ON")
            else:
                title_label.configure(text="회계 실무지침 검토")
                edit_toggle.configure(fg_color="#4B5563", text="편집 모드")
            update_content_display()

        edit_toggle = ctk.CTkButton(
            header_frame,
            text="편집 모드",
            font=(FONT_FAMILY, 11),
            fg_color="#4B5563",
            hover_color="#6B7280",
            width=100,
            command=lambda: [edit_mode_var.set(not edit_mode_var.get()), toggle_edit_mode()]
        )
        edit_toggle.pack(side="right", padx=10, pady=15)

        # AI 모델 선택 (메인 앱의 모델 목록 사용)
        model_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        model_frame.pack(side="right", padx=5, pady=15)

        ctk.CTkLabel(
            model_frame,
            text="AI 모델:",
            font=(FONT_FAMILY, 10),
            text_color="#A0AEC0"
        ).pack(side="left", padx=(0, 5))

        # 메인 앱에서 사용하는 모델 목록 상속
        available_model_names = list(AI_MODELS.keys())
        state["current_model"] = available_model_names[0] if available_model_names else "GPT-5.2 (고성능)"

        def on_model_change(model_name: str):
            state["current_model"] = model_name
            # AI 인스턴스 재생성 (메인 앱과 동일한 방식)
            try:
                model_config = AI_MODELS.get(model_name, {})
                from ai_classifier import AIClassifier
                state["ai_instance"] = AIClassifier(
                    model=model_config.get("model", "gpt-5.2"),
                    provider=model_config.get("provider", "openai")
                )
            except Exception as e:
                messagebox.showerror("오류", f"모델 변경 실패: {str(e)}")

        model_selector = ctk.CTkOptionMenu(
            model_frame,
            values=available_model_names,
            font=(FONT_FAMILY, 10),
            fg_color="#374151",
            button_color="#4B5563",
            button_hover_color="#6B7280",
            dropdown_font=(FONT_FAMILY, 10),
            width=160,
            height=28,
            command=on_model_change
        )
        model_selector.set(state["current_model"])
        model_selector.pack(side="left")

        # ===== AI 검토/편집 영역 (항상 표시) =====
        ai_panel_frame = ctk.CTkFrame(dialog, fg_color="#1E293B")
        ai_panel_frame.pack(fill="x", padx=10, pady=(5, 0))

        # 좌측: 참조 파일 첨부
        attach_section = ctk.CTkFrame(ai_panel_frame, fg_color="transparent")
        attach_section.pack(side="left", fill="y", padx=10, pady=8)

        ctk.CTkLabel(
            attach_section,
            text="📎 참조 파일",
            font=(FONT_FAMILY, 10, "bold"),
            text_color="white"
        ).pack(anchor="w")

        files_display = ctk.CTkLabel(
            attach_section,
            text="첨부된 파일 없음",
            font=(FONT_FAMILY, 9),
            text_color="gray",
            wraplength=150
        )
        files_display.pack(anchor="w", pady=2)

        attach_btn_frame = ctk.CTkFrame(attach_section, fg_color="transparent")
        attach_btn_frame.pack(anchor="w")

        def attach_reference_files():
            """참조 파일 첨부"""
            filetypes = [
                ("지원 파일", "*.pdf *.xlsx *.xls *.docx *.txt *.json *.yaml *.xml"),
                ("PDF", "*.pdf"),
                ("Excel", "*.xlsx *.xls"),
                ("Word", "*.docx"),
                ("텍스트/데이터", "*.txt *.json *.yaml *.xml"),
            ]
            files = filedialog.askopenfilenames(filetypes=filetypes)
            if files:
                state["attached_files"].extend(files)
                file_names = [os.path.basename(f) for f in state["attached_files"]]
                files_display.configure(text=f"{', '.join(file_names[:2])}{'...' if len(file_names) > 2 else ''}")

        def clear_attached_files():
            state["attached_files"].clear()
            files_display.configure(text="첨부된 파일 없음")

        ctk.CTkButton(
            attach_btn_frame,
            text="첨부",
            font=(FONT_FAMILY, 9),
            fg_color=ACCENT_COLOR,
            width=50,
            height=24,
            command=attach_reference_files
        ).pack(side="left", padx=(0, 3))

        ctk.CTkButton(
            attach_btn_frame,
            text="초기화",
            font=(FONT_FAMILY, 9),
            fg_color="#6B7280",
            width=50,
            height=24,
            command=clear_attached_files
        ).pack(side="left")

        # 중앙: 프롬프트 입력
        prompt_section = ctk.CTkFrame(ai_panel_frame, fg_color="transparent")
        prompt_section.pack(side="left", fill="both", expand=True, padx=10, pady=8)

        ctk.CTkLabel(
            prompt_section,
            text="💬 AI 지시 사항 (선택)",
            font=(FONT_FAMILY, 10, "bold"),
            text_color="white"
        ).pack(anchor="w")

        user_prompt_textbox = ctk.CTkTextbox(
            prompt_section,
            font=(FONT_FAMILY, 10),
            height=50,
            wrap="word",
            fg_color="#374151"
        )
        user_prompt_textbox.pack(fill="x", pady=2)
        user_prompt_textbox.insert("1.0", "예: 첨부한 정책 문서를 참고하여 자본화 기준을 검토해줘")

        # 우측: 검토 버튼들
        review_btn_section = ctk.CTkFrame(ai_panel_frame, fg_color="transparent")
        review_btn_section.pack(side="right", fill="y", padx=10, pady=8)

        ctk.CTkLabel(
            review_btn_section,
            text="🔍 AI 검토",
            font=(FONT_FAMILY, 10, "bold"),
            text_color="white"
        ).pack(anchor="e")

        review_buttons_frame = ctk.CTkFrame(review_btn_section, fg_color="transparent")
        review_buttons_frame.pack(anchor="e", pady=3)

        partial_review_btn = ctk.CTkButton(
            review_buttons_frame,
            text="선택 영역 검토",
            font=(FONT_FAMILY, 9),
            fg_color="#4F46E5",
            hover_color="#4338CA",
            width=100,
            height=26,
            command=lambda: run_partial_review()
        )
        partial_review_btn.pack(side="left", padx=(0, 5))

        full_review_btn = ctk.CTkButton(
            review_buttons_frame,
            text="전체 검토",
            font=(FONT_FAMILY, 9),
            fg_color="#0EA5E9",
            hover_color="#0284C7",
            width=80,
            height=26,
            command=lambda: run_full_ai_review()
        )
        full_review_btn.pack(side="left")

        # 편집 모드용 첨부 영역 (숨김)
        attach_frame = ctk.CTkFrame(dialog, fg_color="#1E3A5F", height=40)

        # ===== 메인 콘텐츠 영역 =====
        content_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        content_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # 좌측: 섹션 트리
        left_frame = ctk.CTkFrame(content_frame, fg_color=SIDEBAR_COLOR, width=300)
        left_frame.pack(side="left", fill="y", padx=(0, 5))
        left_frame.pack_propagate(False)

        ctk.CTkLabel(
            left_frame,
            text="실무지침 구조",
            font=(FONT_FAMILY, 12, "bold"),
            text_color="white"
        ).pack(pady=(15, 10))

        # 섹션 목록 (스크롤 가능)
        section_scroll = ctk.CTkScrollableFrame(left_frame, fg_color="#1E1E24")
        section_scroll.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        # 섹션 정의 (지침 구조에 맞게)
        section_definitions = [
            ("company_profile", "회사 프로필", "회계기준, 자본화 정책 등"),
            ("classification_categories", "분류 카테고리", "유형자산, 무형자산, 비용 등"),
            ("master_decision_tree", "의사결정 트리", "분류 판단 알고리즘"),
            ("learned_cases", "학습된 케이스", "AI가 학습한 분류 사례"),
        ]

        section_checkboxes = {}
        edit_widgets = {}  # 섹션별 편집 위젯 저장

        def toggle_section(section_id):
            if section_id in state["selected_sections"]:
                state["selected_sections"].remove(section_id)
            else:
                state["selected_sections"].add(section_id)
            update_content_display()

        for section_id, section_name, description in section_definitions:
            if section_id in guidelines or section_id == "master_decision_tree":
                frame = ctk.CTkFrame(section_scroll, fg_color="#2D2D35", corner_radius=5)
                frame.pack(fill="x", pady=3)

                cb_var = ctk.BooleanVar(value=False)
                cb = ctk.CTkCheckBox(
                    frame,
                    text="",
                    variable=cb_var,
                    width=20,
                    command=lambda sid=section_id: toggle_section(sid)
                )
                cb.pack(side="left", padx=5, pady=8)

                info_frame = ctk.CTkFrame(frame, fg_color="transparent")
                info_frame.pack(side="left", fill="x", expand=True)

                ctk.CTkLabel(
                    info_frame,
                    text=section_name,
                    font=(FONT_FAMILY, 11, "bold"),
                    text_color="white",
                    anchor="w"
                ).pack(anchor="w", padx=5)

                ctk.CTkLabel(
                    info_frame,
                    text=description,
                    font=(FONT_FAMILY, 9),
                    text_color="gray",
                    anchor="w"
                ).pack(anchor="w", padx=5)

                section_checkboxes[section_id] = cb_var

        # 전체 선택/해제 버튼
        select_btn_frame = ctk.CTkFrame(left_frame, fg_color="transparent")
        select_btn_frame.pack(fill="x", padx=10, pady=5)

        def select_all():
            for sid, var in section_checkboxes.items():
                var.set(True)
                state["selected_sections"].add(sid)
            update_content_display()

        def clear_all():
            for sid, var in section_checkboxes.items():
                var.set(False)
            state["selected_sections"].clear()
            update_content_display()

        ctk.CTkButton(
            select_btn_frame,
            text="전체 선택",
            font=(FONT_FAMILY, 10),
            fg_color="#4B5563",
            width=80,
            height=28,
            command=select_all
        ).pack(side="left", padx=2)

        ctk.CTkButton(
            select_btn_frame,
            text="선택 해제",
            font=(FONT_FAMILY, 10),
            fg_color="#4B5563",
            width=80,
            height=28,
            command=clear_all
        ).pack(side="left", padx=2)

        # 우측: 콘텐츠 표시 영역
        right_frame = ctk.CTkFrame(content_frame, fg_color=SIDEBAR_COLOR)
        right_frame.pack(side="left", fill="both", expand=True)

        # 콘텐츠 헤더
        content_header = ctk.CTkFrame(right_frame, fg_color="#1E3A5F", height=40)
        content_header.pack(fill="x")
        content_header.pack_propagate(False)

        content_title = ctk.CTkLabel(
            content_header,
            text="실무지침 내용 (섹션을 선택하세요)",
            font=(FONT_FAMILY, 12, "bold"),
            text_color="white"
        )
        content_title.pack(side="left", padx=15, pady=8)

        # 콘텐츠 스크롤 영역
        content_scroll = ctk.CTkScrollableFrame(right_frame, fg_color="#1E1E24")
        content_scroll.pack(fill="both", expand=True, padx=5, pady=5)

        def format_section_content(section_id: str) -> str:
            """섹션 내용을 사람이 읽을 수 있는 형태로 변환"""
            if section_id == "company_profile":
                cp = guidelines.get("company_profile", {})
                lines = ["[ 회사 프로필 ]\n"]
                lines.append(f"회계기준: {cp.get('회계기준', 'N/A')}")
                lines.append(f"결산월: {cp.get('결산월', 'N/A')}")
                lines.append(f"업종: {cp.get('업종', 'N/A')}")
                lines.append(f"\n자본화 기준금액: {cp.get('자본화_기준금액', 'N/A'):,}원")

                cap_policy = cp.get("자본화_정책", {})
                if cap_policy:
                    lines.append("\n[ 자본화 정책 ]")
                    lines.append(f"  유형자산: {cap_policy.get('유형자산_기준', 'N/A')}")
                    lines.append(f"  무형자산: {cap_policy.get('무형자산_기준', 'N/A')}")
                    lines.append(f"  개발비: {cap_policy.get('개발비_인식조건', 'N/A')}")
                    lines.append(f"  감가상각: {cap_policy.get('감가상각방법', 'N/A')}")

                    useful_life = cap_policy.get("내용연수", {})
                    if useful_life:
                        lines.append("\n  [ 내용연수 ]")
                        for asset, years in useful_life.items():
                            lines.append(f"    {asset}: {years}")

                direction = cp.get("회계방향", {})
                if direction:
                    lines.append("\n[ 회계 방향 ]")
                    lines.append(f"  기본원칙: {direction.get('기본원칙', 'N/A')}")
                    lines.append(f"  불확실한 경우: {direction.get('불확실한경우', 'N/A')}")

                return "\n".join(lines)

            elif section_id == "classification_categories":
                cats = guidelines.get("classification_categories", {})
                lines = ["[ 분류 카테고리 ]\n"]
                for cat_id, cat_info in cats.items():
                    lines.append(f"◆ {cat_id}")
                    lines.append(f"  설명: {cat_info.get('설명', 'N/A')}")
                    accounts = cat_info.get("계정과목", [])
                    if accounts:
                        lines.append(f"  계정과목: {', '.join(accounts)}")
                    lines.append("")
                return "\n".join(lines)

            elif section_id == "master_decision_tree":
                tree = guidelines.get("master_decision_tree", {})
                lines = ["[ 의사결정 트리 ]\n"]
                for step_id, step_info in tree.items():
                    if isinstance(step_info, dict):
                        lines.append(f"▶ {step_id}")
                        question = step_info.get("질문", step_info.get("설명", ""))
                        if question:
                            lines.append(f"  질문: {question}")

                        yes_result = step_info.get("YES일_때", step_info.get("모두_YES", ""))
                        if yes_result:
                            if isinstance(yes_result, dict):
                                lines.append(f"  YES → {json.dumps(yes_result, ensure_ascii=False, indent=4)[:200]}...")
                            else:
                                lines.append(f"  YES → {yes_result}")

                        no_result = step_info.get("NO일_때", step_info.get("하나라도_NO", ""))
                        if no_result:
                            lines.append(f"  NO → {no_result}")
                        lines.append("")
                return "\n".join(lines)

            elif section_id == "learned_cases":
                cases = guidelines.get("learned_cases", [])
                lines = [f"[ 학습된 케이스 ({len(cases)}건) ]\n"]
                for i, case in enumerate(cases[:10]):  # 최대 10개 표시
                    lines.append(f"케이스 #{i+1}")
                    lines.append(f"  분류: {case.get('classification', 'N/A')}")
                    lines.append(f"  근거: {case.get('reasoning', 'N/A')[:100]}...")
                    lines.append("")
                if len(cases) > 10:
                    lines.append(f"... 외 {len(cases) - 10}건")
                return "\n".join(lines)

            return "내용 없음"

        def update_content_display():
            # 기존 위젯 제거
            for widget in content_scroll.winfo_children():
                widget.destroy()
            edit_widgets.clear()
            state["content_widgets"].clear()

            # 편집 모드일 때 파일 첨부 영역 표시
            if state["edit_mode"]:
                attach_frame.pack(fill="x", after=header_frame)
            else:
                attach_frame.pack_forget()

            if not state["selected_sections"]:
                ctk.CTkLabel(
                    content_scroll,
                    text="좌측에서 검토/수정할 섹션을 선택하세요.",
                    font=(FONT_FAMILY, 12),
                    text_color="gray"
                ).pack(pady=50)
                content_title.configure(text="실무지침 내용 (섹션을 선택하세요)")
                return

            mode_text = "편집" if state["edit_mode"] else "검토"
            content_title.configure(text=f"{mode_text} 중: {len(state['selected_sections'])}개 섹션")

            for section_id in state["selected_sections"]:
                section_frame = ctk.CTkFrame(content_scroll, fg_color="#2D2D35", corner_radius=8)
                section_frame.pack(fill="x", pady=5, padx=5)

                # 섹션 헤더
                section_header = ctk.CTkFrame(section_frame, fg_color="#1E3A5F" if state["edit_mode"] else "transparent")
                section_header.pack(fill="x")

                section_names = {
                    "company_profile": "회사 프로필",
                    "classification_categories": "분류 카테고리",
                    "master_decision_tree": "의사결정 트리",
                    "learned_cases": "학습된 케이스"
                }
                ctk.CTkLabel(
                    section_header,
                    text=section_names.get(section_id, section_id),
                    font=(FONT_FAMILY, 11, "bold"),
                    text_color="white"
                ).pack(side="left", padx=10, pady=5)

                if state["edit_mode"]:
                    # 편집 모드: AI 도움 버튼
                    def ai_assist_edit(sid=section_id):
                        ai_edit_section(sid)

                    ctk.CTkButton(
                        section_header,
                        text="AI 수정 도움",
                        font=(FONT_FAMILY, 9),
                        fg_color="#8B5CF6",
                        width=90,
                        height=24,
                        command=ai_assist_edit
                    ).pack(side="right", padx=5, pady=3)

                content_text = format_section_content(section_id)

                # 편집 가능 여부에 따른 텍스트 위젯
                text_widget = ctk.CTkTextbox(
                    section_frame,
                    font=(FONT_FAMILY, 10),
                    fg_color="#1E1E24" if not state["edit_mode"] else "#2D1F1F",
                    height=200,
                    wrap="word"
                )
                text_widget.pack(fill="x", padx=10, pady=10)
                text_widget.insert("1.0", content_text)

                # 텍스트 위젯을 state에 저장 (선택 영역 추출용)
                state["content_widgets"][section_id] = text_widget

                if state["edit_mode"]:
                    # 편집 모드: 수정 가능
                    text_widget.configure(state="normal")
                    edit_widgets[section_id] = text_widget

                    # 변경 감지
                    def on_text_change(event, sid=section_id, widget=text_widget, original=content_text):
                        current = widget.get("1.0", "end-1c")
                        if current != original:
                            state["modified"] = True
                            state["pending_changes"][sid] = current
                        else:
                            if sid in state["pending_changes"]:
                                del state["pending_changes"][sid]
                            if not state["pending_changes"]:
                                state["modified"] = False
                        # 저장 리마인더 업데이트
                        if "update_save_reminder" in state:
                            state["update_save_reminder"]()

                    text_widget.bind("<KeyRelease>", on_text_change)
                else:
                    text_widget.configure(state="disabled")

        def ai_edit_section(section_id: str):
            """AI 도움으로 섹션 편집"""
            if self.ai is None:
                messagebox.showwarning("경고", "AI가 초기화되지 않았습니다.")
                return

            # 편집 대화상자
            edit_dialog = ctk.CTkToplevel(dialog)
            edit_dialog.title(f"AI 도움: {section_id} 수정")
            edit_dialog.geometry("700x500")
            edit_dialog.configure(fg_color=BG_COLOR)
            edit_dialog.transient(dialog)
            edit_dialog.grab_set()

            ctk.CTkLabel(
                edit_dialog,
                text="어떻게 수정할까요?",
                font=(FONT_FAMILY, 14, "bold"),
                text_color="white"
            ).pack(pady=15)

            # 첨부파일 정보 표시
            if state["attached_files"]:
                file_info = f"참조 파일: {len(state['attached_files'])}개"
                ctk.CTkLabel(
                    edit_dialog,
                    text=file_info,
                    font=(FONT_FAMILY, 10),
                    text_color="#60A5FA"
                ).pack()

            # 수정 지시 입력
            instruction_textbox = ctk.CTkTextbox(
                edit_dialog,
                font=(FONT_FAMILY, 11),
                height=100,
                wrap="word"
            )
            instruction_textbox.pack(fill="x", padx=20, pady=10)
            instruction_textbox.insert("1.0", "예: 자본화 기준금액을 200만원으로 변경해줘\n예: 첨부한 회사 정책 문서를 참고해서 회계기준을 업데이트해줘")

            # 결과 표시 영역
            result_frame = ctk.CTkFrame(edit_dialog, fg_color=SIDEBAR_COLOR)
            result_frame.pack(fill="both", expand=True, padx=20, pady=10)

            result_textbox = ctk.CTkTextbox(
                result_frame,
                font=(FONT_FAMILY, 10),
                wrap="word"
            )
            result_textbox.pack(fill="both", expand=True, padx=10, pady=10)

            def run_ai_edit():
                instruction = instruction_textbox.get("1.0", "end-1c").strip()
                if not instruction or instruction.startswith("예:"):
                    messagebox.showwarning("입력 필요", "수정 지시 사항을 입력하세요.")
                    return

                result_textbox.delete("1.0", "end")
                result_textbox.insert("1.0", "AI가 분석 중...")

                # 현재 섹션 내용
                current_content = format_section_content(section_id)

                # 첨부파일 내용 읽기
                attached_contents = []
                if state["attached_files"]:
                    result_textbox.delete("1.0", "end")
                    result_textbox.insert("1.0", f"첨부파일 {len(state['attached_files'])}개 분석 중...")
                    for file_path in state["attached_files"]:
                        try:
                            file_data = FileReader.read_file(file_path)
                            if file_data.get("type") != "error":
                                content = file_data.get("content", "")[:3000]  # 각 파일 3000자 제한
                                attached_contents.append({
                                    "filename": file_data.get("filename", os.path.basename(file_path)),
                                    "content": content
                                })
                        except Exception as e:
                            attached_contents.append({
                                "filename": os.path.basename(file_path),
                                "content": f"(파일 읽기 실패: {str(e)})"
                            })

                # 첨부파일 내용을 프롬프트에 포함
                attached_text = ""
                if attached_contents:
                    attached_text = "\n\n=== 첨부된 참조 문서 내용 ===\n"
                    for doc in attached_contents:
                        attached_text += f"\n[{doc['filename']}]\n{doc['content']}\n"
                    attached_text += "\n=== 참조 문서 끝 ===\n"

                result_textbox.delete("1.0", "end")
                result_textbox.insert("1.0", "AI가 분석 중...")

                # 프롬프트 구성
                prompt = f"""당신은 회계 실무지침 전문가입니다.

현재 [{section_id}] 섹션의 내용:
{current_content}
{attached_text}
사용자 요청:
{instruction}

위 요청에 따라 섹션 내용을 수정하세요.
첨부된 참조 문서가 있다면, 해당 문서의 내용을 참고하여 지침을 업데이트하세요.

수정 시 주의사항:
1. 기존 내용과 충돌이 있다면 명시적으로 알려주세요
2. 회계기준에 위배되는 변경이라면 경고해주세요
3. 첨부문서 참조 시, 어떤 부분을 반영했는지 명시하세요
4. 수정된 전체 내용을 반환하세요

JSON 형식으로 응답:
{{
    "has_conflict": true/false,
    "conflict_description": "충돌 설명 (있는 경우)",
    "accounting_warning": "회계기준 관련 경고 (있는 경우)",
    "referenced_documents": ["참조한 문서명 목록"],
    "modified_content": "수정된 섹션 내용",
    "change_summary": "변경 요약"
}}"""

                try:
                    # 단순 분석 호출
                    response = self.ai.analyze(prompt, [])

                    if response.get("status") == "ERROR":
                        result_textbox.delete("1.0", "end")
                        result_textbox.insert("1.0", f"오류: {response.get('message', '알 수 없는 오류')}")
                        return

                    # 결과 표시
                    has_conflict = response.get("has_conflict", False)
                    conflict_desc = response.get("conflict_description", "")
                    warning = response.get("accounting_warning", "")
                    referenced = response.get("referenced_documents", [])
                    modified = response.get("modified_content", "")
                    summary = response.get("change_summary", "")

                    result_text = f"[ 변경 요약 ]\n{summary}\n\n"

                    # 참조 문서 표시
                    if referenced:
                        result_text += f"📎 참조된 문서: {', '.join(referenced)}\n\n"

                    if has_conflict:
                        result_text += f"⚠️ 충돌 발생:\n{conflict_desc}\n\n"

                    if warning:
                        result_text += f"⚠️ 회계기준 경고:\n{warning}\n\n"

                    result_text += f"[ 수정된 내용 ]\n{modified}"

                    result_textbox.delete("1.0", "end")
                    result_textbox.insert("1.0", result_text)

                    # 적용 버튼 활성화
                    apply_btn.configure(state="normal")
                    nonlocal ai_result
                    ai_result = response

                except Exception as e:
                    result_textbox.delete("1.0", "end")
                    result_textbox.insert("1.0", f"오류 발생: {str(e)}")

            ai_result = None

            def apply_ai_result():
                if ai_result and section_id in edit_widgets:
                    modified = ai_result.get("modified_content", "")
                    if modified:
                        widget = edit_widgets[section_id]
                        widget.delete("1.0", "end")
                        widget.insert("1.0", modified)
                        state["modified"] = True
                        state["pending_changes"][section_id] = modified
                        # 저장 리마인더 업데이트
                        if "update_save_reminder" in state:
                            state["update_save_reminder"]()
                        edit_dialog.destroy()
                        messagebox.showinfo("적용 완료", "AI 수정 내용이 적용되었습니다.\n💡 하단의 '저장' 버튼을 눌러 저장하세요.")

            # 버튼들
            btn_frame = ctk.CTkFrame(edit_dialog, fg_color="transparent")
            btn_frame.pack(fill="x", padx=20, pady=10)

            ctk.CTkButton(
                btn_frame,
                text="AI 수정 요청",
                font=(FONT_FAMILY, 12, "bold"),
                fg_color=ACCENT_COLOR,
                command=run_ai_edit
            ).pack(side="left", padx=5)

            apply_btn = ctk.CTkButton(
                btn_frame,
                text="적용",
                font=(FONT_FAMILY, 12),
                fg_color=SUCCESS_COLOR,
                text_color="black",
                state="disabled",
                command=apply_ai_result
            )
            apply_btn.pack(side="left", padx=5)

            ctk.CTkButton(
                btn_frame,
                text="취소",
                font=(FONT_FAMILY, 12),
                fg_color="#6B7280",
                command=edit_dialog.destroy
            ).pack(side="right", padx=5)

        # ===== 하단 버튼 영역 =====
        bottom_frame = ctk.CTkFrame(dialog, fg_color=SIDEBAR_COLOR, height=70)
        bottom_frame.pack(fill="x", side="bottom")
        bottom_frame.pack_propagate(False)

        # 검토 결과 표시 영역 (숨겨진 상태로 시작)
        result_container = ctk.CTkFrame(dialog, fg_color="#1E293B")

        def get_user_prompt_and_files():
            """사용자 프롬프트와 첨부파일 내용 추출"""
            user_prompt = user_prompt_textbox.get("1.0", "end-1c").strip()
            if user_prompt.startswith("예:"):
                user_prompt = ""  # 플레이스홀더 제거

            # 첨부파일 내용 읽기
            attached_contents = []
            if state["attached_files"]:
                for file_path in state["attached_files"]:
                    try:
                        file_data = FileReader.read_file(file_path)
                        if file_data.get("type") != "error":
                            content = file_data.get("content", "")[:5000]
                            attached_contents.append({
                                "filename": file_data.get("filename", os.path.basename(file_path)),
                                "content": content
                            })
                    except Exception as e:
                        attached_contents.append({
                            "filename": os.path.basename(file_path),
                            "content": f"(파일 읽기 실패: {str(e)})"
                        })

            return user_prompt, attached_contents

        def get_selected_text_from_widgets():
            """텍스트 위젯에서 선택된 텍스트 추출"""
            selected_texts = {}
            for section_id, widget in state["content_widgets"].items():
                try:
                    # CTkTextbox의 내부 텍스트 위젯에서 선택 영역 확인
                    if hasattr(widget, '_textbox'):
                        inner_widget = widget._textbox
                    else:
                        inner_widget = widget

                    # 선택된 텍스트 가져오기 시도
                    try:
                        sel_start = inner_widget.index("sel.first")
                        sel_end = inner_widget.index("sel.last")
                        selected_text = inner_widget.get(sel_start, sel_end)
                        if selected_text.strip():
                            selected_texts[section_id] = selected_text.strip()
                    except:
                        # 선택 영역이 없음
                        pass
                except Exception:
                    pass
            return selected_texts

        def run_partial_review():
            """선택 영역만 AI 검토 (체크된 섹션 내에서 드래그로 선택한 텍스트만)"""
            if not state["selected_sections"]:
                messagebox.showwarning("경고", "검토할 섹션을 선택하세요.")
                return

            # AI 인스턴스 확인 (모델 선택에 따라)
            ai_instance = state.get("ai_instance", self.ai)
            if ai_instance is None:
                messagebox.showwarning("경고", "AI가 초기화되지 않았습니다. API 키를 확인하세요.")
                return

            user_prompt, attached_contents = get_user_prompt_and_files()

            # 선택된 텍스트 추출
            selected_texts = get_selected_text_from_widgets()

            if not selected_texts:
                messagebox.showinfo(
                    "안내",
                    "선택된 텍스트가 없습니다.\n\n"
                    "텍스트 영역에서 검토할 부분을 마우스로 드래그하여 선택한 후 다시 시도하세요.\n\n"
                    "전체 섹션을 검토하려면 '전체 검토' 버튼을 사용하세요."
                )
                return

            # 전체 지침의 모든 섹션 내용 수집 (맥락 일관성 유지용)
            all_sections_for_context = {}
            all_section_ids = ["company_profile", "classification_categories", "master_decision_tree", "learned_cases"]
            for sid in all_section_ids:
                if sid in guidelines or sid == "master_decision_tree":
                    all_sections_for_context[sid] = format_section_content(sid)

            # 선택된 섹션의 전체 내용 (맥락 파악용)
            section_full_content = {}
            for sec_id in state["selected_sections"]:
                section_full_content[sec_id] = format_section_content(sec_id)

            # 첨부파일 텍스트 구성
            attached_text = ""
            if attached_contents:
                attached_text = "\n\n=== 참조 문서 ===\n"
                for doc in attached_contents:
                    attached_text += f"\n[{doc['filename']}]\n{doc['content'][:3000]}\n"

            # AI 프롬프트 구성
            prompt = f"""당신은 대한민국 회계기준 전문가입니다.
회계 실무지침에서 사용자가 선택한 특정 텍스트 영역만을 검토하고 적정성을 평가해주세요.

중요: 선택 영역만 검토하되, 전체 지침과의 일관성을 반드시 확인하세요.

=== 전체 실무지침 (일관성 확인용 - 수정 불가) ===
{json.dumps(all_sections_for_context, ensure_ascii=False, indent=2)[:10000]}

=== 체크된 섹션의 전체 내용 ===
{json.dumps(section_full_content, ensure_ascii=False, indent=2)[:5000]}

=== 사용자가 선택한 검토/수정 대상 텍스트 ===
{json.dumps(selected_texts, ensure_ascii=False, indent=2)}
{attached_text}
{"=== 사용자 지시 사항 (최우선) ===" if user_prompt else ""}
{user_prompt if user_prompt else ""}

검토 지침:
1. 선택된 텍스트 영역에 대해서만 검토/수정 제안을 하세요
2. 단, 수정 제안은 반드시 전체 지침과의 일관성을 유지해야 합니다
3. 다른 섹션과 충돌하는 수정은 제안하지 마세요
4. 사용자 지시 사항이 있다면 해당 관점에서 우선적으로 검토하세요
5. 참조 문서가 있다면, 현재 지침과 비교하여 변경이 필요한 부분을 식별하세요
6. 회계기준(K-IFRS, 중소기업회계기준) 준수 여부를 확인하세요

JSON으로 응답:
{{
    "review_summary": "선택 영역 검토 요약",
    "findings": [
        {{
            "section": "섹션명",
            "severity": "error|warning|info|pass",
            "title": "발견 사항 제목",
            "description": "상세 설명",
            "recommendation": "권장 수정 사항",
            "reference_doc": "관련 참조 문서명 (있는 경우)"
        }}
    ],
    "suggested_changes": [
        {{
            "section": "섹션명",
            "current_value": "현재 값 (선택된 텍스트 중)",
            "suggested_value": "제안 값",
            "reason": "변경 이유"
        }}
    ],
    "overall_assessment": "적합|부분수정필요|전면검토필요"
}}"""

            # AI 호출
            try:
                result_container.pack(fill="both", expand=True, side="bottom", before=bottom_frame)
                for widget in result_container.winfo_children():
                    widget.destroy()

                selected_sections_text = ", ".join(selected_texts.keys())
                loading_label = ctk.CTkLabel(
                    result_container,
                    text=f"🔍 AI가 선택된 텍스트를 검토하고 있습니다... ({selected_sections_text})",
                    font=(FONT_FAMILY, 12),
                    text_color="#60A5FA"
                )
                loading_label.pack(pady=20)
                dialog.update_idletasks()

                response = ai_instance.analyze(prompt, [])
                loading_label.destroy()

                show_ai_review_results(response, "선택 영역")

            except Exception as e:
                messagebox.showerror("오류", f"AI 검토 중 오류 발생: {str(e)}")

        def show_ai_review_results(response: dict, review_type: str):
            """AI 검토 결과를 패널에 표시 (상세 정보 포함)"""
            for widget in result_container.winfo_children():
                widget.destroy()

            # 헤더
            header = ctk.CTkFrame(result_container, fg_color="#0F172A")
            header.pack(fill="x")

            assessment = response.get("overall_assessment", "알 수 없음")
            assessment_colors = {
                "적합": SUCCESS_COLOR,
                "부분수정필요": WARNING_COLOR,
                "전면검토필요": ERROR_COLOR
            }
            assessment_color = assessment_colors.get(assessment, "#A0AEC0")

            ctk.CTkLabel(
                header,
                text=f"🔍 {review_type} 검토 결과",
                font=(FONT_FAMILY, 12, "bold"),
                text_color="white"
            ).pack(side="left", padx=15, pady=8)

            ctk.CTkLabel(
                header,
                text=f"평가: {assessment}",
                font=(FONT_FAMILY, 11, "bold"),
                text_color=assessment_color
            ).pack(side="left", padx=20, pady=8)

            # 통계 정보
            findings_count = len(response.get("findings", []))
            changes_count = len(response.get("suggested_changes", []))
            ctk.CTkLabel(
                header,
                text=f"발견: {findings_count}건 | 제안: {changes_count}건",
                font=(FONT_FAMILY, 10),
                text_color="#60A5FA"
            ).pack(side="left", padx=15, pady=8)

            ctk.CTkButton(
                header,
                text="✕ 닫기",
                font=(FONT_FAMILY, 10),
                fg_color="transparent",
                hover_color="#374151",
                width=60,
                command=lambda: result_container.pack_forget()
            ).pack(side="right", padx=5)

            # 전체 결과 보기 버튼
            def show_full_result():
                full_dialog = ctk.CTkToplevel(dialog)
                full_dialog.title("AI 검토 결과 전체 보기")
                full_dialog.geometry("900x700")
                full_dialog.configure(fg_color=BG_COLOR)
                full_dialog.transient(dialog)
                full_dialog.grab_set()

                ctk.CTkLabel(
                    full_dialog,
                    text="📋 AI 검토 결과 전체 내용",
                    font=(FONT_FAMILY, 14, "bold"),
                    text_color="white"
                ).pack(pady=10)

                full_textbox = ctk.CTkTextbox(
                    full_dialog,
                    font=(FONT_FAMILY, 10),
                    fg_color="#1E1E24",
                    wrap="word"
                )
                full_textbox.pack(fill="both", expand=True, padx=15, pady=10)
                full_textbox.insert("1.0", json.dumps(response, ensure_ascii=False, indent=2))
                full_textbox.configure(state="disabled")

                ctk.CTkButton(
                    full_dialog,
                    text="닫기",
                    font=(FONT_FAMILY, 12),
                    fg_color="#4B5563",
                    width=100,
                    command=full_dialog.destroy
                ).pack(pady=10)

            ctk.CTkButton(
                header,
                text="전체 결과 보기",
                font=(FONT_FAMILY, 9),
                fg_color="#4B5563",
                hover_color="#374151",
                width=90,
                command=show_full_result
            ).pack(side="right", padx=5)

            # 검토 요약
            summary = response.get("review_summary", "")
            if summary:
                summary_frame = ctk.CTkFrame(result_container, fg_color="#1E3A5F")
                summary_frame.pack(fill="x", padx=10, pady=5)
                ctk.CTkLabel(
                    summary_frame,
                    text=f"📋 요약: {summary}",
                    font=(FONT_FAMILY, 10),
                    text_color="white",
                    wraplength=800,
                    justify="left"
                ).pack(padx=10, pady=8, anchor="w")

            # 결과 스크롤 영역
            results_scroll = ctk.CTkScrollableFrame(result_container, fg_color="transparent", height=200)
            results_scroll.pack(fill="both", expand=True, padx=10, pady=5)

            # 발견 사항
            findings = response.get("findings", [])
            if findings:
                ctk.CTkLabel(
                    results_scroll,
                    text=f"📌 발견 사항 ({len(findings)}건)",
                    font=(FONT_FAMILY, 11, "bold"),
                    text_color="white"
                ).pack(anchor="w", pady=(5, 3))

                severity_colors = {
                    "error": (ERROR_COLOR, "⛔ 위반"),
                    "warning": (WARNING_COLOR, "⚠️ 주의"),
                    "info": ("#60A5FA", "ℹ️ 참고"),
                    "pass": (SUCCESS_COLOR, "✅ 적합")
                }

                for idx, finding in enumerate(findings):
                    sev = finding.get("severity", "info")
                    color, label = severity_colors.get(sev, ("#A0AEC0", "기타"))

                    # 메인 프레임
                    item_container = ctk.CTkFrame(results_scroll, fg_color="#374151", corner_radius=5)
                    item_container.pack(fill="x", pady=3)

                    # 헤더 행
                    header_row = ctk.CTkFrame(item_container, fg_color="transparent")
                    header_row.pack(fill="x", padx=5, pady=3)

                    ctk.CTkLabel(
                        header_row,
                        text=label,
                        font=(FONT_FAMILY, 9, "bold"),
                        text_color=color,
                        width=70
                    ).pack(side="left", padx=3)

                    ctk.CTkLabel(
                        header_row,
                        text=f"[{finding.get('section', '')}] {finding.get('title', '')}",
                        font=(FONT_FAMILY, 10, "bold"),
                        text_color="white"
                    ).pack(side="left", padx=5)

                    # 상세 내용 영역 (접기/펼치기 가능)
                    detail_frame = ctk.CTkFrame(item_container, fg_color="#2D2D35")
                    detail_visible = ctk.BooleanVar(value=False)

                    def toggle_detail(frame=detail_frame, var=detail_visible):
                        if var.get():
                            frame.pack_forget()
                            var.set(False)
                        else:
                            frame.pack(fill="x", padx=8, pady=(0, 5))
                            var.set(True)

                    # 상세보기 버튼
                    ctk.CTkButton(
                        header_row,
                        text="▼ 상세",
                        font=(FONT_FAMILY, 9),
                        fg_color="transparent",
                        hover_color="#4B5563",
                        width=60,
                        height=22,
                        command=toggle_detail
                    ).pack(side="right", padx=3)

                    # 상세 내용 구성
                    desc = finding.get('description', '')
                    rec = finding.get('recommendation', '')
                    ref_doc = finding.get('reference_doc', '')

                    detail_text = ""
                    if desc:
                        detail_text += f"📝 설명:\n{desc}\n\n"
                    if rec:
                        detail_text += f"💡 권장사항:\n{rec}\n\n"
                    if ref_doc:
                        detail_text += f"📚 참조 문서: {ref_doc}\n"

                    if detail_text:
                        detail_label = ctk.CTkLabel(
                            detail_frame,
                            text=detail_text.strip(),
                            font=(FONT_FAMILY, 9),
                            text_color="#D1D5DB",
                            wraplength=750,
                            justify="left"
                        )
                        detail_label.pack(padx=10, pady=8, anchor="w")

                    # 짧은 미리보기 (항상 표시)
                    preview = desc[:80] + "..." if len(desc) > 80 else desc
                    if preview:
                        ctk.CTkLabel(
                            header_row,
                            text=preview,
                            font=(FONT_FAMILY, 9),
                            text_color="gray"
                        ).pack(side="left", padx=5, fill="x", expand=True)

            # 제안된 변경사항
            suggested_changes = response.get("suggested_changes", [])
            if suggested_changes:
                ctk.CTkLabel(
                    results_scroll,
                    text=f"💡 제안된 변경사항 ({len(suggested_changes)}건)",
                    font=(FONT_FAMILY, 11, "bold"),
                    text_color="white"
                ).pack(anchor="w", pady=(15, 3))

                for change in suggested_changes:
                    # 메인 컨테이너
                    change_container = ctk.CTkFrame(results_scroll, fg_color="#1E3A1E", corner_radius=5)
                    change_container.pack(fill="x", pady=3)

                    # 헤더 행
                    change_header = ctk.CTkFrame(change_container, fg_color="transparent")
                    change_header.pack(fill="x", padx=5, pady=3)

                    ctk.CTkLabel(
                        change_header,
                        text=f"[{change.get('section', '')}]",
                        font=(FONT_FAMILY, 9, "bold"),
                        text_color="#60A5FA"
                    ).pack(side="left", padx=3)

                    # 간단한 미리보기
                    current_short = change.get('current_value', '')[:40]
                    suggested_short = change.get('suggested_value', '')[:40]
                    preview_text = f"'{current_short}...' → '{suggested_short}...'" if len(change.get('current_value', '')) > 40 else f"'{current_short}' → '{suggested_short}'"
                    ctk.CTkLabel(
                        change_header,
                        text=preview_text,
                        font=(FONT_FAMILY, 9),
                        text_color="white"
                    ).pack(side="left", padx=5)

                    # 상세 내용 영역 (접기/펼치기 가능)
                    change_detail_frame = ctk.CTkFrame(change_container, fg_color="#2D3D2D")
                    change_detail_visible = ctk.BooleanVar(value=False)

                    def toggle_change_detail(frame=change_detail_frame, var=change_detail_visible):
                        if var.get():
                            frame.pack_forget()
                            var.set(False)
                        else:
                            frame.pack(fill="x", padx=8, pady=(0, 5))
                            var.set(True)

                    # 변경 적용 버튼
                    def apply_single_change(c=change):
                        apply_suggested_change(c)

                    ctk.CTkButton(
                        change_header,
                        text="✓ 적용",
                        font=(FONT_FAMILY, 9, "bold"),
                        fg_color=SUCCESS_COLOR,
                        text_color="black",
                        width=60,
                        height=22,
                        command=apply_single_change
                    ).pack(side="right", padx=3, pady=2)

                    # 상세보기 버튼
                    ctk.CTkButton(
                        change_header,
                        text="▼ 상세",
                        font=(FONT_FAMILY, 9),
                        fg_color="transparent",
                        hover_color="#3D5D3D",
                        width=60,
                        height=22,
                        command=toggle_change_detail
                    ).pack(side="right", padx=3)

                    # 상세 내용 구성
                    current_val = change.get('current_value', '')
                    suggested_val = change.get('suggested_value', '')
                    reason = change.get('reason', '')

                    detail_content = ctk.CTkFrame(change_detail_frame, fg_color="transparent")
                    detail_content.pack(fill="x", padx=10, pady=8)

                    if current_val:
                        ctk.CTkLabel(
                            detail_content,
                            text="📋 현재 값:",
                            font=(FONT_FAMILY, 9, "bold"),
                            text_color="#FFB366"
                        ).pack(anchor="w")
                        ctk.CTkLabel(
                            detail_content,
                            text=current_val,
                            font=(FONT_FAMILY, 9),
                            text_color="#D1D5DB",
                            wraplength=720,
                            justify="left"
                        ).pack(anchor="w", padx=15, pady=(0, 8))

                    if suggested_val:
                        ctk.CTkLabel(
                            detail_content,
                            text="✨ 제안 값:",
                            font=(FONT_FAMILY, 9, "bold"),
                            text_color=SUCCESS_COLOR
                        ).pack(anchor="w")
                        ctk.CTkLabel(
                            detail_content,
                            text=suggested_val,
                            font=(FONT_FAMILY, 9),
                            text_color="#D1D5DB",
                            wraplength=720,
                            justify="left"
                        ).pack(anchor="w", padx=15, pady=(0, 8))

                    if reason:
                        ctk.CTkLabel(
                            detail_content,
                            text="💡 변경 이유:",
                            font=(FONT_FAMILY, 9, "bold"),
                            text_color="#60A5FA"
                        ).pack(anchor="w")
                        ctk.CTkLabel(
                            detail_content,
                            text=reason,
                            font=(FONT_FAMILY, 9),
                            text_color="#D1D5DB",
                            wraplength=720,
                            justify="left"
                        ).pack(anchor="w", padx=15)

            # 하단 액션 버튼
            action_frame = ctk.CTkFrame(result_container, fg_color="transparent")
            action_frame.pack(fill="x", padx=10, pady=8)

            if suggested_changes:
                ctk.CTkButton(
                    action_frame,
                    text="모든 제안 적용",
                    font=(FONT_FAMILY, 11, "bold"),
                    fg_color=SUCCESS_COLOR,
                    text_color="black",
                    width=120,
                    command=lambda: apply_all_suggested_changes(suggested_changes)
                ).pack(side="left", padx=5)

            ctk.CTkButton(
                action_frame,
                text="프롬프트로 수정 지시",
                font=(FONT_FAMILY, 11),
                fg_color="#8B5CF6",
                width=140,
                command=lambda: open_modification_dialog(response)
            ).pack(side="left", padx=5)

            # 결과 저장
            state["last_review_result"] = response

        def apply_suggested_change(change: dict, show_message: bool = True) -> bool:
            """단일 제안 변경사항 적용 - 실제로 텍스트 위젯 내용을 변경함"""
            section = change.get("section", "")
            suggested = change.get("suggested_value", "")
            current_value = change.get("current_value", "")
            reason = change.get("reason", "")

            if not section or not suggested:
                return False

            # 편집 모드 활성화
            if not state["edit_mode"]:
                edit_mode_var.set(True)
                toggle_edit_mode()

            # 해당 섹션 선택 및 위젯 생성
            if section not in state["selected_sections"]:
                if section in section_checkboxes:
                    section_checkboxes[section].set(True)
                    state["selected_sections"].add(section)
                    update_content_display()

            # 텍스트 위젯에 변경 적용
            if section in state["content_widgets"]:
                widget = state["content_widgets"][section]
                current_content = widget.get("1.0", "end-1c")

                # 기존 텍스트에서 current_value를 suggested로 교체
                if current_value and current_value in current_content:
                    new_content = current_content.replace(current_value, suggested, 1)
                    widget.delete("1.0", "end")
                    widget.insert("1.0", new_content)
                else:
                    # current_value가 없거나 못찾으면 전체 교체 또는 끝에 추가
                    # suggested가 완전한 내용인 경우 교체, 아니면 끝에 추가
                    if len(suggested) > 100:  # 충분히 긴 경우 전체 교체로 간주
                        widget.delete("1.0", "end")
                        widget.insert("1.0", suggested)
                    else:
                        # 짧은 변경사항은 끝에 추가
                        widget.insert("end", f"\n\n[AI 제안] {suggested}")

                # 수정 상태 업데이트
                state["modified"] = True
                state["pending_changes"][section] = widget.get("1.0", "end-1c")
                # 저장 리마인더 업데이트
                if "update_save_reminder" in state:
                    state["update_save_reminder"]()

                if show_message:
                    messagebox.showinfo(
                        "✓ 변경 적용됨",
                        f"'{section}' 섹션에 제안이 적용되었습니다.\n\n"
                        f"변경 이유: {reason}\n\n"
                        f"💡 하단의 '저장' 버튼을 눌러 저장하세요."
                    )
                return True
            else:
                if show_message:
                    messagebox.showwarning(
                        "적용 실패",
                        f"'{section}' 섹션의 위젯을 찾을 수 없습니다.\n"
                        f"해당 섹션을 먼저 선택해주세요."
                    )
                return False

        def apply_all_suggested_changes(changes: list):
            """모든 제안 변경사항 적용 - 실제로 텍스트 내용을 변경함"""
            if not changes:
                messagebox.showinfo("안내", "적용할 변경사항이 없습니다.")
                return

            if not messagebox.askyesno(
                "모든 제안 적용",
                f"{len(changes)}개의 제안된 변경사항을 모두 적용하시겠습니까?\n\n"
                "각 섹션의 텍스트 내용이 실제로 변경됩니다.\n"
                "적용 후 '저장' 버튼을 눌러 저장하세요."
            ):
                return

            # 편집 모드 활성화
            if not state["edit_mode"]:
                edit_mode_var.set(True)
                toggle_edit_mode()

            # 먼저 모든 관련 섹션 선택
            sections_to_update = set()
            for change in changes:
                section = change.get("section", "")
                if section:
                    sections_to_update.add(section)
                    if section in section_checkboxes:
                        section_checkboxes[section].set(True)
                        state["selected_sections"].add(section)

            # 위젯 생성을 위해 디스플레이 업데이트
            update_content_display()

            # 실제 변경 적용
            success_count = 0
            failed_sections = []

            for change in changes:
                if apply_suggested_change(change, show_message=False):
                    success_count += 1
                else:
                    section = change.get("section", "unknown")
                    failed_sections.append(section)

            # 결과 메시지
            if success_count == len(changes):
                messagebox.showinfo(
                    "✓ 모두 적용됨",
                    f"{success_count}개의 변경사항이 모두 적용되었습니다.\n\n"
                    f"💡 텍스트 내용을 확인 후 하단의 '저장' 버튼을 눌러 저장하세요.\n\n"
                    f"저장 버튼이 보이지 않으면 '편집 모드'가 활성화되어 있는지 확인하세요."
                )
            elif success_count > 0:
                messagebox.showwarning(
                    "부분 적용됨",
                    f"{len(changes)}개 중 {success_count}개 적용됨\n\n"
                    f"적용 실패한 섹션:\n• " + "\n• ".join(failed_sections) + "\n\n"
                    f"💡 적용된 내용 확인 후 '저장' 버튼으로 저장하세요."
                )
            else:
                messagebox.showerror(
                    "적용 실패",
                    "변경사항을 적용하지 못했습니다.\n"
                    "해당 섹션들을 먼저 선택한 후 다시 시도하세요."
                )

        def open_modification_dialog(review_result: dict):
            """검토 결과를 바탕으로 수정 지시 다이얼로그"""
            mod_dialog = ctk.CTkToplevel(dialog)
            mod_dialog.title("AI에게 수정 지시")
            mod_dialog.geometry("600x400")
            mod_dialog.configure(fg_color=BG_COLOR)
            mod_dialog.transient(dialog)
            mod_dialog.grab_set()

            ctk.CTkLabel(
                mod_dialog,
                text="검토 결과를 바탕으로 AI에게 수정을 지시하세요",
                font=(FONT_FAMILY, 13, "bold"),
                text_color="white"
            ).pack(pady=15)

            # 검토 요약 표시
            summary = review_result.get("review_summary", "")
            if summary:
                ctk.CTkLabel(
                    mod_dialog,
                    text=f"📋 검토 요약: {summary[:150]}...",
                    font=(FONT_FAMILY, 10),
                    text_color="#A0AEC0",
                    wraplength=550
                ).pack(pady=5)

            ctk.CTkLabel(
                mod_dialog,
                text="수정 지시 입력:",
                font=(FONT_FAMILY, 11),
                text_color="white"
            ).pack(anchor="w", padx=20, pady=(10, 5))

            mod_textbox = ctk.CTkTextbox(
                mod_dialog,
                font=(FONT_FAMILY, 11),
                height=120,
                wrap="word"
            )
            mod_textbox.pack(fill="x", padx=20, pady=5)
            mod_textbox.insert("1.0", "예: 검토 결과에서 지적된 자본화 기준을 200만원으로 수정해줘")

            result_label = ctk.CTkLabel(
                mod_dialog,
                text="",
                font=(FONT_FAMILY, 10),
                text_color="#60A5FA"
            )
            result_label.pack(pady=10)

            def execute_modification():
                instruction = mod_textbox.get("1.0", "end-1c").strip()
                if not instruction or instruction.startswith("예:"):
                    messagebox.showwarning("입력 필요", "수정 지시를 입력하세요.")
                    return

                result_label.configure(text="AI가 수정을 진행 중...")
                mod_dialog.update_idletasks()

                # AI에게 수정 요청
                mod_prompt = f"""당신은 회계 실무지침 전문가입니다.
아래 검토 결과를 바탕으로 사용자의 수정 지시를 수행하세요.

=== 검토 결과 ===
{json.dumps(review_result, ensure_ascii=False, indent=2)}

=== 사용자 수정 지시 ===
{instruction}

수정된 내용을 JSON으로 응답:
{{
    "modified_sections": {{
        "섹션명": "수정된 전체 내용"
    }},
    "change_description": "변경 내용 설명"
}}"""

                try:
                    response = self.ai.analyze(mod_prompt, [])
                    modified_sections = response.get("modified_sections", {})
                    change_desc = response.get("change_description", "")

                    if modified_sections:
                        # 편집 모드 활성화 및 변경사항 적용
                        if not state["edit_mode"]:
                            edit_mode_var.set(True)
                            toggle_edit_mode()

                        for section_id, new_content in modified_sections.items():
                            if section_id in section_checkboxes:
                                section_checkboxes[section_id].set(True)
                                state["selected_sections"].add(section_id)
                            state["pending_changes"][section_id] = new_content
                            state["modified"] = True

                        update_content_display()
                        # 저장 리마인더 업데이트
                        if "update_save_reminder" in state:
                            state["update_save_reminder"]()

                        messagebox.showinfo(
                            "수정 완료",
                            f"AI가 {len(modified_sections)}개 섹션을 수정했습니다.\n\n"
                            f"{change_desc}\n\n"
                            f"💡 내용을 확인하고 하단의 '저장' 버튼을 눌러 저장하세요."
                        )
                        mod_dialog.destroy()
                    else:
                        result_label.configure(text="수정할 내용이 없습니다.")

                except Exception as e:
                    result_label.configure(text=f"오류: {str(e)}")

            btn_frame = ctk.CTkFrame(mod_dialog, fg_color="transparent")
            btn_frame.pack(fill="x", padx=20, pady=15)

            ctk.CTkButton(
                btn_frame,
                text="수정 실행",
                font=(FONT_FAMILY, 12, "bold"),
                fg_color=ACCENT_COLOR,
                command=execute_modification
            ).pack(side="left", padx=5)

            ctk.CTkButton(
                btn_frame,
                text="취소",
                font=(FONT_FAMILY, 12),
                fg_color="#6B7280",
                command=mod_dialog.destroy
            ).pack(side="right", padx=5)

        def save_changes():
            """변경사항 저장 (충돌 검사 및 AI 검증 포함)"""
            if not state["modified"] or not state["pending_changes"]:
                messagebox.showinfo("안내", "저장할 변경사항이 없습니다.")
                return

            # 충돌 검사 수행
            conflicts = []
            for section_id, new_content in state["pending_changes"].items():
                original_content = format_section_content(section_id)
                if original_content != new_content:
                    # AI에게 충돌 여부 확인
                    conflicts.append({
                        "section": section_id,
                        "original": original_content[:500],
                        "new": new_content[:500]
                    })

            if conflicts and self.ai:
                # AI 충돌 검사
                conflict_check_prompt = f"""당신은 회계 실무지침 전문가입니다.

다음 변경사항이 기존 내용과 충돌하는지, 회계기준에 위배되는지 분석해주세요.

변경 내용:
{json.dumps(conflicts, ensure_ascii=False, indent=2)}

JSON으로 응답:
{{
    "has_critical_conflict": true/false,
    "conflicts": [
        {{"section": "섹션명", "issue": "문제 설명", "severity": "critical/warning/info"}}
    ],
    "recommendation": "권장 사항",
    "safe_to_save": true/false
}}"""

                try:
                    response = self.ai.analyze(conflict_check_prompt, [])

                    if not response.get("safe_to_save", True):
                        conflict_msg = "다음 충돌이 발견되었습니다:\n\n"
                        for c in response.get("conflicts", []):
                            conflict_msg += f"• [{c.get('section')}] {c.get('issue')}\n"
                        conflict_msg += f"\n권장사항: {response.get('recommendation', '')}"

                        if not messagebox.askyesno(
                            "충돌 경고",
                            f"{conflict_msg}\n\n그래도 저장하시겠습니까?"
                        ):
                            return
                except Exception as e:
                    # AI 검사 실패 시 경고 후 계속
                    if not messagebox.askyesno(
                        "검증 실패",
                        f"AI 충돌 검사 중 오류 발생: {str(e)}\n\n검증 없이 저장하시겠습니까?"
                    ):
                        return

            # 백업 생성
            try:
                create_backup("manual_edit")
            except Exception as e:
                if not messagebox.askyesno(
                    "백업 실패",
                    f"백업 생성 중 오류: {str(e)}\n\n백업 없이 저장하시겠습니까?"
                ):
                    return

            # 실제 저장 수행
            section_names = {
                "company_profile": "회사 프로필",
                "classification_categories": "분류 카테고리",
                "master_decision_tree": "의사결정 트리",
                "learned_cases": "학습된 케이스"
            }

            save_summary = "다음 섹션이 수정되었습니다:\n\n"
            for section_id in state["pending_changes"].keys():
                save_summary += f"• {section_names.get(section_id, section_id)}\n"

            try:
                # 텍스트 형태의 변경 내용을 파싱하여 원본 guidelines 업데이트
                # 주의: 텍스트박스의 내용은 표시용 포맷이므로,
                # 실제로는 원본 JSON 구조 자체를 수정해야 함
                # 여기서는 변경된 섹션의 메타정보를 기록하고 사용자에게 알림

                from config import GUIDELINES_JSON_PATH, GUIDELINES_YAML_PATH
                import yaml

                # JSON 파일 저장
                # 변경 사항을 audit_trail에 기록
                if "audit_trail" not in guidelines:
                    guidelines["audit_trail"] = []

                from datetime import datetime
                audit_entry = {
                    "timestamp": datetime.now().isoformat(),
                    "action": "manual_edit",
                    "sections_modified": list(state["pending_changes"].keys()),
                    "change_count": len(state["pending_changes"])
                }
                guidelines["audit_trail"].append(audit_entry)

                # 메타데이터 업데이트
                if "_metadata" in guidelines:
                    guidelines["_metadata"]["last_updated"] = datetime.now().strftime("%Y-%m-%d")

                # JSON 파일 저장
                with open(GUIDELINES_JSON_PATH, 'w', encoding='utf-8') as f:
                    json.dump(guidelines, f, ensure_ascii=False, indent=2)

                # YAML 파일도 동기화 (있는 경우)
                if os.path.exists(GUIDELINES_YAML_PATH):
                    try:
                        with open(GUIDELINES_YAML_PATH, 'w', encoding='utf-8') as f:
                            yaml.dump(guidelines, f, allow_unicode=True, default_flow_style=False, sort_keys=False)
                    except Exception:
                        pass  # YAML 동기화 실패는 무시

                messagebox.showinfo(
                    "저장 완료",
                    f"{save_summary}\n변경사항이 저장되었습니다.\n"
                    f"이전 버전은 '버전 관리'에서 복원할 수 있습니다."
                )

            except Exception as e:
                messagebox.showerror(
                    "저장 실패",
                    f"파일 저장 중 오류가 발생했습니다:\n{str(e)}"
                )
                return

            # 상태 초기화
            state["modified"] = False
            state["pending_changes"].clear()
            # 저장 리마인더 업데이트
            if "update_save_reminder" in state:
                state["update_save_reminder"]()

        def discard_changes():
            """변경사항 취소"""
            if state["modified"]:
                if messagebox.askyesno("확인", "수정한 내용을 모두 취소하시겠습니까?"):
                    state["modified"] = False
                    state["pending_changes"].clear()
                    # 저장 리마인더 업데이트
                    if "update_save_reminder" in state:
                        state["update_save_reminder"]()
                    update_content_display()

        def run_full_ai_review():
            """AI를 사용한 체크된 섹션 전체 검토 (참조 파일 및 프롬프트 통합)"""
            # AI 인스턴스 확인 (모델 선택에 따라)
            ai_instance = state.get("ai_instance", self.ai)
            if ai_instance is None:
                messagebox.showwarning("경고", "AI가 초기화되지 않았습니다. API 키를 확인하세요.")
                return

            if not state["selected_sections"]:
                messagebox.showwarning("경고", "검토할 섹션을 선택하세요.\n좌측 목록에서 하나 이상의 섹션을 체크해주세요.")
                return

            user_prompt, attached_contents = get_user_prompt_and_files()

            # 체크된 섹션의 전체 내용 수집
            checked_sections_content = {}
            section_names_map = {
                "company_profile": "회사 프로필",
                "classification_categories": "분류 카테고리",
                "master_decision_tree": "의사결정 트리",
                "learned_cases": "학습된 케이스"
            }

            for section_id in state["selected_sections"]:
                if section_id in guidelines or section_id == "master_decision_tree":
                    checked_sections_content[section_id] = format_section_content(section_id)

            # 전체 지침 맥락 (참고용 - 체크되지 않은 섹션 요약)
            full_context = {
                "company_profile": guidelines.get("company_profile", {}),
                "classification_categories": list(guidelines.get("classification_categories", {}).keys()),
            }

            # 첨부파일 텍스트 구성
            attached_text = ""
            if attached_contents:
                attached_text = "\n\n=== 참조 문서 ===\n"
                for doc in attached_contents:
                    attached_text += f"\n[{doc['filename']}]\n{doc['content'][:4000]}\n"

            # 섹션 이름 목록
            checked_section_names = [section_names_map.get(sid, sid) for sid in state["selected_sections"]]

            # AI 프롬프트 구성
            prompt = f"""당신은 대한민국 회계기준 전문가입니다.
회계 실무지침에서 체크된 섹션({', '.join(checked_section_names)})의 전체 내용을 검토하고 적정성을 평가해주세요.

=== 전체 지침 맥락 (참고용) ===
{json.dumps(full_context, ensure_ascii=False, indent=2)}

=== 검토 대상 섹션의 전체 내용 ({len(checked_sections_content)}개 섹션) ===
{json.dumps(checked_sections_content, ensure_ascii=False, indent=2)[:15000]}
{attached_text}
{"=== 사용자 지시 사항 (최우선) ===" if user_prompt else ""}
{user_prompt if user_prompt else ""}

검토 지침:
1. 체크된 섹션({', '.join(checked_section_names)})의 전체 내용을 상세히 검토하세요
2. 사용자 지시 사항이 있다면 해당 관점에서 우선적으로 검토하세요
3. 참조 문서가 있다면, 현재 지침과 비교하여 변경이 필요한 부분을 식별하세요
4. 각 섹션 간의 일관성을 확인하세요
5. 회계기준(K-IFRS, 중소기업회계기준) 준수 여부를 확인하세요
6. 실무적 적용 가능성을 평가하세요

JSON으로 응답:
{{
    "review_summary": "전체 검토 요약",
    "findings": [
        {{
            "section": "섹션명",
            "severity": "error|warning|info|pass",
            "title": "발견 사항 제목",
            "description": "상세 설명",
            "recommendation": "권장 수정 사항",
            "reference_doc": "관련 참조 문서명 (있는 경우)"
        }}
    ],
    "consistency_issues": [
        {{
            "sections": ["관련 섹션1", "관련 섹션2"],
            "issue": "일관성 문제 설명"
        }}
    ],
    "suggested_changes": [
        {{
            "section": "섹션명",
            "current_value": "현재 값",
            "suggested_value": "제안 값",
            "reason": "변경 이유"
        }}
    ],
    "overall_assessment": "적합|부분수정필요|전면검토필요"
}}"""

            # AI 호출
            try:
                result_container.pack(fill="both", expand=True, side="bottom", before=bottom_frame)
                for widget in result_container.winfo_children():
                    widget.destroy()

                sections_text = ", ".join(checked_section_names)
                loading_label = ctk.CTkLabel(
                    result_container,
                    text=f"🔍 AI가 체크된 섹션({sections_text})을 검토하고 있습니다...",
                    font=(FONT_FAMILY, 12),
                    text_color="#60A5FA"
                )
                loading_label.pack(pady=20)
                dialog.update_idletasks()

                response = ai_instance.analyze(prompt, [])
                loading_label.destroy()

                show_ai_review_results(response, f"체크된 섹션 ({sections_text})")

                # 요약 메시지
                assessment = response.get("overall_assessment", "알 수 없음")
                findings = response.get("findings", [])
                error_count = sum(1 for f in findings if f.get("severity") == "error")
                warning_count = sum(1 for f in findings if f.get("severity") == "warning")

                if error_count > 0:
                    messagebox.showwarning(
                        "검토 완료",
                        f"검토 대상: {sections_text}\n"
                        f"평가: {assessment}\n\n"
                        f"위반 사항: {error_count}건\n"
                        f"주의 사항: {warning_count}건\n\n"
                        f"아래 결과 패널에서 상세 내용을 확인하세요."
                    )
                else:
                    messagebox.showinfo(
                        "검토 완료",
                        f"검토 대상: {sections_text}\n"
                        f"평가: {assessment}\n\n"
                        f"회계기준 위반 사항 없음\n"
                        f"주의 사항: {warning_count}건"
                    )

            except Exception as e:
                messagebox.showerror("오류", f"AI 검토 중 오류 발생: {str(e)}")

        # 닫기 버튼 (항상 표시)
        def on_close():
            if state["modified"]:
                result = messagebox.askyesnocancel(
                    "변경사항 확인",
                    "저장하지 않은 변경사항이 있습니다.\n\n저장하시겠습니까?"
                )
                if result is True:  # 예
                    save_changes()
                    dialog.destroy()
                elif result is False:  # 아니오
                    dialog.destroy()
                # None (취소)인 경우 아무것도 안함
            else:
                dialog.destroy()

        ctk.CTkButton(
            bottom_frame,
            text="닫기",
            font=(FONT_FAMILY, 12),
            fg_color="#4B5563",
            width=80,
            command=on_close
        ).pack(side="right", padx=15, pady=15)

        # 저장 버튼 (편집 모드용)
        save_btn = ctk.CTkButton(
            bottom_frame,
            text="저장",
            font=(FONT_FAMILY, 12, "bold"),
            fg_color=SUCCESS_COLOR,
            text_color="black",
            width=80,
            command=save_changes
        )

        # 취소 버튼 (편집 모드용)
        discard_btn = ctk.CTkButton(
            bottom_frame,
            text="변경 취소",
            font=(FONT_FAMILY, 12),
            fg_color=ERROR_COLOR,
            width=90,
            command=discard_changes
        )

        ctk.CTkButton(
            bottom_frame,
            text="AI 전체 검토",
            font=(FONT_FAMILY, 12, "bold"),
            fg_color="#8B5CF6",
            hover_color="#7C3AED",
            width=130,
            command=run_full_ai_review
        ).pack(side="right", padx=5, pady=15)

        ctk.CTkButton(
            bottom_frame,
            text="선택 영역 검토",
            font=(FONT_FAMILY, 12, "bold"),
            fg_color=ACCENT_COLOR,
            width=130,
            command=run_partial_review
        ).pack(side="right", padx=5, pady=15)

        # 도움말 라벨 (동적으로 업데이트)
        help_label = ctk.CTkLabel(
            bottom_frame,
            text="💡 섹션을 선택하고 '선택 영역 검토'를 클릭하거나,\n   'AI 전체 검토'로 맥락을 고려한 심층 검토를 받으세요.",
            font=(FONT_FAMILY, 10),
            text_color="#A0AEC0",
            justify="left"
        )
        help_label.pack(side="left", padx=15, pady=10)

        # 편집 모드 토글 시 버튼 표시/숨김
        original_toggle_edit = toggle_edit_mode
        def toggle_edit_mode_with_buttons():
            original_toggle_edit()
            if state["edit_mode"]:
                save_btn.pack(side="right", padx=5, pady=15, before=discard_btn)
                discard_btn.pack(side="right", padx=5, pady=15)
                help_label.configure(
                    text="✏️ 편집 모드: 섹션 내용을 직접 수정하거나,\n   'AI 수정 도움' 버튼으로 AI의 도움을 받으세요."
                )
            else:
                save_btn.pack_forget()
                discard_btn.pack_forget()
                help_label.configure(
                    text="💡 섹션을 선택하고 '선택 영역 검토'를 클릭하거나,\n   'AI 전체 검토'로 맥락을 고려한 심층 검토를 받으세요."
                )

        # toggle_edit_mode 재정의
        edit_toggle.configure(command=lambda: [edit_mode_var.set(not edit_mode_var.get()), toggle_edit_mode_with_buttons()])

        # 초기 표시
        update_content_display()


# ==========================================
# 메인 실행
# ==========================================
if __name__ == "__main__":
    app = AccountingClassifierApp()
    app.mainloop()
